import json
import os
import re
import sys
import time
import logging
import io
import zipfile
import base64
import xml.etree.ElementTree as ET
from datetime import date, datetime
from decimal import Decimal

import resend
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, Response, send_file, has_request_context, stream_with_context
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from dotenv import load_dotenv

load_dotenv()


def _configure_weasyprint_dll_paths():
    """Registra rutas de Pango/GTK para WeasyPrint en Windows (antes del import)."""
    candidates = []
    env_val = str(os.environ.get('WEASYPRINT_DLL_DIRECTORIES', '') or '').strip()
    if env_val:
        candidates.extend(p.strip() for p in re.split(r'[;]', env_val) if p.strip())

    for default_path in (
        r'C:\msys64\mingw64\bin',
        r'C:\Program Files\GTK3-Runtime Win64\bin',
    ):
        if default_path not in candidates:
            candidates.append(default_path)

    valid = [p for p in candidates if os.path.isdir(p)]
    if not valid:
        return []

    os.environ['WEASYPRINT_DLL_DIRECTORIES'] = os.pathsep.join(valid)
    path_parts = [p for p in os.environ.get('PATH', '').split(os.pathsep) if p]
    for dll_dir in valid:
        if dll_dir not in path_parts:
            path_parts.insert(0, dll_dir)
        if hasattr(os, 'add_dll_directory'):
            try:
                os.add_dll_directory(dll_dir)
            except OSError:
                pass
    os.environ['PATH'] = os.pathsep.join(path_parts)
    return valid


_WEASYPRINT_DLL_DIRS = _configure_weasyprint_dll_paths()

try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except Exception as _weasy_err:
    HTML = None
    WEASYPRINT_AVAILABLE = False
    _WEASYPRINT_IMPORT_ERROR = _weasy_err

from database import User, get_datos_usuario_web, cambiar_password, get_db_connection, get_config_empresa, get_listado_generar_boletas, get_listado_certificado_quinta
from plame_sunat_parser import ARCHIVOS_SUNAT, parse_filename, parse_sunat_xml


def _env_var(*names, default=''):
    """Lee variable de entorno (Render, .env local). Quita espacios y comillas."""
    for name in names:
        raw = os.getenv(name)
        if raw is None:
            continue
        val = str(raw).strip()
        if len(val) >= 2 and val[0] == val[-1] and val[0] in ('"', "'"):
            val = val[1:-1].strip()
        if val:
            return val
    return default


def _resend_api_key():
    return _env_var('RESEND_API_KEY', 'RESEND_KEY', 'RESEND_APIKEY')


def _resend_api_key_diagnostico():
    """Ayuda segura cuando falta la API key (sin exponer secretos)."""
    candidatas = []
    for k, v in os.environ.items():
        if 'RESEND' not in k.upper():
            continue
        candidatas.append(f'{k}={"set" if str(v).strip() else "vacía"}')
    if candidatas:
        return ' Variables detectadas: ' + ', '.join(candidatas) + '.'
    return ' No hay ninguna variable de entorno cuyo nombre contenga RESEND.'


app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'dev-key-123')

logging.getLogger('werkzeug').setLevel(logging.ERROR)
sys.stdout.reconfigure(line_buffering=True)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'


def _cache_session_user(user):
    if not user:
        return
    session['_user_login'] = {
        'id': user.id,
        'username': user.username,
        'email': getattr(user, 'email', None),
        'nombre': getattr(user, 'nombre', None),
        'client_database': session.get('client_database'),
    }


def _user_from_session_cache(user_id):
    cached = session.get('_user_login') or {}
    if not cached or str(cached.get('id')) != str(user_id):
        return None
    if cached.get('client_database') and not session.get('client_database'):
        session['client_database'] = cached['client_database']
    return User(
        cached['id'],
        cached.get('username'),
        cached.get('email'),
        cached.get('nombre'),
    )


@login_manager.unauthorized_handler
def unauthorized():
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Sesión expirada o no autorizado. Vuelva a iniciar sesión.'}), 401
    return redirect(url_for('login', next=request.url))


def ensure_user_session():
    """Asegura que company y person estén en sesión."""
    if not session.get('company') or not session.get('person'):
        info = get_datos_usuario_web(current_user.id)
        if info:
            session['company'], session['person'] = info['company'], info['person']
            return info
    return {'company': session.get('company'), 'person': session.get('person')}


@app.template_filter('importe')
def format_importe(value):
    try:
        return '{:,.2f}'.format(float(value or 0))
    except Exception:
        return '0.00'


@app.template_filter('pct')
def format_pct(value):
    try:
        return '{:.2f} %'.format(float(value or 0))
    except Exception:
        return '0.00 %'


@app.template_filter('fecha')
def fecha_filter(value):
    if not value:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    s = str(value).strip()
    if not s:
        return ''
    try:
        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
            return datetime.strptime(s[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        pass
    return s


@app.template_filter('dias')
def format_dias(value):
    try:
        v = float(value or 0)
        if abs(v - round(v)) < 0.00005:
            return '{:,.0f}'.format(v)
        return '{:,.2f}'.format(v)
    except Exception:
        return '0'


@app.context_processor
def inject_now():
    from database import get_client_database_from_session
    active_db = get_client_database_from_session() or (os.getenv('SQL_DATABASE') or '').strip()
    return {
        'now': datetime.now(),
        'sql_database': active_db,
    }


def _jsonable_value(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    return value


def _report_column_name(name):
    """La primera columna del SP no tiene alias; pyodbc puede devolver '' → periodo_fmt."""
    if name is None:
        return 'periodo_fmt'
    if isinstance(name, str) and not name.strip():
        return 'periodo_fmt'
    return name


def _normalize_pr_period_vacacion(period_raw):
    """Acepta YYYYMM o YYYYMMDD; devuelve solo dígitos (6 u 8)."""
    s = str(period_raw or '').strip().replace('-', '').replace('/', '')
    if len(s) >= 8 and s[:8].isdigit():
        return s[:8]
    if len(s) >= 6 and s[:6].isdigit():
        return s[:6]
    return s


def _format_prperiod_mes(period_raw):
    """YYYYMM o YYYYMMDD → YYYY-MM para pantalla."""
    s = str(period_raw or '').strip().replace('-', '').replace('/', '')
    if len(s) >= 6 and s[:6].isdigit():
        return f'{s[:4]}-{s[4:6]}'
    return str(period_raw or '').strip()


def _xlastuser_id():
    """UserID de sesión (no nombre de usuario)."""
    return str(getattr(current_user, 'id', '') or '')[:20]


def _normalize_pr_period(period_raw):
    """
    PRPeriod en BD es yyyymmdd (8 dígitos), p. ej. 20251212.
    Acepta también '2025-12-12' o '2025/12/12' por si el valor llegó formateado.
    """
    s = str(period_raw or '').strip().replace('-', '').replace('/', '')
    if len(s) >= 8 and s[:8].isdigit():
        return s[:8]
    return str(period_raw or '').strip()


def _boleta_pdf_filename(person, period_raw):
    """Nombre estándar: boleta_{person}_{yyyymmdd}.pdf"""
    person_safe = re.sub(r'[^A-Za-z0-9_\\-]+', '_', str(person or 'preview').strip()).strip('_') or 'preview'
    period_safe = re.sub(r'[^0-9]+', '', _normalize_pr_period(period_raw)) or 'periodo'
    return f'boleta_{person_safe}_{period_safe}.pdf'


def _certificado_quinta_pdf_filename(person, anio):
    person_safe = re.sub(r'[^A-Za-z0-9_\\-]+', '_', str(person or 'preview').strip()).strip('_') or 'preview'
    anio_safe = re.sub(r'[^0-9]+', '', str(anio or '')) or 'anio'
    return f'certificado_quinta_{person_safe}_{anio_safe}.pdf'


def _certificado_trabajo_pdf_filename(person, period_raw):
    person_safe = re.sub(r'[^A-Za-z0-9_\\-]+', '_', str(person or 'preview').strip()).strip('_') or 'preview'
    period_safe = re.sub(r'[^0-9]+', '', _normalize_pr_period(period_raw)) or 'periodo'
    return f'certificado_trabajo_{person_safe}_{period_safe}.pdf'


def _tratamiento_certificado_trabajo(sex):
    """PowerBuilder dw r058: sex = '1' → el Sr., caso contrario la Srta."""
    return 'el Sr. ' if str(sex or '').strip() == '1' else 'la Srta '


def _fecha_emision_certificado_trabajo(cert):
    """Línea de fecha del certificado: Lima, día de mes del año (fecha de cese)."""
    cert = cert or {}
    try:
        dia = int(cert.get('ceasedate_day') or 0)
    except (TypeError, ValueError):
        dia = 0
    mes = str(cert.get('ceasedate_month') or '').strip() or 'Mes'
    try:
        anio = int(cert.get('ceasedate_year') or 0)
    except (TypeError, ValueError):
        anio = 0
    if dia and anio:
        return f'Lima, {dia} de {mes} del {anio}'
    return 'Lima'


def _certificado_retiro_cts_pdf_filename(person, period_raw):
    person_safe = re.sub(r'[^A-Za-z0-9_\\-]+', '_', str(person or 'preview').strip()).strip('_') or 'preview'
    period_safe = re.sub(r'[^0-9]+', '', _normalize_pr_period(period_raw)) or 'periodo'
    return f'certificado_retiro_cts_{person_safe}_{period_safe}.pdf'


def _tratamiento_retiro_cts(sex):
    """PowerBuilder dw r063: sex = '1' → al Sr., caso contrario a la Sra."""
    return 'al Sr. ' if str(sex or '').strip() == '1' else 'a la Sra '


def _tipo_doc_retiro_cts(type_pdt):
    return ' CE' if str(type_pdt or '').strip() == '04' else ' DNI'


def _texto_autorizacion_retiro_cts(cert):
    cert = cert or {}
    tratamiento = _tratamiento_retiro_cts(cert.get('sex'))
    tipo_doc = _tipo_doc_retiro_cts(cert.get('type_pdt'))
    try:
        dia = int(cert.get('fecha_cese_day') or 0)
    except (TypeError, ValueError):
        dia = 0
    mes = str(cert.get('fecha_cese_month') or '').strip()
    fecha_cese = f'desde el {dia} de {mes}' if dia and mes else ''
    return (
        f"{str(cert.get('company_name') or '').strip()} con RUC N° "
        f"{str(cert.get('company_ruc') or '').strip()}, por medio de la presente autorizamos "
        f"{tratamiento}{str(cert.get('person_name') or '').strip()}, con{tipo_doc} Nº "
        f"{str(cert.get('person_document') or '').strip()} retirar el íntegro de su depósito "
        f"de Compensación de Tiempo de Servicio (CTS), de la cuenta "
        f"{str(cert.get('cts_account') or '').strip()}, ya que ha dejado de laborar en nuestra "
        f"empresa {fecha_cese}.".replace('  ', ' ').strip()
    )


def _listar_trabajadores_liquidacion(cia, payroll_type, period, person='0', nombre=None):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'EXEC sp_pr_listadocertificadotrabajo_web @cia=?, @payrolltype=?, @period=?, @person=?, @nombre=?',
            (cia, payroll_type, period, person, nombre),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        trabajadores = []
        for r in rows:
            fi = _jsonable_value(r.get('fechaingreso'))
            fc = _jsonable_value(r.get('fechacese'))
            trabajadores.append(
                {
                    'person': str(r.get('person') or '').strip(),
                    'nombre': str(r.get('nombre') or '').strip(),
                    'email': str(r.get('email') or '').strip(),
                    'ingreso': fi if fi is not None else '',
                    'cese': fc if fc is not None else '',
                    'sex': r.get('sex', 0),
                }
            )
        return trabajadores
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


_MESES_ES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}


def _mes_nombre_es(month_num):
    try:
        return _MESES_ES.get(int(month_num), 'Mes')
    except Exception:
        return 'Mes'


def _parse_fecha_flexible(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y%m%d'):
        try:
            return datetime.strptime(s[:10] if fmt != '%Y%m%d' else s[:8], fmt).date()
        except Exception:
            continue
    return None


def _fecha_emision_certificado_quinta(fecha_emision=None):
    ref = _parse_fecha_flexible(fecha_emision) or date.today()
    return f'Lima {ref.day} de {_mes_nombre_es(ref.month)} del {ref.year}'


def _sexo_tratamiento_certificado(sexo):
    try:
        return 'Doña ' if int(sexo) == 2 else 'Don '
    except Exception:
        return 'Don '


def _calcular_impuesto_renta_quinta(c_renta_imponible, uit):
    """
    Tramos x1..x5 del certificado de quinta (PowerBuilder dw_pr_r073).
    uit = PRREP_UIT; c_renta_imponible = renta neta (total bruta - 7 UIT).
    """
    try:
        uit_val = float(uit or 0)
        c = float(c_renta_imponible or 0)
    except (TypeError, ValueError):
        return 0.0

    minuit = uit_val * 5
    mediauit = uit_val * 20
    media1uit = uit_val * 35
    maxuit = uit_val * 45

    if c <= minuit:
        x1 = c * 0.08
    else:
        x1 = minuit * 0.08

    if c <= mediauit:
        x2 = (c - minuit) * 0.14
    else:
        x2 = (mediauit - minuit) * 0.14
    if x2 < 0:
        x2 = 0.0

    if c <= media1uit:
        x3 = (c - mediauit) * 0.17
    else:
        x3 = (media1uit - mediauit) * 0.17
    if x3 < 0:
        x3 = 0.0

    if c <= maxuit:
        x4 = (c - media1uit) * 0.20
    else:
        x4 = (maxuit - media1uit) * 0.20
    if x4 < 0:
        x4 = 0.0

    if c > maxuit:
        x5 = (c - maxuit) * 0.30
    else:
        x5 = 0.0
    if x5 < 0:
        x5 = 0.0

    return x1 + x2 + x3 + x4 + x5


_MESES_5TA_TRAB = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Setiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


def _float_quinta(value, default=0.0):
    try:
        return float(value if value is not None else default)
    except (TypeError, ValueError):
        return default


def _periodo_titulo_quinta_trab(period):
    period_val = re.sub(r'[^0-9]', '', str(period or ''))
    if len(period_val) < 6:
        return str(period or '')
    try:
        mes_num = int(period_val[4:6])
        anio = period_val[:4]
        mes_nombre = _MESES_5TA_TRAB[mes_num - 1].upper() if 1 <= mes_num <= 12 else 'MES'
        return f'{mes_nombre} - {anio}'
    except Exception:
        return period_val


def _calcular_escala_seguimiento_5ta(deducible, par_uit):
    par = _float_quinta(par_uit)
    d = _float_quinta(deducible)
    tope5 = par * 5
    tope20 = par * 20
    tope35 = par * 35
    tope45 = par * 45

    imp1 = d if tope5 > d else tope5
    imp2 = (d - tope5) if tope20 > d else (tope20 - tope5)
    if imp2 < 0:
        imp2 = 0.0
    imp3 = (d - tope20) if tope35 > d else (tope35 - tope20)
    if imp3 < 0:
        imp3 = 0.0
    imp4 = (d - tope35) if tope45 > d else (tope45 - tope35)
    if imp4 < 0:
        imp4 = 0.0
    imp5 = (d - tope45) if d > tope45 else 0.0
    if imp5 < 0:
        imp5 = 0.0

    tramos = [
        {'label': 'Hasta 5 UIT', 'tope': tope5, 'base': imp1, 'pct': 8, 'retencion': imp1 * 0.08},
        {'label': 'Hasta 20 UIT', 'tope': tope20, 'base': imp2, 'pct': 14, 'retencion': imp2 * 0.14},
        {'label': 'Hasta 35 UIT', 'tope': tope35, 'base': imp3, 'pct': 17, 'retencion': imp3 * 0.17},
        {'label': 'Hasta 45 UIT', 'tope': tope45, 'base': imp4, 'pct': 20, 'retencion': imp4 * 0.20},
        {'label': 'Mas 45 UIT', 'tope': 0.0, 'base': imp5, 'pct': 30, 'retencion': imp5 * 0.30},
    ]
    total_retencion = sum(t['retencion'] for t in tramos)
    return tramos, total_retencion


def _armar_reporte_quinta_trabajador(row, period):
    row = row or {}

    meses = []
    for idx in range(1, 13):
        sfx = f'{idx:02d}'
        meses.append(
            {
                'mes': _MESES_5TA_TRAB[idx - 1],
                'ingresos': _float_quinta(row.get(f'ingreso{sfx}')),
                'liquidacion': _float_quinta(row.get(f'descuento{sfx}')),
                'utilidades': _float_quinta(row.get(f'utilidad{sfx}')),
                'retenciones': _float_quinta(row.get(f'salida{sfx}')),
            }
        )

    totales_meses = {
        'ingresos': sum(m['ingresos'] for m in meses),
        'liquidacion': sum(m['liquidacion'] for m in meses),
        'utilidades': sum(m['utilidades'] for m in meses),
        'retenciones': sum(m['retenciones'] for m in meses),
    }

    proy_ingresos = _float_quinta(row.get('proy_ingresos'))
    rem_otra_empresa = _float_quinta(row.get('rem_otra_empresa'))
    grati_julio = _float_quinta(row.get('grati_julio'))
    grati_dic = _float_quinta(row.get('grati_dic'))
    rem_acumulada = _float_quinta(row.get('rem_acumulada'))
    ingresos_5ta = _float_quinta(row.get('ingresos_5ta'))
    otros_ingresos_5ta = _float_quinta(row.get('otros_ingresos_5ta'))
    renta_neta = (
        proy_ingresos + rem_otra_empresa + grati_julio + grati_dic
        + rem_acumulada + ingresos_5ta + otros_ingresos_5ta
    )

    uit_deduccion = _float_quinta(row.get('uit'))
    par_uit = _float_quinta(row.get('par_uit'))
    deducible = renta_neta - uit_deduccion

    escala, total_renta_anual = _calcular_escala_seguimiento_5ta(deducible, par_uit)

    ret_anteriores = _float_quinta(row.get('ret_anteriores'))
    ret_otra_empresa = _float_quinta(row.get('ret_otra_empresa'))
    devolucion_quinta = _float_quinta(row.get('devolucion_quinta'))
    impuesto_anual = total_renta_anual - (ret_anteriores - devolucion_quinta) - ret_otra_empresa

    num_meses = _float_quinta(row.get('meses'))
    if num_meses <= 0:
        num_meses = 1.0
    renta_mes = impuesto_anual / num_meses

    diferenciasemana = _float_quinta(row.get('diferenciasemana'))
    numerosemana = _float_quinta(row.get('numerosemana'))
    retencion_semanal = round(diferenciasemana / numerosemana, 2) if numerosemana else 0.0

    doc_tipo = str(row.get('documenttype') or 'DNI:').strip()
    if doc_tipo and not doc_tipo.endswith(':'):
        doc_tipo = f'{doc_tipo}:'

    return {
        'nombre': str(row.get('name') or '').strip(),
        'documenttype': doc_tipo,
        'documentnumber': str(row.get('documentnumber') or row.get('docno_persona') or '').strip(),
        'cargo': str(row.get('cargo') or row.get('pr_position_cargo') or '').strip(),
        'meses_pendientes': int(_float_quinta(row.get('meses_pendientes'))),
        'meses': meses,
        'totales_meses': totales_meses,
        'proyeccion': {
            'meses_pendientes': int(_float_quinta(row.get('meses_pendientes'))),
            'proy_ingresos': proy_ingresos,
            'rem_otra_empresa': rem_otra_empresa,
            'grati_julio': grati_julio,
            'grati_dic': grati_dic,
            'rem_acumulada': rem_acumulada,
            'ingresos_5ta': ingresos_5ta,
            'otros_ingresos_5ta': otros_ingresos_5ta,
            'renta_neta': renta_neta,
            'uit_valor': par_uit,
            'uit_deduccion': uit_deduccion,
            'deducible': deducible,
        },
        'escala': escala,
        'par_uit': par_uit,
        'resumen': {
            'total_renta_anual': total_renta_anual,
            'devolucion_quinta': devolucion_quinta,
            'ret_anteriores': ret_anteriores,
            'ret_otra_empresa': ret_otra_empresa,
            'impuesto_anual': impuesto_anual,
            'num_meses': num_meses,
            'renta_mes': renta_mes,
            'ret_renta_acum': _float_quinta(row.get('ret_renta_acum')),
            'diferenciasemana': diferenciasemana,
            'numerosemana': numerosemana,
            'retencion_semanal': retencion_semanal,
        },
        'periodo_titulo': _periodo_titulo_quinta_trab(period),
    }


def _get_company_header_quinta(cia):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT
                LTRIM(RTRIM(ISNULL(Description, ''))) AS descripcion,
                LTRIM(RTRIM(ISNULL(TaxID, ''))) AS ruc
            FROM SY_Company (NOLOCK)
            WHERE Company = ?
            """,
            (cia,),
        )
        row = cursor.fetchone()
        if not row:
            return {'nombre': cia, 'ruc': ''}
        cols = [c[0] for c in cursor.description]
        data = dict(zip(cols, row))
        return {
            'nombre': str(data.get('descripcion') or cia).strip(),
            'ruc': str(data.get('ruc') or '').strip(),
        }
    except Exception:
        logging.exception('_get_company_header_quinta')
        return {'nombre': cia, 'ruc': ''}
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


def _asignacion_concepto_pk_from_json(body):
    """Clave PR_EmployeeConcept desde JSON del cliente."""
    body = body or {}
    return {
        'company': str(body.get('company') or body.get('cia') or '').strip(),
        'person': str(body.get('person') or '').strip(),
        'concept': str(body.get('concept') or '').strip(),
        'payrolltype': str(body.get('payrolltype') or body.get('payroll_type') or '').strip(),
        'prperiodstart': _normalize_pr_period(body.get('prperiodstart') or body.get('period_start')),
        'costcenter': str(body.get('costcenter') if body.get('costcenter') is not None else '').strip(),
    }


def _normalize_flagapplyformula_asig(raw):
    if raw is True or str(raw).strip().upper() in ('Y', '1', 'TRUE', 'SI', 'SÍ'):
        return 'Y'
    return 'N'


def _normalize_flagfrecuencytype_asig(raw):
    v = str(raw or 'P').strip().upper()
    return 'T' if v == 'T' else 'P'


def _normalize_conceptcurrency_asig(raw):
    v = str(raw or 'LO').strip().upper()
    return 'EX' if v == 'EX' else 'LO'


def _normalize_tipo_concepto_asig(raw):
    """Filtro listado: 0 = todos, P = permanente, T = temporal."""
    v = str(raw or '0').strip().upper()
    if v == 'P':
        return 'P'
    if v == 'T':
        return 'T'
    return '0'


def _normalize_replicationunit_asig(raw):
    """Filtro listado: 0 = todas las unidades (SY_Person.ReplicationUnit)."""
    v = str(raw or '0').strip()
    return '0' if v in ('', '0') else v


PLAME_ARCHIVO_EXTENSION = {
    '14': 'jor',
    '15': 'snl',
    '18': 'rem',
    '26': 'toc',
}


def _plame_period_yyyymm(period_raw):
    """Periodo tributario PLAME: YYYYMM (6 dígitos)."""
    s = str(period_raw or '').strip().replace('-', '').replace('/', '')
    if len(s) >= 6 and s[:6].isdigit():
        return s[:6]
    return ''


def _plame_params_from_json(body):
    body = body or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    period = _plame_period_yyyymm(body.get('period') or body.get('periodo'))
    return {'cia': cia, 'period': period}


def _plame_validar_params(p):
    if not p.get('cia'):
        return 'Seleccione compañía.'
    if not p.get('period') or len(p['period']) != 6:
        return 'Indique un periodo tributario válido (YYYYMM).'
    return None


def _plame_split_horas_minutos(valor):
    try:
        v = float(valor or 0)
    except (TypeError, ValueError):
        v = 0.0
    if v < 0:
        v = 0.0
    horas = int(v)
    minutos = int(round((v - horas) * 60))
    if minutos >= 60:
        horas += minutos // 60
        minutos = minutos % 60
    return horas, minutos


def _plame_linea_archivo14(row):
    """Genera línea PLAME Archivo 14: TT|DOC|H|m|H|m|"""
    doc_type = str(row.get('documenttype') or '').strip()
    if doc_type.isdigit():
        doc_type = doc_type.zfill(2)
    doc_num = str(row.get('documentnumber') or '').strip()
    wh, wm = _plame_split_horas_minutos(row.get('workinghours'))
    eh, em = _plame_split_horas_minutos(row.get('extrahours'))
    wmi = row.get('workingminutes')
    emi = row.get('extraminutes')
    if wmi is not None and str(wmi).strip() not in ('', '0'):
        try:
            wm = max(0, int(float(wmi)))
        except (TypeError, ValueError):
            pass
    if emi is not None and str(emi).strip() not in ('', '0'):
        try:
            em = max(0, int(float(emi)))
        except (TypeError, ValueError):
            pass
    return '|'.join([
        doc_type,
        doc_num,
        str(wh),
        str(wm),
        str(eh),
        str(em),
    ]) + '|'


def _plame_linea_archivo26(row):
    """Genera línea PLAME Archivo 26 (.toc): TT|DOC|INDICADOR|"""
    doc_type = str(row.get('documenttype') or '').strip()
    if doc_type.isdigit():
        doc_type = doc_type.zfill(2)
    doc_num = str(row.get('documentnumber') or '').strip()
    indicator = str(row.get('pensionmembership') or '0').strip()
    if indicator not in ('0', '1'):
        indicator = '0'
    return '|'.join([doc_type, doc_num, indicator]) + '|'


def _plame_rows_archivo26_from_json(body):
    rows = body.get('rows')
    if not isinstance(rows, list):
        return []
    allowed_docs = {'01', '04', '07', '09'}
    resultado = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        doc_num = str(r.get('documentnumber') or '').strip()
        if not doc_num:
            continue
        doc_type = str(r.get('documenttype') or '').strip()
        if doc_type.isdigit():
            doc_type = doc_type.zfill(2)
        if doc_type not in allowed_docs:
            continue
        indicator = str(r.get('pensionmembership') or '0').strip()
        if indicator not in ('0', '1'):
            indicator = '0'
        resultado.append({
            'documenttype': doc_type,
            'documentnumber': doc_num,
            'pensionmembership': indicator,
        })
    return resultado


def _plame_linea_archivo15(row):
    """Genera línea PLAME Archivo 15 (.snl): TT|DOC|MOTIVO|DIAS|"""
    doc_type = str(row.get('documenttype') or '').strip()
    if doc_type.isdigit():
        doc_type = doc_type.zfill(2)
    doc_num = str(row.get('documentnumber') or '').strip()
    susp = str(row.get('suspensiontype') or row.get('pdt') or '').strip()
    if susp.isdigit():
        susp = susp.zfill(2)
    try:
        days = int(float(row.get('days') or 0))
    except (TypeError, ValueError):
        days = 0
    if days < 0:
        days = 0
    return '|'.join([
        doc_type,
        doc_num,
        susp,
        str(days).zfill(2),
    ]) + '|'


def _plame_validar_archivo14_incidencias(cursor, p):
    """sp_pr_plame_validar_archivo14_web → mensajes y personas con incidencia por fila."""
    cursor.execute(
        'EXEC sp_pr_plame_validar_archivo14_web @cia=?, @period=?',
        (p['cia'], p['period']),
    )
    rows = _dicts_first_nonempty_resultset(cursor)
    mensajes = []
    personas = set()
    for r in rows or []:
        person = str(r.get('person') or '').strip()
        msg = str(r.get('mensaje') or r.get('observacion') or '').strip()
        if not msg:
            continue
        mensajes.append(msg)
        if person:
            personas.add(person)
    return mensajes, personas


def _plame_validar_archivo18_incidencias(cursor, p):
    """sp_pr_plame_validar_archivo18_web → mensajes y personas con incidencia por fila."""
    cursor.execute(
        'EXEC sp_pr_plame_validar_archivo18_web '
        '@cia=?, @period=?, @payroll_all=?, @payroll=?, @cesados=?',
        (p['cia'], p['period'], p['payroll_all'], p['payroll'] or None, p['cesados']),
    )
    rows = _dicts_first_nonempty_resultset(cursor)
    mensajes = []
    personas = set()
    for r in rows or []:
        person = str(r.get('person') or '').strip()
        msg = str(r.get('mensaje') or r.get('observacion') or '').strip()
        if not msg:
            continue
        mensajes.append(msg)
        if person:
            personas.add(person)
    return mensajes, personas


def _plame_params_archivo18_from_json(body):
    base = _plame_params_from_json(body)
    body = body or {}
    payroll = str(body.get('payroll') or body.get('payrolltype') or body.get('payroll_type') or '').strip()
    payroll_all = str(body.get('payroll_all') or '').strip().upper()
    if payroll_all not in ('Y', 'N'):
        payroll_all = 'Y' if not payroll or payroll in ('0', 'T', 'TODOS', 'TODAS') else 'N'
    if payroll_all == 'Y':
        payroll = ''
    cesados = _normalize_cesados_telecredito(body.get('cesados'))
    base['payroll_all'] = payroll_all
    base['payroll'] = payroll
    base['cesados'] = cesados
    return base


def _plame_pdt_monto_cero_obligatorio(pdt):
    """PDT que deben exportarse con 0|0 aunque no tengan movimiento (comisión, ret. 5ta)."""
    s = str(pdt or '').strip()
    if s.isdigit():
        s = s.zfill(4)
    return s in ('0601', '0605')


def _plame_format_monto_rem(valor, pdt=None):
    """Monto PLAME .rem: entero sin decimales o hasta 2 decimales con punto."""
    pdt_norm = str(pdt or '').strip()
    if pdt_norm.isdigit():
        pdt_norm = pdt_norm.zfill(4)
    if valor is None or valor == '':
        if _plame_pdt_monto_cero_obligatorio(pdt_norm):
            return '0'
        return ''
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return ''
    if abs(v) < 0.00005:
        if _plame_pdt_monto_cero_obligatorio(pdt_norm):
            return '0'
        return ''
    redondeado = round(v, 2)
    if abs(redondeado - round(redondeado)) < 0.00005:
        return str(int(round(redondeado)))
    return f'{redondeado:.2f}'


def _plame_es_descuento_tabla22(pdt):
    """Códigos 07xx (descuentos): solo monto pagado si no hay devengado en el .rem."""
    s = str(pdt or '').strip()
    if s.isdigit():
        s = s.zfill(4)
    return len(s) >= 2 and s[:2] == '07' and s != '0700'


def _plame_linea_archivo18(row):
    """Genera línea PLAME Archivo 18 (.rem): TT|DOC|PDT|DEV|PAG|"""
    doc_type = str(row.get('documenttype') or '').strip()
    if doc_type.isdigit():
        doc_type = doc_type.zfill(2)
    doc_num = str(row.get('documentnumber') or '').strip()
    pdt = str(row.get('pdt') or '').strip()
    if pdt.isdigit():
        pdt = pdt.zfill(4)
    devengado_val = row.get('conceptvalue', row.get('devengado'))
    pagado_val = row.get('conceptvaluelo', row.get('pagado'))
    devengado = _plame_format_monto_rem(devengado_val, pdt)
    pagado = _plame_format_monto_rem(pagado_val, pdt)
    if not (_plame_es_descuento_tabla22(pdt) and not devengado):
        if devengado and not pagado:
            pagado = devengado
        elif pagado and not devengado:
            devengado = pagado
    return '|'.join([doc_type, doc_num, pdt, devengado, pagado]) + '|'


def _plame_rows_archivo18_from_json(body):
    rows = body.get('rows')
    if not isinstance(rows, list):
        return []
    resultado = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        doc_num = str(r.get('documentnumber') or '').strip()
        pdt = str(r.get('pdt') or '').strip()
        if not doc_num or not pdt:
            continue
        try:
            cv = float(r.get('conceptvalue') or r.get('devengado') or 0)
            cl = float(r.get('conceptvaluelo') or r.get('pagado') or 0)
        except (TypeError, ValueError):
            cv, cl = 0.0, 0.0
        pdt_norm = pdt.zfill(4) if pdt.isdigit() else pdt
        if abs(cv) < 0.00005 and abs(cl) < 0.00005 and not _plame_pdt_monto_cero_obligatorio(pdt_norm):
            continue
        resultado.append({
            'documenttype': str(r.get('documenttype') or '').strip(),
            'documentnumber': doc_num,
            'pdt': pdt,
            'conceptvalue': cv,
            'conceptvaluelo': cl,
        })
    return resultado


def _declaracion_afp_params_from_json(body):
    body = body or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    period = str(body.get('period') or '').strip()
    if len(period) >= 6:
        period = period[:6]

    payroll = str(body.get('payroll') or body.get('payrolltype') or body.get('payroll_type') or '').strip()
    payroll_all = str(body.get('payroll_all') or '').strip().upper()
    if payroll_all not in ('Y', 'N'):
        payroll_all = 'Y' if not payroll or payroll in ('0', 'T', 'TODOS', 'TODAS') else 'N'
    if payroll_all == 'Y':
        payroll = ''

    afp = str(body.get('afp') or '').strip()
    afp_all = str(body.get('afp_all') or '').strip().upper()
    if afp_all not in ('Y', 'N'):
        afp_all = 'Y' if not afp or afp in ('0', 'T', 'TODOS', 'TODAS') else 'N'
    if afp_all == 'Y':
        afp = ''

    return {
        'cia': cia,
        'period': period,
        'payroll_all': payroll_all,
        'payroll': payroll,
        'afp_all': afp_all,
        'afp': afp,
    }


def _declaracion_afp_validar_params(p):
    if not p.get('cia'):
        return 'Seleccione la compañía.'
    period = str(p.get('period') or '').strip()
    if not re.fullmatch(r'\d{6}', period):
        return 'Seleccione un periodo válido (YYYYMM).'
    return None


def _declaracion_afp_row_dict(r):
    aporte_riesgo = _jsonable_value(r.get('aporte_riesgo_trab'))
    return {
        'person': _jsonable_value(r.get('person')),
        'afp_description': _jsonable_value(r.get('afp_description')),
        'cuspp': _jsonable_value(r.get('cuspp')),
        'documentnumber': _jsonable_value(r.get('documentnumber')),
        'lastname1': _jsonable_value(r.get('lastname1')),
        'lastname2': _jsonable_value(r.get('lastname2')),
        'names': _jsonable_value(r.get('names')),
        'fecha_cese': _jsonable_value(r.get('fecha_cese')),
        'entrydate': _jsonable_value(r.get('entrydate')),
        'ceasedate': _jsonable_value(r.get('ceasedate')),
        'inicio_relacion': _jsonable_value(r.get('inicio_relacion')),
        'cese_relacion': _jsonable_value(r.get('cese_relacion')),
        'relacion_laboral': _jsonable_value(r.get('relacion_laboral')),
        'excepcion_aportar': _jsonable_value(r.get('excepcion_aportar')),
        'tipo_trabajo': _jsonable_value(r.get('tipo_trabajo')),
        'remuneracion': _jsonable_value(r.get('remuneracion')),
        'topafp': _jsonable_value(r.get('topafp')),
        'insuredpercentage': _jsonable_value(r.get('insuredpercentage')),
        'aporte_obligatorio': _jsonable_value(r.get('aporte_obligatorio')),
        'aporte_empleador': _jsonable_value(r.get('aporte_empleador')),
        'total_fondo_pensiones': _jsonable_value(r.get('total_fondo_pensiones')),
        'seguro': _jsonable_value(r.get('seguro')),
        'seguro_esperado': _jsonable_value(r.get('seguro_esperado')),
        'comision': _jsonable_value(r.get('comision')),
        'total_retenciones': _jsonable_value(r.get('total_retenciones')),
        'aporte_riesgo_trab': aporte_riesgo,
        'aporte_riesgo_emp': aporte_riesgo,
        'tipodoc': _jsonable_value(r.get('tipodoc')),
    }


def _declaracion_afp_rows_from_json(body):
    rows = body.get('rows')
    if not isinstance(rows, list):
        return []
    return [r for r in rows if isinstance(r, dict)]


def _afpnet_limpia_texto(valor):
    return str(valor or '').strip()


def _cuspp_afpnet_es_valido(cuspp):
    return bool(re.fullmatch(r'[A-Za-z0-9Ññ]{12}', str(cuspp or '').strip()))


def _fecha_ddmmyyyy_en_periodo(fecha_txt, period_yyyymm):
    """True si la fecha dd/mm/yyyy cae en el periodo YYYYMM."""
    period = str(period_yyyymm or '').strip()[:6]
    if not re.fullmatch(r'\d{6}', period):
        return False
    s = str(fecha_txt or '').strip()
    m = re.fullmatch(r'(\d{2})/(\d{2})/(\d{4})', s)
    if not m:
        return False
    _d, mes, anio = m.group(1), m.group(2), m.group(3)
    return f'{anio}{mes}' == period


def _declaracion_afp_nombre_completo_trabajador(row):
    partes = [
        str(row.get('lastname1') or '').strip(),
        str(row.get('lastname2') or '').strip(),
        str(row.get('names') or '').strip(),
    ]
    nombre = ' '.join(p for p in partes if p)
    if nombre:
        return nombre
    return str(row.get('nombre') or '').strip()


def _declaracion_afp_etiqueta_trabajador(row):
    dni = str(row.get('documentnumber') or '').strip()
    nombre = _declaracion_afp_nombre_completo_trabajador(row)
    if dni and nombre:
        return f'DNI {dni} — {nombre}'
    return dni or nombre or 'Trabajador'


def _declaracion_afp_etiqueta_fila(row, fila=None):
    etiqueta = _declaracion_afp_etiqueta_trabajador(row)
    if fila is not None:
        return f'Fila {fila} ({etiqueta})'
    return etiqueta


def _declaracion_afp_validar_fila_afpnet(row, period, fila=None):
    """Validaciones críticas AFPnet antes de generar el XLS."""
    errores = []
    etiqueta = _declaracion_afp_etiqueta_fila(row, fila)
    period = str(period or '').strip()[:6]

    excepcion = str(row.get('excepcion_aportar') or '').strip().upper()
    cuspp = str(row.get('cuspp') or '').strip()
    if excepcion not in ('J', 'I', 'O', 'L', 'U', 'P') and not _cuspp_afpnet_es_valido(cuspp):
        errores.append(
            f'{etiqueta}: CUSPP inválido o vacío («{cuspp or "vacío"}»). '
            'Debe tener exactamente 12 caracteres alfanuméricos (incluye Ñ).'
        )
    elif cuspp and not _cuspp_afpnet_es_valido(cuspp):
        errores.append(
            f'{etiqueta}: CUSPP inválido («{cuspp}»). '
            'Debe tener exactamente 12 caracteres alfanuméricos (incluye Ñ).'
        )

    inicio = str(row.get('inicio_relacion') or '').strip().upper()
    entrydate = str(row.get('entrydate') or '').strip()
    if inicio == 'S':
        if not entrydate:
            errores.append(
                f'{etiqueta}: Inicio de relación laboral marcado «S» sin fecha de ingreso.'
            )
        elif not _fecha_ddmmyyyy_en_periodo(entrydate, period):
            errores.append(
                f'{etiqueta}: Inicio de relación laboral «S» solo aplica si el ingreso '
                f'({entrydate}) cae en el periodo {period[:4]}-{period[4:6]}.'
            )

    cese = str(row.get('cese_relacion') or '').strip().upper()
    ceasedate = str(row.get('ceasedate') or '').strip()
    if cese == 'S':
        if not ceasedate:
            errores.append(
                f'{etiqueta}: Cese de relación laboral marcado «S» sin fecha de cese.'
            )
        elif not _fecha_ddmmyyyy_en_periodo(ceasedate, period):
            errores.append(
                f'{etiqueta}: Cese de relación laboral «S» solo aplica si el cese '
                f'({ceasedate}) cae en el periodo {period[:4]}-{period[4:6]}.'
            )

    return errores


def _declaracion_afp_validar_afpnet(filas, period):
    mensajes = []
    for idx, row in enumerate(filas or [], start=1):
        mensajes.extend(_declaracion_afp_validar_fila_afpnet(row, period, fila=idx))
    return mensajes


def _declaracion_afp_aplicar_validaciones_filas(filas, period):
    resultado = []
    todas = []
    for idx, row in enumerate(filas or [], start=1):
        errores = _declaracion_afp_validar_fila_afpnet(row, period, fila=idx)
        item = dict(row)
        item['validacion_ok'] = len(errores) == 0
        item['validacion_errores'] = errores
        resultado.append(item)
        todas.extend(errores)
    return resultado, todas


def _declaracion_afp_validar_regimen_pension_planilla(cursor, p):
    """Advertencias: trabajadores en planilla del periodo sin régimen ONP/AFP."""
    cursor.execute(
        'EXEC sp_pr_trabajadores_sin_regimen_pension_afp_web '
        '@cia=?, @period=?, @payroll_all=?, @payroll=?',
        (p['cia'], p['period'], p['payroll_all'], p['payroll'] or None),
    )
    rows = _dicts_first_nonempty_resultset(cursor)
    mensajes = []
    for r in rows or []:
        etiqueta = _declaracion_afp_etiqueta_trabajador(r)
        mensajes.append(
            f'{etiqueta}: sin régimen de pensión en la planilla (debe tener ONP o AFP).'
        )
    return mensajes


def _declaracion_afp_es_jubilado_afpnet(row):
    return str(row.get('afp_description') or '').strip() == '(Jubilado)'


def _declaracion_afp_validar_jubilados_filas(filas):
    """Advertencias: jubilados incluidos en el archivo AFPnet (remuneración 0)."""
    jubilados = [r for r in (filas or []) if _declaracion_afp_es_jubilado_afpnet(r)]
    if not jubilados:
        return []
    mensajes = [
        'Jubilados en AFPnet ({}): se incluyen con remuneración 0 y excepción de aportar (J o L).'.format(
            len(jubilados)
        )
    ]
    for r in jubilados:
        etiqueta = _declaracion_afp_etiqueta_trabajador(r)
        exc = str(r.get('excepcion_aportar') or 'J').strip().upper() or 'J'
        mensajes.append(
            f'{etiqueta}: jubilado AFPnet (excepción {exc}, remuneración 0).'
        )
    return mensajes


def _declaracion_afp_validaciones_completas(cursor, filas, p):
    filas_val, validaciones = _declaracion_afp_aplicar_validaciones_filas(filas, p['period'])
    validaciones.extend(_declaracion_afp_validar_regimen_pension_planilla(cursor, p))
    validaciones.extend(_declaracion_afp_validar_jubilados_filas(filas_val))
    return filas_val, validaciones


def _afpnet_sn(valor, default='N'):
    """AFPnet col. H/I/J: solo S o N."""
    v = str(valor or '').strip().upper()
    return 'S' if v == 'S' else default


def _afpnet_relacion_laboral(row):
    return _afpnet_sn(row.get('relacion_laboral'))


def _afpnet_inicio_relacion(row):
    return _afpnet_sn(row.get('inicio_relacion'))


def _afpnet_cese_relacion(row):
    return _afpnet_sn(row.get('cese_relacion'))


def _afpnet_excepcion_aportar(row):
    """AFPnet col. K: vacío o L/U/J/I/P/O."""
    v = str(row.get('excepcion_aportar') or '').strip().upper()
    return v if v in ('L', 'U', 'J', 'I', 'P', 'O') else ''


def _afpnet_remuneracion(row):
    try:
        return round(float(row.get('remuneracion') or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _afpnet_tipo_trabajo(row):
    v = str(row.get('tipo_trabajo') or '').strip().upper()
    return v if v in ('N', 'C', 'M', 'P') else 'N'


def _afpnet_documento(valor):
    """Col. D: texto para conservar ceros a la izquierda."""
    return _afpnet_limpia_texto(valor)


def _afpnet_fila_excel(secuencia, row):
    return [
        secuencia,
        _afpnet_limpia_texto(row.get('cuspp')),
        _afpnet_limpia_texto(row.get('tipodoc')),
        _afpnet_documento(row.get('documentnumber')),
        _afpnet_limpia_texto(row.get('lastname1')),
        _afpnet_limpia_texto(row.get('lastname2')),
        _afpnet_limpia_texto(row.get('names')),
        _afpnet_relacion_laboral(row),
        _afpnet_inicio_relacion(row),
        _afpnet_cese_relacion(row),
        _afpnet_excepcion_aportar(row),
        _afpnet_remuneracion(row),
        0.0,
        0.0,
        0.0,
        _afpnet_tipo_trabajo(row),
    ]


def _declaracion_afp_actualizar_datos(cursor, conn, p, xlastuser='WEB'):
    """sp_pr_actualizar_datos_afp_web: control de datos AFP antes del Excel AFPnet."""
    cursor.execute(
        "EXEC sp_pr_actualizar_datos_afp_web "
        "@cia=?, @period=?, @payroll_all=?, @payroll=?, @xlastuser=?",
        (p['cia'], p['period'], p['payroll_all'], p['payroll'] or None, xlastuser),
    )
    rows = _dicts_first_nonempty_resultset(cursor)
    conn.commit()
    if rows:
        return rows[0]
    return {'actualizado': 0, 'mensaje': 'Sin respuesta del control de datos AFP.'}


def _declaracion_afp_xlastuser():
    return str(session.get('person') or 'WEB').strip() or 'WEB'


def _declaracion_afp_sincronizar_planilla(cursor, conn, p):
    """Alimenta PR_EmployeeAFP / PR_EmployeeAFPHeader desde la planilla del periodo."""
    sync = _declaracion_afp_actualizar_datos(cursor, conn, p, _declaracion_afp_xlastuser())
    if sync.get('actualizado') == 0:
        msg = str(sync.get('mensaje') or '').strip()
        if msg:
            logging.info('AFPnet sync periodo %s cia %s: %s', p.get('period'), p.get('cia'), msg)
    return sync


def _declaracion_afp_validaciones_sync_afp(validaciones, sync):
    """Incluye advertencia si no hubo planilla/proceso con TOTAL_REM_AFP."""
    if sync.get('actualizado') != 0:
        return validaciones
    msg = str(sync.get('mensaje') or '').strip()
    if not msg:
        msg = 'No hay combinaciones planilla/proceso con concepto TOTAL_REM_AFP para el periodo.'
    if msg not in (validaciones or []):
        return [msg] + list(validaciones or [])
    return validaciones


def _declaracion_afp_ejecutar_listado(cursor, p):
    cursor.execute(
        "EXEC sp_pr_listado_declaracion_afp_web "
        "@cia=?, @period=?, @payroll_all=?, @payroll=?, "
        "@afp_all=?, @afp=?, @repunit_all=?, @repunit=?, "
        "@flagcostcenter=?, @costcenter=?, @employee_all=?, @employee=?",
        (
            p['cia'], p['period'], p['payroll_all'], p['payroll'] or None,
            p['afp_all'], p['afp'] or None, 'Y', None, 'Y', None, 'Y', None,
        ),
    )
    rows = _dicts_first_nonempty_resultset(cursor)
    return [_declaracion_afp_row_dict(r) for r in rows]


def _declaracion_afp_generar_xlsx_bytes(filas):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'AFPnet'
    for idx, row in enumerate(filas, start=1):
        valores = _afpnet_fila_excel(idx, row)
        ws.append(valores)
        fila_excel = ws.max_row
        for col in (3, 4):
            celda = ws.cell(row=fila_excel, column=col)
            celda.number_format = '@'
            celda.value = str(valores[col - 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _declaracion_afp_etiqueta_proceso(shortname):
    labels = {
        'FIN_DE_MES': 'FIN DE MES',
        'LIQUIDACION': 'LIQUIDACION',
        'SEMANAL': 'SEMANAL',
    }
    key = str(shortname or '').strip().upper()
    return labels.get(key, key)


def _declaracion_afp_inicio_relacion_mes(row):
    return str(row.get('inicio_relacion') or '').strip().upper() == 'S'


def _declaracion_afp_cese_relacion_mes(row):
    return str(row.get('cese_relacion') or '').strip().upper() == 'S'


def _declaracion_afp_ingreso_cese_mismo_mes(row):
    return (
        _declaracion_afp_inicio_relacion_mes(row)
        and _declaracion_afp_cese_relacion_mes(row)
    )


def _declaracion_afp_diferencia_activa(diferencia, clave):
    return _declaracion_afp_resumen_int((diferencia or {}).get(clave)) != 0


def _declaracion_afp_flag_si(valor):
    if valor in (1, '1', True):
        return True
    return str(valor or '').strip().upper() in ('S', 'Y', '1')


def _declaracion_afp_etiqueta_trabajador_datos(documentnumber, nombre):
    dni = str(documentnumber or '').strip()
    nombre = str(nombre or '').strip()
    if dni and nombre:
        return f'DNI {dni} — {nombre}'
    return dni or nombre or 'Trabajador'


def _declaracion_afp_mensaje_diferencia_trabajador(documentnumber, nombre, motivo):
    etiqueta = _declaracion_afp_etiqueta_trabajador_datos(documentnumber, nombre)
    return f'Trabajador {etiqueta}: {motivo}'


def _declaracion_afp_nombre_completo_desde_partes(row):
    partes = [
        str(row.get('lastname1') or '').strip(),
        str(row.get('lastname2') or '').strip(),
        str(row.get('names') or '').strip(),
    ]
    nombre = ' '.join(p for p in partes if p)
    return nombre or _declaracion_afp_etiqueta_trabajador(row)


def _declaracion_afp_flags_afpnet(row):
    return (
        _declaracion_afp_inicio_relacion_mes(row),
        _declaracion_afp_cese_relacion_mes(row),
    )


def _declaracion_afp_flags_planilla(reg):
    inicio = _declaracion_afp_flag_si(reg.get('inicio_en_periodo'))
    cese = _declaracion_afp_flag_si(reg.get('cese_en_periodo'))
    return inicio, cese


def _declaracion_afp_describe_flags(inicio, cese):
    partes = []
    if inicio:
        partes.append('nuevo')
    if cese:
        partes.append('cesado')
    if not partes:
        partes.append('antiguo')
    return ' y '.join(partes)


def _declaracion_afp_gap_total_vs_planilla(conteo):
    """Diferencia interna: Total (suma categorías) menos Total Planilla (personas únicas)."""
    c = conteo or {}
    return (
        _declaracion_afp_resumen_int(c.get('total'))
        - _declaracion_afp_resumen_int(c.get('total_planilla'))
    )


def _declaracion_afp_diferencia_con_gap_interno(diferencia, planilla, afpnet):
    """Si planilla y AFPnet coinciden pero Total supera a Total Planilla, exponer -1 en total_planilla."""
    dif = dict(diferencia or {})
    claves = ('nuevos', 'cesados', 'antiguos', 'total', 'total_planilla')
    if any(_declaracion_afp_diferencia_activa(dif, k) for k in claves):
        return dif
    gap = _declaracion_afp_gap_total_vs_planilla(planilla)
    if gap == 0:
        gap = _declaracion_afp_gap_total_vs_planilla(afpnet)
    if gap != 0:
        dif['total_planilla'] = -gap
    return dif


def _declaracion_afp_detalle_diferencias_trabajadores(
    filas, diferencia, planilla_trabajadores=None, planilla=None, afpnet=None,
):
    """Identifica trabajadores que explican diferencias planilla vs AFPnet o Total vs Total Planilla."""
    dif = diferencia or {}
    claves = ('nuevos', 'cesados', 'antiguos', 'total', 'total_planilla')
    gap_planilla = _declaracion_afp_gap_total_vs_planilla(planilla)
    gap_afpnet = _declaracion_afp_gap_total_vs_planilla(afpnet)
    hay_gap_interno = gap_planilla != 0 or gap_afpnet != 0
    if (
        not any(_declaracion_afp_diferencia_activa(dif, k) for k in claves)
        and not hay_gap_interno
    ):
        return []

    detalles = []
    vistos = set()

    def agregar(person, documentnumber, nombre, motivo):
        person_key = str(person or '').strip()
        clave = (person_key, motivo)
        if clave in vistos:
            return
        vistos.add(clave)
        detalles.append({
            'person': _jsonable_value(person),
            'documentnumber': _jsonable_value(documentnumber),
            'nombre': nombre,
            'mensaje': _declaracion_afp_mensaje_diferencia_trabajador(documentnumber, nombre, motivo),
        })

    filas = filas or []
    filas_por_persona = {}
    for row in filas:
        person = str(row.get('person') or '').strip()
        if person:
            filas_por_persona[person] = row

    planilla_por_persona = {}
    for reg in planilla_trabajadores or []:
        person = str(reg.get('person') or '').strip()
        if person:
            planilla_por_persona[person] = reg

    personas_planilla = set(planilla_por_persona)
    personas_afpnet = set(filas_por_persona)
    solo_planilla = personas_planilla - personas_afpnet
    solo_afpnet = personas_afpnet - personas_planilla
    overlap_personas = set()
    for row in filas or []:
        if _declaracion_afp_ingreso_cese_mismo_mes(row):
            person = str(row.get('person') or '').strip()
            if person:
                overlap_personas.add(person)
    for reg in planilla_trabajadores or []:
        inicio, cese = _declaracion_afp_flags_planilla(reg)
        if inicio and cese:
            person = str(reg.get('person') or '').strip()
            if person:
                overlap_personas.add(person)

    def aplica_diferencia_categoria(inicio, cese):
        if _declaracion_afp_diferencia_activa(dif, 'total'):
            return True
        if _declaracion_afp_diferencia_activa(dif, 'total_planilla'):
            return True
        if inicio and _declaracion_afp_diferencia_activa(dif, 'nuevos'):
            return True
        if cese and _declaracion_afp_diferencia_activa(dif, 'cesados'):
            return True
        if not inicio and not cese and _declaracion_afp_diferencia_activa(dif, 'antiguos'):
            return True
        return False

    for person in sorted(solo_planilla):
        reg = planilla_por_persona[person]
        inicio, cese = _declaracion_afp_flags_planilla(reg)
        if not aplica_diferencia_categoria(inicio, cese):
            continue
        agregar(
            person,
            reg.get('documentnumber'),
            _declaracion_afp_nombre_completo_desde_partes(reg),
            'está en la planilla del periodo pero no aparece en el listado AFPnet '
            f'(se contabiliza como {_declaracion_afp_describe_flags(inicio, cese)} en planilla).',
        )

    for person in sorted(solo_afpnet):
        row = filas_por_persona[person]
        inicio, cese = _declaracion_afp_flags_afpnet(row)
        if not aplica_diferencia_categoria(inicio, cese):
            continue
        agregar(
            person,
            row.get('documentnumber'),
            _declaracion_afp_nombre_completo_trabajador(row),
            'aparece en el listado AFPnet pero no en la población del resumen de planilla '
            f'(se contabiliza como {_declaracion_afp_describe_flags(inicio, cese)} en AFPnet).',
        )

    if not solo_planilla and not solo_afpnet and overlap_personas and hay_gap_interno:
        objetivo = abs(gap_planilla) or abs(gap_afpnet)
        if not objetivo or len(overlap_personas) == objetivo:
            motivo_ingreso_cese_mismo_mes = (
                'ingresó y cesó el mismo mes (se contabiliza en nuevos y en cesados; '
                'por eso Total supera a Total Planilla en 1).'
            )
            for person in sorted(overlap_personas):
                row = filas_por_persona.get(person)
                reg = planilla_por_persona.get(person)
                agregar(
                    person,
                    (row or reg or {}).get('documentnumber'),
                    _declaracion_afp_nombre_completo_trabajador(row or reg or {}),
                    motivo_ingreso_cese_mismo_mes,
                )

    for person in sorted(personas_planilla & personas_afpnet):
        reg = planilla_por_persona[person]
        row = filas_por_persona[person]
        flags_planilla = _declaracion_afp_flags_planilla(reg)
        flags_afpnet = _declaracion_afp_flags_afpnet(row)
        if flags_planilla == flags_afpnet:
            continue
        objetivo = 0
        if flags_planilla[0] != flags_afpnet[0] and _declaracion_afp_diferencia_activa(dif, 'nuevos'):
            objetivo = abs(_declaracion_afp_resumen_int(dif.get('nuevos')))
        elif flags_planilla[1] != flags_afpnet[1] and _declaracion_afp_diferencia_activa(dif, 'cesados'):
            objetivo = abs(_declaracion_afp_resumen_int(dif.get('cesados')))
        elif _declaracion_afp_diferencia_activa(dif, 'antiguos'):
            objetivo = abs(_declaracion_afp_resumen_int(dif.get('antiguos')))
        if not objetivo:
            continue
        agregar(
            person,
            row.get('documentnumber'),
            _declaracion_afp_nombre_completo_trabajador(row),
            'clasificación distinta: planilla '
            f'{_declaracion_afp_describe_flags(*flags_planilla)} vs AFPnet '
            f'{_declaracion_afp_describe_flags(*flags_afpnet)}.',
        )

    return detalles


def _declaracion_afp_flags_por_persona(filas):
    """Unifica banderas inicio/cese por trabajador (evita duplicar filas PR_EmployeeAFP)."""
    flags = {}
    for row in filas or []:
        person = str(row.get('person') or '').strip()
        if not person:
            continue
        inicio = _declaracion_afp_inicio_relacion_mes(row)
        cese = _declaracion_afp_cese_relacion_mes(row)
        if person not in flags:
            flags[person] = {'inicio': inicio, 'ceso': cese}
            continue
        flags[person]['inicio'] = flags[person]['inicio'] or inicio
        flags[person]['ceso'] = flags[person]['ceso'] or cese
    return flags


def _declaracion_afp_conteo_trabajadores_afpnet(filas):
    """AFPnet: nuevos/cesados por bandera S; total suma categorías (puede duplicar ingreso+ceso)."""
    flags = _declaracion_afp_flags_por_persona(filas)
    nuevos = sum(1 for f in flags.values() if f['inicio'])
    cesados = sum(1 for f in flags.values() if f['ceso'])
    antiguos = sum(
        1 for f in flags.values()
        if not f['inicio'] and not f['ceso']
    )
    total_sum = nuevos + cesados + antiguos
    return {
        'nuevos': nuevos,
        'cesados': cesados,
        'antiguos': antiguos,
        'total': total_sum,
        'total_planilla': len(flags),
    }


def _declaracion_afp_ejecutar_resumen_planilla(cursor, p):
    cursor.execute(
        "EXEC sp_pr_resumen_declaracion_afp_web "
        "@cia=?, @period=?, @payroll_all=?, @payroll=?, @afp_all=?, @afp=?",
        (
            p['cia'], p['period'], p['payroll_all'], p['payroll'] or None,
            p['afp_all'], p['afp'] or None,
        ),
    )
    sets = _dicts_collect_nonempty_resultsets(cursor)
    montos_rows = sets[0] if len(sets) > 0 else []
    planilla_row = sets[1][0] if len(sets) > 1 and sets[1] else {}
    planilla_trabajadores = sets[2] if len(sets) > 2 else []
    return montos_rows, planilla_row, planilla_trabajadores


def _declaracion_afp_resumen_int(valor):
    try:
        return int(valor or 0)
    except (TypeError, ValueError):
        return 0


def _declaracion_afp_build_resumen(montos_rows, planilla_row, filas, planilla_trabajadores=None):
    montos = []
    total_planilla = 0.0
    for r in montos_rows or []:
        proc = str(r.get('proceso') or '').strip()
        try:
            monto = round(float(r.get('monto') or 0), 2)
        except (TypeError, ValueError):
            monto = 0.0
        total_planilla += monto
        montos.append({
            'proceso': proc,
            'proceso_label': _declaracion_afp_etiqueta_proceso(proc),
            'monto': monto,
        })
    total_planilla = round(total_planilla, 2)
    total_afpnet = round(sum(_afpnet_remuneracion(r) for r in (filas or [])), 2)

    planilla_nuevos = _declaracion_afp_resumen_int(planilla_row.get('nuevos'))
    planilla_cesados = _declaracion_afp_resumen_int(planilla_row.get('cesados'))
    planilla_antiguos = _declaracion_afp_resumen_int(planilla_row.get('antiguos'))
    planilla_total_planilla = _declaracion_afp_resumen_int(
        planilla_row.get('total_planilla', planilla_row.get('total'))
    )
    planilla_total_sum = planilla_nuevos + planilla_cesados + planilla_antiguos
    planilla = {
        'nuevos': planilla_nuevos,
        'cesados': planilla_cesados,
        'antiguos': planilla_antiguos,
        'total': planilla_total_sum,
        'total_planilla': planilla_total_planilla,
    }
    afpnet = _declaracion_afp_conteo_trabajadores_afpnet(filas)
    diferencia = {
        'nuevos': planilla['nuevos'] - afpnet['nuevos'],
        'cesados': planilla['cesados'] - afpnet['cesados'],
        'antiguos': planilla['antiguos'] - afpnet['antiguos'],
        'total': planilla['total'] - afpnet['total'],
        'total_planilla': planilla['total_planilla'] - afpnet['total_planilla'],
    }
    detalle_diferencias = _declaracion_afp_detalle_diferencias_trabajadores(
        filas, diferencia, planilla_trabajadores, planilla=planilla, afpnet=afpnet,
    )
    diferencia = _declaracion_afp_diferencia_con_gap_interno(diferencia, planilla, afpnet)
    return {
        'montos_proceso': montos,
        'total_afecto_planilla': total_planilla,
        'remuneracion_afecta_afpnet': total_afpnet,
        'diferencia_montos': round(total_planilla - total_afpnet, 2),
        'trabajadores_planilla': planilla,
        'trabajadores_afpnet': afpnet,
        'diferencia_trabajadores': diferencia,
        'detalle_diferencias_trabajadores': detalle_diferencias,
    }


def _declaracion_afp_resumen_tiene_diferencias(resumen):
    """True si hay diferencias de montos o conteo de trabajadores."""
    if not resumen:
        return False
    try:
        if abs(float(resumen.get('diferencia_montos') or 0)) >= 0.005:
            return True
    except (TypeError, ValueError):
        pass
    dif = resumen.get('diferencia_trabajadores') or {}
    for key in ('nuevos', 'cesados', 'antiguos', 'total', 'total_planilla'):
        if _declaracion_afp_resumen_int(dif.get(key)) != 0:
            return True
    if resumen.get('detalle_diferencias_trabajadores'):
        return True
    return False


def _plame_rows_archivo15_from_json(body):
    rows = body.get('rows')
    if not isinstance(rows, list):
        return []
    resultado = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        doc_num = str(r.get('documentnumber') or '').strip()
        susp = str(r.get('suspensiontype') or r.get('pdt') or '').strip()
        if not doc_num or not susp:
            continue
        try:
            days = int(float(r.get('days') or 0))
        except (TypeError, ValueError):
            days = 0
        if days <= 0:
            continue
        resultado.append({
            'documenttype': str(r.get('documenttype') or '').strip(),
            'documentnumber': doc_num,
            'suspensiontype': susp,
            'days': days,
        })
    return resultado


def _plame_filename(ruc, yyyymm, codigo_archivo):
    digits = ''.join(ch for ch in str(ruc or '') if ch.isdigit())
    ruc11 = digits.zfill(11)[-11:] if digits else '00000000000'
    yyyy = yyyymm[:4]
    mm = yyyymm[4:6]
    ext = PLAME_ARCHIVO_EXTENSION.get(str(codigo_archivo), 'txt')
    if str(codigo_archivo) in ('14', '15', '18', '26'):
        return f'0601{yyyy}{mm}{ruc11}.{ext}'
    return f'RP_{ruc11}_{yyyy}_{mm}_{codigo_archivo}.{ext}'


def _obtener_ruc_compania(cursor, cia):
    """Obtiene RUC de SY_Company para nomenclatura de archivos PLAME."""
    queries = (
        "SELECT TOP 1 LTRIM(RTRIM(ISNULL(RUC, ''))) FROM SY_Company WHERE Company = ?",
        "SELECT TOP 1 LTRIM(RTRIM(ISNULL(Ruc, ''))) FROM SY_Company WHERE Company = ?",
    )
    for sql in queries:
        try:
            cursor.execute(sql, (cia,))
            row = cursor.fetchone()
            if row and row[0]:
                return str(row[0]).strip()
        except Exception:
            continue
    return ''


def _plame_rows_from_json(body):
    rows = body.get('rows')
    if not isinstance(rows, list):
        return []
    resultado = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        doc_num = str(r.get('documentnumber') or '').strip()
        if not doc_num:
            continue
        resultado.append({
            'documenttype': str(r.get('documenttype') or '').strip(),
            'documentnumber': doc_num,
            'workinghours': r.get('workinghours'),
            'workingminutes': r.get('workingminutes'),
            'extrahours': r.get('extrahours'),
            'extraminutes': r.get('extraminutes'),
        })
    return resultado


def _asignacion_concepto_detalle_dict(r):
    conceptvalue = r.get('conceptvalue')
    try:
        conceptvalue_num = float(conceptvalue) if conceptvalue is not None else 0.0
    except Exception:
        conceptvalue_num = 0.0
    return {
        "person": _jsonable_value(r.get('person')),
        "nombre": _jsonable_value(r.get('nombre')),
        "employeecode": _jsonable_value(r.get('employeecode')),
        "company": _jsonable_value(r.get('company')),
        "concept": _jsonable_value(r.get('concept')),
        "conceptname": _jsonable_value(r.get('conceptname')),
        "payrolltype": _jsonable_value(r.get('payrolltype')),
        "prperiodstart": _jsonable_value(r.get('prperiodstart')),
        "prperiodend": _jsonable_value(r.get('prperiodend')),
        "conceptvalue": conceptvalue_num,
        "conceptcurrency": _jsonable_value(r.get('conceptcurrency')),
        "flagapplyformula": _jsonable_value(r.get('flagapplyformula')),
        "flagfrecuencytype": _jsonable_value(r.get('flagfrecuencytype')),
        "costcenter": _jsonable_value(r.get('costcenter')),
        "costcentercode": _jsonable_value(r.get('costcentercode')),
    }


def _fmt_periodo_yyyy_mm(val):
    """Periodo para columnas de reporte: YYYY-MM (p. ej. 20250301 → 2025-03)."""
    if val is None:
        return ''
    s = re.sub(r'\D', '', str(val).strip())
    if len(s) >= 6:
        return f'{s[:4]}-{s[4:6]}'
    return str(val).strip()


def _parse_report_date(fecha_raw):
    """Parsea fecha del filtro (YYYY-MM-DD o dd/mm/yyyy); hoy si viene vacía o inválida."""
    s = str(fecha_raw or '').strip()
    if not s:
        return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)


def _anios_saldo_vacaciones(fecha_dt):
    """Años de las columnas saldo1..saldo5 según lógica del SP (year(@date)+1)."""
    ref = fecha_dt.year + 1
    return [ref - 5, ref - 4, ref - 3, ref - 2, ref - 1]


def _normalize_cesados_saldo_vacaciones(raw):
    """Todos=T, Si=Y, No=N (acepta letra directa del SP)."""
    s = str(raw or 'T').strip().upper()
    if s in ('Y', 'S', 'SI', 'SÍ'):
        return 'Y'
    if s in ('N', 'NO'):
        return 'N'
    return 'T'


def _normalize_estado_trabajador(raw):
    """Todos=T, Activo=A, Inactivo=I (por defecto solo activos)."""
    s = str(raw or 'A').strip().upper()
    if s in ('T', 'A', 'I'):
        return s
    return 'A'


def _normalize_cesados_telecredito(raw, default='T'):
    v = str(raw or default).strip().upper()
    return v if v in ('T', 'Y', 'N') else default


def _parse_optional_date(fecha_raw):
    """Fecha opcional YYYY-MM-DD o dd/mm/yyyy; None si vacía o inválida."""
    s = str(fecha_raw or '').strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _sql_date_str_param(fecha_raw):
    """Cadena YYYY-MM-DD para ODBC legacy; vacío si no hay fecha válida."""
    if fecha_raw is None:
        return ''
    if isinstance(fecha_raw, date):
        return fecha_raw.strftime('%Y-%m-%d')
    parsed = _parse_optional_date(fecha_raw)
    return parsed.strftime('%Y-%m-%d') if parsed else ''


def _trabajadores_fecha_ingreso_from_json(body):
    """Sin filtro o rango con ISNULL(ReEntryDate, EntryDate) si está activo el checkbox."""
    body = body or {}
    activo = body.get('fecha_ingreso_activo')
    if activo is None:
        activo = body.get('fechaIngresoActivo')
    if activo is None:
        modo = str(
            body.get('fecha_ingreso_modo')
            or body.get('fechaIngresoModo')
            or body.get('fecha_ingreso_all')
            or 'T'
        ).strip().upper()
        activo = modo in ('N', 'R', 'RANGO', 'RANGE')
    elif isinstance(activo, str):
        activo = activo.strip().lower() in ('1', 'true', 'y', 's', 'si', 'sí')
    if not activo:
        return 'Y', None, None
    desde = _parse_optional_date(body.get('fecha_ingreso_desde') or body.get('fechaIngresoDesde'))
    hasta = _parse_optional_date(body.get('fecha_ingreso_hasta') or body.get('fechaIngresoHasta'))
    return 'N', desde, hasta


def _normalize_todos_bancos_banbif(raw, default='N'):
    v = str(raw or default).strip().upper()
    if v in ('Y', 'S', '1', 'TRUE', 'ON'):
        return 'Y'
    return 'N'


def _telecredito_params_from_json(body):
    """Parámetros comunes para listado y generación Telecrédito BCP."""
    body = body or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    currency = str(body.get('currency') or body.get('par_currency') or 'LO').strip().upper()
    concept = str(body.get('concept') or body.get('par_concept') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period') or body.get('par_period'))
    processtype = str(body.get('processtype') or body.get('process') or '').strip()
    paydate = _parse_report_date(body.get('paydate') or body.get('par_paydate'))
    return {
        'cia': cia,
        'currency': currency,
        'concept': concept,
        'payrolltype': payrolltype,
        'period': period,
        'processtype': processtype,
        'paydate': paydate,
    }


def _pago_haberes_cargar_personas_temp(cursor, persons, temp_table):
    """Carga tabla temporal #TelecreditoPersonas o #InterbankPersonas en lotes."""
    allowed = {'TelecreditoPersonas', 'InterbankPersonas', 'ContinentalPersonas', 'BanbifPersonas'}
    if temp_table not in allowed:
        raise ValueError('Tabla temporal no permitida.')
    cursor.execute(
        f"""
        IF OBJECT_ID('tempdb..#{temp_table}') IS NOT NULL
            DROP TABLE #{temp_table};
        CREATE TABLE #{temp_table} (person VARCHAR(30) NOT NULL PRIMARY KEY);
        """
    )
    if not persons:
        return
    batch_size = 200
    for i in range(0, len(persons), batch_size):
        chunk = [str(p)[:30] for p in persons[i:i + batch_size]]
        placeholders = ','.join(['(?)'] * len(chunk))
        cursor.execute(
            f"INSERT INTO #{temp_table} (person) VALUES {placeholders}",
            chunk,
        )


def _telecredito_cargar_personas_temp(cursor, persons):
    _pago_haberes_cargar_personas_temp(cursor, persons, 'TelecreditoPersonas')


def _telecredito_persons_from_json(body):
    raw = body.get('persons') or body.get('trabajadores') or body.get('person') or []
    if isinstance(raw, str):
        parts = [p.strip() for p in raw.split(',') if p.strip()]
    elif isinstance(raw, list):
        parts = [str(p).strip() for p in raw if str(p).strip()]
    else:
        parts = []
    return parts


def _telecredito_validar_params(params):
    if not params['cia']:
        return 'Seleccione una compañía.'
    if not params['payrolltype']:
        return 'Seleccione tipo de planilla.'
    if not params['processtype']:
        return 'Seleccione proceso.'
    if not params['period']:
        return 'Seleccione periodo.'
    if not params['concept']:
        return 'Indique el concepto de pago.'
    if params['currency'] not in ('LO', 'EX'):
        return 'Moneda inválida (use LO o EX).'
    return None


def _telecredito_campos_faltantes(conn, cia, lineas, currency='LO'):
    """Campos de configuración aún no resueltos o con valor por defecto."""
    faltantes = []
    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT
                ISNULL(ba.BankAccountNumber, '') AS bankaccountnumber,
                ISNULL(m.SalaryAccountType, '') AS salaryaccounttype,
                ISNULL(m.Creditobank, '') AS creditobank,
                ISNULL(tat.abrev, '') AS accounttypeabrev
            FROM pr_mapping m
                LEFT JOIN TE_BankAccount ba
                    ON ba.Company = m.Company
                   AND ba.Bank = m.Creditobank
                   AND ba.accountcurrency = ?
                LEFT JOIN TE_accounttype tat
                    ON tat.AccountType = m.SalaryAccountType
            WHERE m.Company = ?
            """,
            (currency, cia),
        )
        row = cursor.fetchone()
        if not row:
            faltantes.append('pr_mapping: no hay configuración para la compañía.')
        else:
            if not str(row[0] or '').strip():
                faltantes.append(
                    'Cuenta origen empresa (TE_BankAccount.BankAccountNumber '
                    'según CreditoBank y moneda).'
                )
            if not str(row[1] or '').strip():
                faltantes.append('Tipo cuenta origen empresa (pr_mapping.SalaryAccountType).')
            elif not str(row[3] or '').strip():
                faltantes.append('Abreviatura tipo cuenta origen (TE_accounttype.abrev).')
            if not str(row[2] or '').strip():
                faltantes.append('Banco Telecrédito (pr_mapping.Creditobank).')
    except Exception:
        faltantes.append(
            'Configuración TE_BankAccount / pr_mapping (BankAccountNumber, '
            'SalaryAccountType, Creditobank): verificar nombres de columnas en BD.'
        )

    if not lineas:
        faltantes.append('No se generaron líneas de detalle para los trabajadores seleccionados.')

    return faltantes


def _telecredito_filename(period):
    periodo = re.sub(r'[^0-9]', '', str(period or ''))[:8]
    stamp = datetime.now().strftime('%Y%m%d%H%M')
    return f'Telecredito_{periodo}_{stamp}.txt'


def _interbank_filename(period):
    periodo = re.sub(r'[^0-9]', '', str(period or ''))[:8]
    stamp = datetime.now().strftime('%Y%m%d%H%M')
    return f'Interbank_{periodo}_{stamp}.txt'


def _continental_filename(period):
    periodo = re.sub(r'[^0-9]', '', str(period or ''))[:8]
    stamp = datetime.now().strftime('%Y%m%d%H%M')
    return f'Continental_{periodo}_{stamp}.txt'


def _banbif_filename(period):
    periodo = re.sub(r'[^0-9]', '', str(period or ''))[:8]
    stamp = datetime.now().strftime('%Y%m%d%H%M')
    return f'Banbif_{periodo}_{stamp}.txt'


def _report_params_from_json(req):
    """Extrae y normaliza los 4 parámetros del SP (mismo orden que SSMS)."""
    body = req.get_json(silent=True) or {}
    cia = str(body.get('cia') or '').strip()
    payrolltype = str(body.get('payrolltype') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    person = str(body.get('person') or '').strip()
    if not (cia and payrolltype and period and person):
        return None
    return (cia, payrolltype, period, person)


def _fetch_first_nonempty_resultset(cursor):
    """
    Algunos SP devuelven resultsets vacíos antes del SELECT final;
    avanza con nextset() hasta encontrar filas (o se acaban los sets).
    """
    columns = []
    rows = []
    while True:
        if cursor.description:
            columns = [_report_column_name(c[0]) for c in cursor.description]
            rows = cursor.fetchall()
            if rows:
                return columns, rows
        if not cursor.nextset():
            break
    return columns, []


def _dicts_first_nonempty_resultset(cursor):
    """
    Igual que _fetch_first_nonempty_resultset pero devuelve filas como dicts
    con claves en minúsculas (robusto con pyodbc / alias del SP).
    """
    while True:
        if cursor.description:
            cols = [str(c[0]).strip() for c in cursor.description]
            rows = cursor.fetchall()
            if rows:
                out = []
                for row in rows:
                    rd = {}
                    for i, cname in enumerate(cols):
                        key = (cname or f"col{i}").lower()
                        rd[key] = row[i]
                    out.append(rd)
                return out
        if not cursor.nextset():
            break
    return []


def _dicts_collect_nonempty_resultsets(cursor, max_sets=10):
    """Lista de resultsets no vacíos; cada uno es una lista de dicts."""
    result = []
    sets_read = 0
    while sets_read < max_sets:
        if cursor.description:
            cols = [str(c[0]).strip() for c in cursor.description]
            rows = cursor.fetchall()
            if rows:
                out = []
                for row in rows:
                    rd = {}
                    for i, cname in enumerate(cols):
                        key = (cname or f"col{i}").lower()
                        rd[key] = row[i]
                    out.append(rd)
                result.append(out)
            sets_read += 1
        if not cursor.nextset():
            break
    return result


def _sanitize_dynamic_procedure_name(name):
    """
    Valida ProcedureName leído de PR_ProcessType antes de usarlo en {{CALL ...}}.
    Permite esquema.procedimiento (segmentos alfanuméricos / guión bajo).
    """
    s = str(name or "").strip()
    if not s or len(s) > 200 or ".." in s:
        return None
    for part in s.split("."):
        if not part or not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", part):
            return None
    return s


def _drain_pyodbc_cursor(cursor):
    """Consume resultsets pendientes tras EXEC/CALL (evita errores en la siguiente ejecución)."""
    try:
        while True:
            if cursor.description:
                cursor.fetchall()
            if not cursor.nextset():
                break
    except Exception:
        logging.debug("drenado de cursor", exc_info=True)


def _is_comm_link_failure(err):
    """Detecta caídas transitorias de enlace ODBC/SQL Server."""
    s = str(err or "").lower()
    return ("08s01" in s) or ("communication link failure" in s)


def _is_transient_sql_error(err):
    """Errores reintentables: enlace caído o timeout de comando."""
    s = str(err or "").lower()
    return _is_comm_link_failure(err) or ("hyt00" in s) or ("timeout expired" in s)


def _sql_call_timeout_seconds():
    raw = str(os.getenv("SQL_CALL_TIMEOUT_SEC", "35")).strip()
    try:
        n = int(raw)
    except Exception:
        n = 35
    return max(10, min(n, 180))


def _sql_call_timeout_payroll_seconds():
    """Timeout por trabajador en cálculo de planilla (SP puede ser más lento)."""
    raw = str(os.getenv("SQL_CALL_TIMEOUT_PAYROLL_SEC", "120")).strip()
    try:
        n = int(raw)
    except Exception:
        n = 120
    return max(30, min(n, 600))


def _sql_call_timeout_report_seconds():
    """Timeout para reportes pesados (saldo vacaciones, planilla vertical, etc.)."""
    raw = str(os.getenv("SQL_CALL_TIMEOUT_REPORT_SEC", "100")).strip()
    try:
        n = int(raw)
    except Exception:
        n = 100
    return max(30, min(n, 300))


def _set_cursor_timeout(cursor):
    """Timeout por ejecución de SP (segundos) para evitar cuelgues largos."""
    try:
        cursor.timeout = _sql_call_timeout_seconds()
    except Exception:
        logging.debug("No se pudo fijar timeout en cursor", exc_info=True)


def _set_cursor_timeout_payroll(cursor):
    """Timeout ampliado para procesar planilla trabajador por trabajador."""
    try:
        cursor.timeout = _sql_call_timeout_payroll_seconds()
    except Exception:
        logging.debug("No se pudo fijar timeout payroll en cursor", exc_info=True)


def _set_cursor_timeout_report(cursor):
    """Timeout ampliado para reportes que ejecutan SP lentos en SQL Server."""
    try:
        cursor.timeout = _sql_call_timeout_report_seconds()
    except Exception:
        logging.debug("No se pudo fijar timeout reporte en cursor", exc_info=True)


def _reporte_sql_error_message(err):
    if isinstance(err, SystemError) or _is_transient_sql_error(err):
        return (
            'El reporte tardó demasiado o se interrumpió la consulta. '
            'Intente filtrar por un trabajador específico o una fecha más reciente.'
        )
    return str(err)


def _rows_to_dual_dicts(columns, rows):
    out = []
    for row in rows:
        item = {}
        for i, col in enumerate(columns):
            key = str(col or f'col{i}').strip()
            val = row[i]
            item[key] = val
            item[key.lower()] = val
        out.append(item)
    return out


def _escape_pdf_text(value):
    return str(value or '').replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _generar_pdf_fallback_basico(meta):
    """
    PDF mínimo sin dependencias externas.
    Se usa cuando WeasyPrint no está disponible en el sistema.
    """
    titulo = _escape_pdf_text("BOLETA DE PAGO - MODO COMPATIBLE")
    person = _escape_pdf_text(meta.get("person"))
    nombre = _escape_pdf_text(meta.get("nombre_trabajador") or meta.get("nombre"))
    cia = _escape_pdf_text(meta.get("cia"))
    payroll = _escape_pdf_text(meta.get("payroll_type"))
    proc = _escape_pdf_text(meta.get("process"))
    period = _escape_pdf_text(meta.get("period"))
    fecha = _escape_pdf_text(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

    lines = [
        f"BT /F1 16 Tf 50 790 Td ({titulo}) Tj ET",
        f"BT /F1 11 Tf 50 760 Td (Persona: {person}) Tj ET",
        f"BT /F1 11 Tf 50 742 Td (Nombre: {nombre}) Tj ET",
        f"BT /F1 11 Tf 50 724 Td (Compania: {cia}) Tj ET",
        f"BT /F1 11 Tf 50 706 Td (Tipo planilla: {payroll}) Tj ET",
        f"BT /F1 11 Tf 50 688 Td (Proceso: {proc}) Tj ET",
        f"BT /F1 11 Tf 50 670 Td (Periodo: {period}) Tj ET",
        f"BT /F1 9 Tf 50 640 Td (Generado: {fecha}) Tj ET",
    ]
    content_stream = ("\n".join(lines)).encode("latin-1", errors="replace")

    objects = []
    objects.append(b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n")
    objects.append(b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n")
    objects.append(
        b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n"
    )
    objects.append(b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n")
    objects.append(
        f"5 0 obj << /Length {len(content_stream)} >> stream\n".encode("latin-1")
        + content_stream
        + b"\nendstream endobj\n"
    )

    pdf = io.BytesIO()
    pdf.write(b"%PDF-1.4\n")
    xref_positions = [0]
    for obj in objects:
        xref_positions.append(pdf.tell())
        pdf.write(obj)
    xref_start = pdf.tell()
    pdf.write(f"xref\n0 {len(xref_positions)}\n".encode("latin-1"))
    pdf.write(b"0000000000 65535 f \n")
    for pos in xref_positions[1:]:
        pdf.write(f"{pos:010d} 00000 n \n".encode("latin-1"))
    pdf.write(
        (
            "trailer << /Size "
            + str(len(xref_positions))
            + " /Root 1 0 R >>\nstartxref\n"
            + str(xref_start)
            + "\n%%EOF"
        ).encode("latin-1")
    )
    pdf.seek(0)
    return pdf


def _exec_sp_rows_dicts(cursor, sql, params):
    cursor.execute(sql, params)
    cols, rows = _fetch_first_nonempty_resultset(cursor)
    if not rows:
        return []
    return _rows_to_dual_dicts(cols, rows)


def get_image_base64(file_path):
    if not file_path or not os.path.exists(file_path):
        return ''
    try:
        with open(file_path, 'rb') as f:
            return base64.b64encode(f.read()).decode('utf-8')
    except Exception:
        logging.exception('get_image_base64')
        return ''


def _image_data_uri(file_path):
    """Data URI con MIME según extensión (para WeasyPrint / HTML embebido)."""
    if not file_path or not os.path.exists(file_path):
        return ''
    ext = os.path.splitext(file_path)[1].lower()
    mime = {
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.webp': 'image/webp',
    }.get(ext, 'image/png')
    try:
        with open(file_path, 'rb') as f:
            encoded = base64.b64encode(f.read()).decode('utf-8')
        return f'data:{mime};base64,{encoded}'
    except Exception:
        logging.exception('_image_data_uri')
        return ''


def _boleta_imagen_ruta(img_dir, nombre_archivo):
    """Resuelve ruta en static/img; solo el nombre de archivo (sin rutas)."""
    nombre = os.path.basename(str(nombre_archivo or '').strip())
    if not nombre:
        return ''
    ruta = os.path.join(img_dir, nombre)
    return ruta if os.path.exists(ruta) else ''


def _boleta_imagenes_paths(cia):
    """Logo y firma desde SY_Company (logoname, signaturename) en static/img."""
    img_dir = os.path.join(app.root_path, 'static', 'img')
    cfg = get_config_empresa(cia)
    nombre_logo = str(cfg[0]).strip() if cfg and len(cfg) > 0 and cfg[0] else ''
    nombre_firma = str(cfg[1]).strip() if cfg and len(cfg) > 1 and cfg[1] else ''
    return (
        _boleta_imagen_ruta(img_dir, nombre_logo),
        _boleta_imagen_ruta(img_dir, nombre_firma),
    )


def _bool_env(name, default=False):
    raw = str(os.getenv(name, str(default))).strip().lower()
    return raw in ('1', 'true', 'yes', 'on')


def formatear_periodo_texto(periodo_str):
    # Asumiendo formato YYYYMM... (ej: 20251212)
    meses = {
        "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
        "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
        "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre",
    }
    try:
        periodo_val = str(periodo_str or "").strip()
        anio = periodo_val[:4]
        mes_num = periodo_val[4:6]
        nombre_mes = meses.get(mes_num, "Mes")
        return f"{nombre_mes} {anio}"
    except Exception:
        return str(periodo_str or "")


def enviar_correo_boleta(destinatario, nombre_empleado, periodo, sexo, pdf_io, person=None):
    """Envía boleta por Resend API con PDF adjunto."""
    if not destinatario or '@' not in str(destinatario):
        return False, "Sin correo"

    resend.api_key = _resend_api_key()
    if not resend.api_key:
        return False, 'RESEND_API_KEY no configurada.' + _resend_api_key_diagnostico()
    remitente = _env_var('MAIL_FROM', 'EMAIL_FROM', default='onboarding@resend.dev')

    try:
        sexo_val = int(sexo)
    except Exception:
        sexo_val = 0
    trato = "Estimada" if sexo_val == 2 else "Estimado"
    periodo_legible = formatear_periodo_texto(periodo)
    pdf_base64 = base64.b64encode(pdf_io.getvalue()).decode('utf-8')

    try:
        params = {
            "from": f"Recursos Humanos <{remitente}>",
            "to": destinatario,
            "subject": f"Boleta de Pago - {periodo_legible} - {nombre_empleado}",
            "html": f"""
                <p>{trato} {nombre_empleado},</p>
                <p>Le hacemos entrega de su boleta de pago correspondiente al periodo de <b>{periodo_legible}</b>.</p>
                <p>Saludos,<br>Recursos Humanos</p>
            """,
            "attachments": [
                {
                    "content": pdf_base64,
                    "filename": _boleta_pdf_filename(person or nombre_empleado, periodo),
                }
            ],
        }
        resend.Emails.send(params)
        return True, "Enviado"
    except Exception as e:
        logging.error("Error en Resend: %s", str(e))
        return False, str(e)


def enviar_correo_certificado_quinta(destinatario, nombre_empleado, anio, sexo, pdf_io, person=None):
    """Envía certificado de quinta por Resend API con PDF adjunto."""
    if not destinatario or '@' not in str(destinatario):
        return False, 'Sin correo'

    resend.api_key = _resend_api_key()
    if not resend.api_key:
        return False, 'RESEND_API_KEY no configurada.' + _resend_api_key_diagnostico()
    remitente = _env_var('MAIL_FROM', 'EMAIL_FROM', default='onboarding@resend.dev')

    try:
        sexo_val = int(sexo)
    except Exception:
        sexo_val = 0
    trato = 'Estimada' if sexo_val == 2 else 'Estimado'
    anio_txt = str(anio or '').strip() or 'N/D'
    pdf_base64 = base64.b64encode(pdf_io.getvalue()).decode('utf-8')

    try:
        params = {
            'from': f'Recursos Humanos <{remitente}>',
            'to': destinatario,
            'subject': f'Certificado de Quinta Categoría - Ejercicio {anio_txt} - {nombre_empleado}',
            'html': f"""
                <p>{trato} {nombre_empleado},</p>
                <p>Le hacemos entrega de su certificado de remuneraciones y retenciones sobre rentas de quinta categoría correspondiente al ejercicio gravable <b>{anio_txt}</b>.</p>
                <p>Saludos,<br>Recursos Humanos</p>
            """,
            'attachments': [
                {
                    'content': pdf_base64,
                    'filename': _certificado_quinta_pdf_filename(person or nombre_empleado, anio_txt),
                }
            ],
        }
        resend.Emails.send(params)
        return True, 'Enviado'
    except Exception as e:
        logging.error('Error en Resend certificado quinta: %s', str(e))
        return False, str(e)


def generar_pdf_en_memoria(params):
    cia_param = str(params.get('cia') or '').strip()
    if not cia_param and has_request_context():
        ensure_user_session()
    cia = str(cia_param or (session.get('company') if has_request_context() else '') or '').strip()
    payroll_type = str(params.get('payroll_type') or '').strip()
    processtype = str(params.get('process') or params.get('processtype') or '').strip()
    period = _normalize_pr_period(params.get('period'))
    person = str(params.get('person') or '').strip()
    if not (cia and payroll_type and processtype and period and person):
        raise ValueError('Faltan parámetros para generar boleta.')

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout(cursor)

        cab_rows = _exec_sp_rows_dicts(
            cursor,
            'EXEC sp_pr_generarboleta_web @cia=?, @process=?, @payrolltype=?, @period=?, @person=?',
            (cia, processtype, payroll_type, period, person),
        )
        cabecera = cab_rows[0] if cab_rows else {}

        ingresos = _exec_sp_rows_dicts(
            cursor,
            'EXEC sp_pr_detalleboletaingresos_web @cia=?, @process=?, @payrolltype=?, @period=?, @person=?',
            (cia, processtype, payroll_type, period, person),
        )
        descuentos = _exec_sp_rows_dicts(
            cursor,
            'EXEC sp_pr_detalleboletadescuentos_web @cia=?, @process=?, @payrolltype=?, @period=?, @person=?',
            (cia, processtype, payroll_type, period, person),
        )
        aportes = _exec_sp_rows_dicts(
            cursor,
            'EXEC sp_pr_detalleboletaaportes_web @cia=?, @process=?, @payrolltype=?, @period=?, @person=?',
            (cia, processtype, payroll_type, period, person),
        )
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    ruta_logo, ruta_firma = _boleta_imagenes_paths(cia)
    logo_src = _image_data_uri(ruta_logo)
    firma_src = _image_data_uri(ruta_firma)
    if _bool_env('LOG_BOLETA_ASSETS', False):
        logging.info(
            '[boleta assets] cia=%s logo="%s" exists=%s | firma="%s" exists=%s',
            cia,
            ruta_logo,
            os.path.exists(ruta_logo),
            ruta_firma,
            os.path.exists(ruta_firma),
        )

    if WEASYPRINT_AVAILABLE:
        html_renderizado = render_template(
            'boleta_moderna.html',
            cabecera=cabecera,
            ingresos=ingresos,
            descuentos=descuentos,
            aportes=aportes,
            logo_src=logo_src,
            firma_src=firma_src,
        )
        pdf_io = io.BytesIO()
        HTML(string=html_renderizado).write_pdf(pdf_io)
        pdf_io.seek(0)
        return pdf_io

    logging.warning(
        'WeasyPrint no disponible; usando PDF fallback básico. Motivo: %s',
        _WEASYPRINT_IMPORT_ERROR,
    )
    fallback_meta = dict(cabecera or {})
    fallback_meta['person'] = person
    fallback_meta['cia'] = cia
    fallback_meta['payroll_type'] = payroll_type
    fallback_meta['process'] = processtype
    fallback_meta['period'] = period
    return _generar_pdf_fallback_basico(fallback_meta)


def generar_pdf_certificado_quinta(params):
    cia_param = str(params.get('cia') or '').strip()
    if not cia_param and has_request_context():
        ensure_user_session()
    cia = str(cia_param or (session.get('company') if has_request_context() else '') or '').strip()
    payroll_type = str(params.get('payroll_type') or '').strip()
    anio = str(params.get('anio') or params.get('year') or '').strip()
    person = str(params.get('person') or '').strip()
    if not (cia and payroll_type and anio and person):
        raise ValueError('Faltan parámetros para generar certificado de quinta.')
    if len(anio) != 4 or not anio.isdigit():
        raise ValueError('Año inválido para certificado de quinta.')

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout(cursor)
        rows = _exec_sp_rows_dicts(
            cursor,
            'EXEC sp_pr_certificadoquinta_web @cia=?, @payrolltype=?, @payrolltype_all=?, @anio=?, @person=?, @employee_all=?, @activo=?',
            (cia, payroll_type, 'N', anio, person, 'N', 'N'),
        )
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    cert = rows[0] if rows else {}
    if not cert:
        raise ValueError('No se encontraron datos para el certificado de quinta.')

    ruta_logo, ruta_firma = _boleta_imagenes_paths(cia)
    logo_src = _image_data_uri(ruta_logo)
    firma_src = _image_data_uri(ruta_firma)

    try:
        cert['importe_remuneracion_bruta_total'] = (
            float(cert.get('importe_sueldos_asignaciones') or 0)
            + float(cert.get('importe_participacion_utilidades') or 0)
            + float(cert.get('importe_remuneracion_otras_empresas') or 0)
        )
    except (TypeError, ValueError):
        cert['importe_remuneracion_bruta_total'] = 0.0

    try:
        deduccion_7uit = float(cert.get('importe_deduccion_7uit') or 0)
        renta_neta = cert['importe_remuneracion_bruta_total'] - deduccion_7uit
        cert['importe_renta_neta'] = renta_neta if renta_neta > 0 else 0.0
    except (TypeError, ValueError):
        cert['importe_renta_neta'] = 0.0

    cert['importe_impuesto_renta'] = _calcular_impuesto_renta_quinta(
        cert.get('importe_renta_neta'),
        cert.get('uit'),
    )

    try:
        cert['importe_impuesto_retenido_exceso'] = (
            float(cert.get('importe_impuesto_renta') or 0)
            - float(cert.get('importe_impuesto_total_retenido') or 0)
        )
    except (TypeError, ValueError):
        cert['importe_impuesto_retenido_exceso'] = 0.0

    monto_sueldos = float(cert.get('importe_sueldos_asignaciones') or 0)
    texto_remuneracion = f'S/. {format_importe(monto_sueldos)}'
    monto_retenido = float(cert.get('importe_impuesto_total_retenido') or 0)
    texto_retenido = f'S/. {format_importe(monto_retenido)}'

    html_renderizado = render_template(
        'certificado_quinta_pdf.html',
        cert=cert,
        anio=anio,
        logo_src=logo_src,
        firma_src=firma_src,
        tratamiento=_sexo_tratamiento_certificado(cert.get('sexo')),
        texto_retenido=texto_retenido,
        texto_remuneracion=texto_remuneracion,
        fecha_emision_texto=_fecha_emision_certificado_quinta(params.get('fecha_emision')),
    )

    if WEASYPRINT_AVAILABLE:
        pdf_io = io.BytesIO()
        HTML(string=html_renderizado).write_pdf(pdf_io)
        pdf_io.seek(0)
        return pdf_io

    raise RuntimeError(
        'WeasyPrint no está disponible para generar el certificado de quinta. '
        + str(_WEASYPRINT_IMPORT_ERROR or '')
    )


def generar_pdf_certificado_trabajo(params):
    cia_param = str(params.get('cia') or '').strip()
    if not cia_param and has_request_context():
        ensure_user_session()
    cia = str(cia_param or (session.get('company') if has_request_context() else '') or '').strip()
    payroll_type = str(params.get('payroll_type') or '').strip()
    period = _normalize_pr_period(params.get('period'))
    person = str(params.get('person') or '').strip()
    if not (cia and payroll_type and period and person):
        raise ValueError('Faltan parámetros para generar certificado de trabajo.')

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout(cursor)
        rows = _exec_sp_rows_dicts(
            cursor,
            'EXEC sp_pr_certificadotrabajo_web @cia=?, @payrolltype=?, @period=?, @person=?',
            (cia, payroll_type, period, person),
        )
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    cert = rows[0] if rows else {}
    if not cert:
        raise ValueError('No se encontraron datos para el certificado de trabajo.')

    ruta_logo, ruta_firma = _boleta_imagenes_paths(cia)
    logo_src = _image_data_uri(ruta_logo)
    firma_src = _image_data_uri(ruta_firma)

    html_renderizado = render_template(
        'certificado_trabajo_pdf.html',
        cert=cert,
        logo_src=logo_src,
        firma_src=firma_src,
        tratamiento=_tratamiento_certificado_trabajo(cert.get('sex')),
        fecha_emision_texto=_fecha_emision_certificado_trabajo(cert),
    )

    if WEASYPRINT_AVAILABLE:
        pdf_io = io.BytesIO()
        HTML(string=html_renderizado).write_pdf(pdf_io)
        pdf_io.seek(0)
        return pdf_io

    raise RuntimeError(
        'WeasyPrint no está disponible para generar el certificado de trabajo. '
        + str(_WEASYPRINT_IMPORT_ERROR or '')
    )


def enviar_correo_certificado_trabajo(destinatario, nombre_empleado, periodo, sexo, pdf_io, person=None):
    """Envía certificado de trabajo por Resend API con PDF adjunto."""
    if not destinatario or '@' not in str(destinatario):
        return False, 'Sin correo'

    resend.api_key = _resend_api_key()
    if not resend.api_key:
        return False, 'RESEND_API_KEY no configurada.' + _resend_api_key_diagnostico()
    remitente = _env_var('MAIL_FROM', 'EMAIL_FROM', default='onboarding@resend.dev')

    try:
        sexo_val = int(sexo)
    except Exception:
        sexo_val = 0
    trato = 'Estimada' if sexo_val == 2 else 'Estimado'
    periodo_legible = formatear_periodo_texto(periodo)
    pdf_base64 = base64.b64encode(pdf_io.getvalue()).decode('utf-8')

    try:
        params = {
            'from': f'Recursos Humanos <{remitente}>',
            'to': destinatario,
            'subject': f'Certificado de Trabajo - {periodo_legible} - {nombre_empleado}',
            'html': f"""
                <p>{trato} {nombre_empleado},</p>
                <p>Le hacemos entrega de su certificado de trabajo correspondiente al periodo de <b>{periodo_legible}</b>.</p>
                <p>Saludos,<br>Recursos Humanos</p>
            """,
            'attachments': [
                {
                    'content': pdf_base64,
                    'filename': _certificado_trabajo_pdf_filename(person or nombre_empleado, periodo),
                }
            ],
        }
        resend.Emails.send(params)
        return True, 'Enviado'
    except Exception as e:
        logging.error('Error en Resend certificado trabajo: %s', str(e))
        return False, str(e)


def generar_pdf_certificado_retiro_cts(params):
    cia_param = str(params.get('cia') or '').strip()
    if not cia_param and has_request_context():
        ensure_user_session()
    cia = str(cia_param or (session.get('company') if has_request_context() else '') or '').strip()
    payroll_type = str(params.get('payroll_type') or '').strip()
    period = _normalize_pr_period(params.get('period'))
    person = str(params.get('person') or '').strip()
    if not (cia and payroll_type and period and person):
        raise ValueError('Faltan parámetros para generar certificado retiro CTS.')

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout(cursor)
        rows = _exec_sp_rows_dicts(
            cursor,
            'EXEC sp_pr_certificadoretirocts_web @cia=?, @payrolltype=?, @period=?, @person=?',
            (cia, payroll_type, period, person),
        )
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    cert = rows[0] if rows else {}
    if not cert:
        raise ValueError('No se encontraron datos para el certificado retiro CTS.')

    ruta_logo, ruta_firma = _boleta_imagenes_paths(cia)
    logo_src = _image_data_uri(ruta_logo)
    firma_src = _image_data_uri(ruta_firma)

    html_renderizado = render_template(
        'certificado_retiro_cts_pdf.html',
        cert=cert,
        logo_src=logo_src,
        firma_src=firma_src,
        texto_autorizacion=_texto_autorizacion_retiro_cts(cert),
    )

    if WEASYPRINT_AVAILABLE:
        pdf_io = io.BytesIO()
        HTML(string=html_renderizado).write_pdf(pdf_io)
        pdf_io.seek(0)
        return pdf_io

    raise RuntimeError(
        'WeasyPrint no está disponible para generar el certificado retiro CTS. '
        + str(_WEASYPRINT_IMPORT_ERROR or '')
    )


def _formato_liquidacion_pdf_filename(person, period_raw):
    person_safe = re.sub(r'[^A-Za-z0-9_\\-]+', '_', str(person or 'preview').strip()).strip('_') or 'preview'
    period_safe = re.sub(r'[^0-9]+', '', _normalize_pr_period(period_raw)) or 'periodo'
    return f'formato_liquidacion_{person_safe}_{period_safe}.pdf'


def _formato_liquidacion_fecha(val):
    ref = _parse_fecha_flexible(val)
    return ref.strftime('%d/%m/%Y') if ref else ''


def _formato_liquidacion_moneda(val):
    try:
        n = float(val or 0)
    except (TypeError, ValueError):
        n = 0.0
    return f'S/ {n:,.2f}'


def _formato_liquidacion_porcentaje(val):
    try:
        n = float(val or 0)
    except (TypeError, ValueError):
        n = 0.0
    if n <= 0:
        return ''
    text = f'{n:.2f}'.rstrip('0').rstrip('.')
    return f'{text}%'


def _regimen_pensionario_formato_liquidacion(liq):
    liq = liq or {}
    tipo_pension = str(liq.get('type_pension') or '').strip()
    tipo_comision = str(liq.get('tipo_comision') or '').strip()
    if tipo_pension and tipo_comision:
        return f'{tipo_pension} - {tipo_comision}'
    return tipo_pension or tipo_comision


_FORMATO_LIQ_REMUNERACION_DEF = (
    {'label': 'Básico', 'cts': 'BASICO_CTS', 'grati': 'BASICO_GRATI', 'vaca': 'BASICO_VACA'},
    {'label': 'Asig. fam.', 'cts': 'ASIG_FAM_CTS', 'grati': 'ASIG_FAM_GRATI', 'vaca': 'ASIG_FAM_VACA'},
    {'label': 'Sexto grati.', 'cts': 'PROMEDIO_GRATI', 'grati': None, 'vaca': None},
    {'label': 'Promedio noche', 'cts': 'BONO_PROD_CTS', 'grati': 'BONO_PROD_GRATI', 'vaca': 'BONO_PROD_VAC'},
    {'label': 'Promedio HE', 'cts': 'HRS_EXTRAS_25_CTS', 'grati': 'HRS_EXTRAS_25_GRA', 'vaca': 'HRS_EXTRAS_25_VAC'},
    {'label': 'Promedio feriado', 'cts': 'LIQ_PROM_COMI_CTS', 'grati': 'LIQ_PROM_COMI_GRA', 'vaca': 'LIQ_PROM_COMI_VAC'},
)


def _fetch_formato_liquidacion_formulacodes(cursor, cia, payroll_type, period, person):
    codes = sorted({
        code
        for row in _FORMATO_LIQ_REMUNERACION_DEF
        for code in (row.get('cts'), row.get('grati'), row.get('vaca'))
        if code
    })
    if not codes:
        return {}

    placeholders = ','.join('?' for _ in codes)
    sql = f"""
        SELECT
            C.FormulaCode AS formula_code,
            SUM(ISNULL(EC.ConceptValueLo, EC.ConceptValue)) AS valor
        FROM PR_EmployeePayRollConcept EC (NOLOCK)
            INNER JOIN PR_Concept C (NOLOCK) ON EC.Concept = C.Concept
        WHERE EC.Company = ?
          AND EC.Person = ?
          AND EC.PayRollType = ?
          AND EC.PRPeriod = ?
          AND EXISTS (
                SELECT 1
                FROM PR_ProcessType PT (NOLOCK)
                WHERE PT.ShortName = 'LIQUIDACION'
                  AND PT.ProcessType = EC.ProcessType
            )
          AND C.FormulaCode IN ({placeholders})
        GROUP BY C.FormulaCode
    """
    cursor.execute(sql, (cia, person, payroll_type, period, *codes))
    rows = _dicts_first_nonempty_resultset(cursor)
    valores = {}
    for row in rows:
        fc = str(row.get('formula_code') or '').strip()
        if not fc:
            continue
        try:
            valores[fc] = float(row.get('valor') or 0)
        except (TypeError, ValueError):
            valores[fc] = 0.0
    return valores


def _build_formato_liquidacion_remuneracion(formula_values):
    formula_values = formula_values or {}
    filas = []
    totales = {'cts': 0.0, 'grati': 0.0, 'vaca': 0.0}
    for defn in _FORMATO_LIQ_REMUNERACION_DEF:
        fila = {'label': defn['label']}
        for col in ('cts', 'grati', 'vaca'):
            fc = defn.get(col)
            valor = float(formula_values.get(fc, 0) or 0) if fc else 0.0
            totales[col] += valor
            fila[f'{col}_fmt'] = _formato_liquidacion_moneda(valor)
        filas.append(fila)
    totales_fmt = {
        col: _formato_liquidacion_moneda(totales[col])
        for col in ('cts', 'grati', 'vaca')
    }
    return filas, totales_fmt


def generar_pdf_formato_liquidacion(params):
    cia_param = str(params.get('cia') or '').strip()
    if not cia_param and has_request_context():
        ensure_user_session()
    cia = str(cia_param or (session.get('company') if has_request_context() else '') or '').strip()
    payroll_type = str(params.get('payroll_type') or '').strip()
    period = _normalize_pr_period(params.get('period'))
    person = str(params.get('person') or '').strip()
    if not (cia and payroll_type and period and person):
        raise ValueError('Faltan parámetros para generar el formato de liquidación.')

    conn = None
    formula_values = {}
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout(cursor)
        rows = _exec_sp_rows_dicts(
            cursor,
            'EXEC sp_pr_formatoliquidacion_web @cia=?, @payrolltype=?, @period=?, @person=?',
            (cia, payroll_type, period, person),
        )
        formula_values = _fetch_formato_liquidacion_formulacodes(
            cursor, cia, payroll_type, period, person
        )
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    liq = rows[0] if rows else {}
    if not liq:
        raise ValueError('No se encontraron datos para el formato de liquidación.')

    ruta_logo, _ruta_firma = _boleta_imagenes_paths(cia)
    logo_src = _image_data_uri(ruta_logo)
    cero = _formato_liquidacion_moneda(0)
    cero_pct = '0.00%'
    porc_onp_val = float(liq.get('porc_onp') or 0)
    remuneracion_rows, remuneracion_totales = _build_formato_liquidacion_remuneracion(formula_values)

    html_renderizado = render_template(
        'formato_liquidacion_pdf.html',
        liq=liq,
        logo_src=logo_src,
        regimen_pensionario=_regimen_pensionario_formato_liquidacion(liq),
        basico_fmt=_formato_liquidacion_moneda(liq.get('basico')),
        entry_date_fmt=_formato_liquidacion_fecha(liq.get('entry_date')),
        cease_date_fmt=_formato_liquidacion_fecha(liq.get('cease_date')),
        porc_aporte=_formato_liquidacion_porcentaje(liq.get('porc_aporte')),
        porc_comision_fija=_formato_liquidacion_porcentaje(liq.get('porc_comision_fija')),
        porc_comision_mixta=_formato_liquidacion_porcentaje(liq.get('porc_comision_mixta')),
        porc_seguro=_formato_liquidacion_porcentaje(liq.get('porc_seguro')),
        porc_onp=_formato_liquidacion_porcentaje(porc_onp_val) if porc_onp_val > 0 else '',
        cero=cero,
        cero_pct=cero_pct,
        remuneracion_rows=remuneracion_rows,
        remuneracion_totales=remuneracion_totales,
    )

    if WEASYPRINT_AVAILABLE:
        pdf_io = io.BytesIO()
        HTML(string=html_renderizado).write_pdf(pdf_io)
        pdf_io.seek(0)
        return pdf_io

    raise RuntimeError(
        'WeasyPrint no está disponible para generar el formato de liquidación. '
        + str(_WEASYPRINT_IMPORT_ERROR or '')
    )


def enviar_correo_formato_liquidacion(destinatario, nombre_empleado, periodo, sexo, pdf_io, person=None):
    if not destinatario or '@' not in str(destinatario):
        return False, 'Sin correo'

    resend.api_key = _resend_api_key()
    if not resend.api_key:
        return False, 'RESEND_API_KEY no configurada.' + _resend_api_key_diagnostico()
    remitente = _env_var('MAIL_FROM', 'EMAIL_FROM', default='onboarding@resend.dev')

    try:
        sexo_val = int(sexo)
    except Exception:
        sexo_val = 0
    trato = 'Estimada' if sexo_val == 2 else 'Estimado'
    periodo_legible = formatear_periodo_texto(periodo)
    pdf_base64 = base64.b64encode(pdf_io.getvalue()).decode('utf-8')

    try:
        params = {
            'from': f'Recursos Humanos <{remitente}>',
            'to': destinatario,
            'subject': f'Formato de Liquidación - {periodo_legible} - {nombre_empleado}',
            'html': f"""
                <p>{trato} {nombre_empleado},</p>
                <p>Le hacemos entrega de su formato de liquidación correspondiente al periodo de <b>{periodo_legible}</b>.</p>
                <p>Saludos,<br>Recursos Humanos</p>
            """,
            'attachments': [
                {
                    'content': pdf_base64,
                    'filename': _formato_liquidacion_pdf_filename(person or nombre_empleado, periodo),
                }
            ],
        }
        resend.Emails.send(params)
        return True, 'Enviado'
    except Exception as e:
        logging.error('Error en Resend formato liquidación: %s', str(e))
        return False, str(e)


def enviar_correo_certificado_retiro_cts(destinatario, nombre_empleado, periodo, sexo, pdf_io, person=None):
    """Envía certificado retiro CTS por Resend API con PDF adjunto."""
    if not destinatario or '@' not in str(destinatario):
        return False, 'Sin correo'

    resend.api_key = _resend_api_key()
    if not resend.api_key:
        return False, 'RESEND_API_KEY no configurada.' + _resend_api_key_diagnostico()
    remitente = _env_var('MAIL_FROM', 'EMAIL_FROM', default='onboarding@resend.dev')

    try:
        sexo_val = int(sexo)
    except Exception:
        sexo_val = 0
    trato = 'Estimada' if sexo_val == 2 else 'Estimado'
    periodo_legible = formatear_periodo_texto(periodo)
    pdf_base64 = base64.b64encode(pdf_io.getvalue()).decode('utf-8')

    try:
        params = {
            'from': f'Recursos Humanos <{remitente}>',
            'to': destinatario,
            'subject': f'Certificado Retiro CTS - {periodo_legible} - {nombre_empleado}',
            'html': f"""
                <p>{trato} {nombre_empleado},</p>
                <p>Le hacemos entrega de su certificado de retiro CTS correspondiente al periodo de <b>{periodo_legible}</b>.</p>
                <p>Saludos,<br>Recursos Humanos</p>
            """,
            'attachments': [
                {
                    'content': pdf_base64,
                    'filename': _certificado_retiro_cts_pdf_filename(person or nombre_empleado, periodo),
                }
            ],
        }
        resend.Emails.send(params)
        return True, 'Enviado'
    except Exception as e:
        logging.error('Error en Resend certificado retiro CTS: %s', str(e))
        return False, str(e)


@login_manager.user_loader
def load_user(user_id):
    user = User.get_user_by_id(user_id)
    if user:
        _cache_session_user(user)
        return user
    return _user_from_session_cache(user_id)


@app.route('/')
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('login.html')


@app.route('/login', methods=['POST'])
def login_post():
    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    user = User.validate_user(username, password)
    if user:
        login_user(user)
        _cache_session_user(user)
        ensure_user_session()
        return redirect(url_for('dashboard'))
    flash('Usuario o contraseña incorrectos.', 'error')
    return redirect(url_for('login'))


@app.route('/cambiar-password', methods=['POST'])
def change_password_route():
    username = (request.form.get('username') or '').strip()
    old_password = request.form.get('old_password') or ''
    new_password = request.form.get('new_password') or ''
    confirm = request.form.get('confirm_password') or ''
    if new_password != confirm:
        flash('Las contraseñas nuevas no coinciden.', 'error')
        return redirect(url_for('login'))
    user = User.validate_user(username, old_password)
    if not user:
        flash('Usuario o contraseña anterior incorrectos.', 'error')
        return redirect(url_for('login'))
    ok, msg = cambiar_password(user.id, old_password, new_password)
    flash(msg, 'success' if ok else 'error')
    return redirect(url_for('login'))


@app.route('/logout')
@login_required
def logout():
    session.clear()
    logout_user()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')


@app.route('/trabajadores')
@login_required
def trabajadores_page():
    return render_template('trabajadores.html')


def _selector_items_from_sp(cursor, sp_sql, params):
    """Ejecuta un SP selector y devuelve lista {id, text}."""
    cursor.execute(sp_sql, params)
    col_names = [str(c[0]).strip() for c in (cursor.description or [])]
    items = []
    for row in cursor.fetchall():
        rd = _row_dict_from_columns(col_names, row)
        item_id = rd.get('id') if rd.get('id') is not None else rd.get('bank')
        item_text = rd.get('text') if rd.get('text') is not None else rd.get('name')
        if item_id is None:
            continue
        items.append({'id': str(item_id).strip(), 'text': str(item_text or item_id).strip()})
    return items


def _cargar_selectores_bancario(cursor, cia):
    bancos = _selector_items_from_sp(cursor, 'EXEC sp_pr_selectorbancos_web @cia=?', (cia,))
    formas_pago = _selector_items_from_sp(cursor, 'EXEC sp_pr_selectorformapago_web @cia=?', (cia,))
    tipos_cuenta = _selector_items_from_sp(cursor, 'EXEC sp_pr_selectortipocuenta_web @cia=?', (cia,))
    return bancos, formas_pago, tipos_cuenta


def _cargar_selectores_pensiones(cursor, cia):
    pension_types = _selector_items_from_sp(cursor, 'EXEC sp_pr_selectorpensiontype_web @cia=?', (cia,))
    regime_health = _selector_items_from_sp(cursor, 'EXEC sp_pr_selectorregimehealth_web @cia=?', (cia,))
    return pension_types, regime_health


def _trabajadores_editar_seccion(raw):
    seccion = str(raw or 'bancario').strip().lower()
    return seccion if seccion in ('bancario', 'pensiones') else 'bancario'


def _empleado_pensiones_desde_form(form):
    return {
        'pensiontype': str(form.get('pensiontype') or '').strip(),
        'pensioninscriptiondate': str(form.get('pensioninscriptiondate') or '').strip(),
        'regimehealth': str(form.get('regimehealth') or '').strip(),
        'flagmixta': 'Y' if form.get('flagmixta') == 'Y' else 'N',
        'flagasigfamiliar': 'Y' if form.get('flagasigfamiliar') == 'Y' else 'N',
        'cuspp': str(form.get('cuspp') or '').strip().upper()[:20],
    }


def _empleado_pensiones_para_form(empleado):
    """Fechas en YYYY-MM-DD para <input type=\"date\"> (pyodbc devuelve datetime)."""
    if not empleado:
        return empleado
    out = dict(empleado)
    out['pensioninscriptiondate'] = _sql_date_str_param(out.get('pensioninscriptiondate'))
    return out


@app.route('/trabajadores/editar/<person_id>', methods=['GET', 'POST'])
@login_required
def trabajadores_editar(person_id):
    """Edición de datos bancarios, CTS y pensiones del trabajador."""
    person_id = str(person_id or '').strip()
    cia = str(request.args.get('cia') or request.form.get('cia') or session.get('company') or '').strip()
    seccion = _trabajadores_editar_seccion(request.args.get('seccion') or request.form.get('seccion'))

    if not person_id:
        flash('Trabajador no indicado.', 'warning')
        return redirect(url_for('trabajadores_page'))
    if not cia:
        flash('Indique la compañía (cia) para editar el trabajador.', 'warning')
        return redirect(url_for('trabajadores_page'))

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        if request.method == 'POST' and seccion == 'pensiones':
            datos = _empleado_pensiones_desde_form(request.form)
            cursor.execute(
                'EXEC sp_pr_actualizar_pensiones_trabajador_web '
                '@cia=?, @person=?, @pensiontype=?, @pensioninscriptiondate=?, '
                '@regimehealth=?, @flagmixta=?, @flagasigfamiliar=?, @cuspp=?, @xlastuser=?',
                (
                    cia,
                    person_id,
                    datos['pensiontype'],
                    _sql_date_str_param(datos['pensioninscriptiondate']),
                    datos['regimehealth'],
                    datos['flagmixta'],
                    datos['flagasigfamiliar'],
                    datos['cuspp'],
                    str(getattr(current_user, 'username', '') or '')[:20],
                ),
            )
            conn.commit()
            flash('Datos de pensiones actualizados correctamente.', 'success')
            return redirect(url_for('trabajadores_editar', person_id=person_id, cia=cia, seccion='pensiones'))

        if request.method == 'POST':
            collectionform = str(request.form.get('collectionform') or '').strip()
            cci = re.sub(r'\D', '', str(request.form.get('cci') or ''))[:20]

            forma_es_deposito = (
                '000000000009' in collectionform
                or 'DEPOSITO' in collectionform.upper()
            )
            if not forma_es_deposito and collectionform:
                cursor.execute(
                    "SELECT TOP 1 LTRIM(RTRIM(ISNULL(description, ''))) FROM te_collectionform WHERE collectionform = ? AND company = ?",
                    (collectionform, cia),
                )
                cf_row = cursor.fetchone()
                if cf_row and 'DEPOSITO' in str(cf_row[0] or '').upper():
                    forma_es_deposito = True

            if forma_es_deposito and len(cci) != 20:
                flash('El CCI debe tener 20 dígitos cuando la forma de pago es depósito.', 'danger')
                bancos, formas_pago, tipos_cuenta = _cargar_selectores_bancario(cursor, cia)
                cursor.execute(
                    'EXEC sp_pr_obtener_bancario_trabajador_web @cia=?, @person=?',
                    (cia, person_id),
                )
                rows = _dicts_first_nonempty_resultset(cursor)
                empleado = rows[0] if rows else {}
                empleado['cci'] = cci
                empleado['collectionform'] = collectionform
                empleado['salarybank'] = str(request.form.get('salarybank') or '').strip()
                empleado['salaryaccounttype'] = str(request.form.get('salaryaccounttype') or '').strip()
                empleado['salaryaccount'] = str(request.form.get('salaryaccount') or '').strip()
                empleado['ctsbank'] = str(request.form.get('ctsbank') or '').strip()
                empleado['ctsaccount'] = str(request.form.get('ctsaccount') or '').strip()
                empleado['ctscurrency'] = str(request.form.get('ctscurrency') or 'LO').strip() or 'LO'
                return render_template(
                    'trabajadores_editar.html',
                    cia=cia,
                    person_id=person_id,
                    seccion=seccion,
                    empleado=empleado,
                    bancos=bancos,
                    formas_pago=formas_pago,
                    tipos_cuenta=tipos_cuenta,
                    pension_types=[],
                    regime_health=[],
                )

            cursor.execute(
                'EXEC sp_pr_actualizar_bancario_trabajador_web '
                '@cia=?, @person=?, @collectionform=?, @salarybank=?, @salaryaccounttype=?, '
                '@salaryaccount=?, @cci=?, @ctsbank=?, @ctsaccount=?, @ctscurrency=?, @xlastuser=?',
                (
                    cia,
                    person_id,
                    str(request.form.get('collectionform') or '').strip(),
                    str(request.form.get('salarybank') or '').strip(),
                    str(request.form.get('salaryaccounttype') or '').strip(),
                    str(request.form.get('salaryaccount') or '').strip(),
                    cci,
                    str(request.form.get('ctsbank') or '').strip(),
                    str(request.form.get('ctsaccount') or '').strip(),
                    str(request.form.get('ctscurrency') or 'LO').strip() or 'LO',
                    str(getattr(current_user, 'username', '') or '')[:20],
                ),
            )
            conn.commit()
            flash('Datos bancarios actualizados correctamente.', 'success')
            return redirect(url_for('trabajadores_editar', person_id=person_id, cia=cia, seccion='bancario'))

        if seccion == 'pensiones':
            cursor.execute(
                'EXEC sp_pr_obtener_pensiones_trabajador_web @cia=?, @person=?',
                (cia, person_id),
            )
        else:
            cursor.execute(
                'EXEC sp_pr_obtener_bancario_trabajador_web @cia=?, @person=?',
                (cia, person_id),
            )
        rows = _dicts_first_nonempty_resultset(cursor)
        if not rows:
            flash('No se encontró el trabajador indicado.', 'warning')
            return redirect(url_for('trabajadores_page'))

        empleado = rows[0]
        if seccion == 'pensiones':
            empleado = _empleado_pensiones_para_form(empleado)
        bancos, formas_pago, tipos_cuenta = _cargar_selectores_bancario(cursor, cia)
        pension_types, regime_health = _cargar_selectores_pensiones(cursor, cia)

        return render_template(
            'trabajadores_editar.html',
            cia=cia,
            person_id=person_id,
            seccion=seccion,
            empleado=empleado,
            bancos=bancos,
            formas_pago=formas_pago,
            tipos_cuenta=tipos_cuenta,
            pension_types=pension_types,
            regime_health=regime_health,
        )
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        logging.exception('trabajadores_editar')
        flash(f'Error al procesar la solicitud: {e}', 'danger')
        return redirect(url_for('trabajadores_page'))
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/reporte-liquidaciones')
@login_required
def reporte_liquidaciones():
    # La página carga vacía; los filtros se llenan por JS vía APIs.
    return render_template('reporte_liquidaciones.html')


@app.route('/reporte-planilla-vertical')
@login_required
def reporte_planilla_vertical_page():
    return render_template('reporte_planilla_vertical.html')


@app.route('/reporte-vacaciones-detalle')
@login_required
def reporte_vacaciones_detalle_page():
    return render_template('reporte_vacaciones_detalle.html')


@app.route('/reporte-saldo-vacaciones')
@login_required
def reporte_saldo_vacaciones_page():
    return render_template('reporte_saldo_vacaciones.html')


@app.route('/reporte-descansos-medicos-detalle')
@login_required
def reporte_descansos_medicos_detalle_page():
    return render_template('reporte_descansos_medicos_detalle.html')


@app.route('/reporte-log-calculo')
@login_required
def reporte_log_calculo_page():
    return render_template('reporte_log_calculo.html')


@app.route('/reporte-listado-pagos')
@login_required
def reporte_listado_pagos_page():
    return render_template('reporte_listado_pagos.html')


@app.route('/procesar_planilla')
@login_required
def procesar_planilla_page():
    return render_template('procesar_planilla.html')


@app.route('/aperturar-periodos')
@login_required
def aperturar_periodos_page():
    return render_template('aperturar_periodos.html')


@app.route('/asignacion-conceptos')
@login_required
def asignacion_conceptos_page():
    return render_template('asignacion_conceptos.html')


@app.route('/registro-vacaciones')
@login_required
def registro_vacaciones_page():
    return render_template('registro_vacaciones.html')


@app.route('/conceptos')
@login_required
def conceptos_page():
    return render_template('maestro_conceptos.html')


def _concepto_lista_dict(r):
    return {
        'concept': _jsonable_value(r.get('concept')),
        'description': _jsonable_value(r.get('description')),
        'pdt': _jsonable_value(r.get('pdt')),
        'tiposhortname': _jsonable_value(r.get('tiposhortname')),
        'tipodescription': _jsonable_value(r.get('tipodescription')),
        'formulacode': _jsonable_value(r.get('formulacode')),
        'reporden': _jsonable_value(r.get('reporden')),
        'xlastdate': _jsonable_datetime(r.get('xlastdate')),
        'puede_eliminar': _jsonable_value(r.get('puede_eliminar')) or 'N',
    }


def _concepto_detalle_dict(r):
    if not r:
        return None
    return {
        'concept': _jsonable_value(r.get('concept')),
        'company': _jsonable_value(r.get('company')),
        'description': _jsonable_value(r.get('description')),
        'printtext': _jsonable_value(r.get('printtext')),
        'formulacode': _jsonable_value(r.get('formulacode')),
        'concepttype': _jsonable_value(r.get('concepttype')),
        'concepttypename': _jsonable_value(r.get('concepttypename')),
        'conceptcurrency': _jsonable_value(r.get('conceptcurrency')) or 'LO',
        'flagismonetary': _jsonable_value(r.get('flagismonetary')) or 'N',
        'flagassign': _jsonable_value(r.get('flagassign')) or 'N',
        'conceptorder': _jsonable_value(r.get('conceptorder')),
        'flagpayrollticket': _jsonable_value(r.get('flagpayrollticket')) or 'N',
        'reporden': _jsonable_value(r.get('reporden')),
        'flagconceptdeclare': _jsonable_value(r.get('flagconceptdeclare')) or 'N',
        'pdt': _jsonable_value(r.get('pdt')),
        'flagcontract': _jsonable_value(r.get('flagcontract')) or 'N',
        'status': _jsonable_value(r.get('status')) or 'A',
        'flaginsertar': _jsonable_value(r.get('flaginsertar')) or 'N',
        'flagafecto5ta': _jsonable_value(r.get('flagafecto5ta')) or 'N',
        'flagafectoafp': _jsonable_value(r.get('flagafectoafp')) or 'N',
        'xlastuser': _jsonable_value(r.get('xlastuser')),
        'xlastdate': _jsonable_datetime(r.get('xlastdate')),
    }


def _normalize_flag_yn(value, default='N'):
    s = str(value or default).strip().upper()
    return s if s in ('Y', 'N') else default


@app.route('/api/conceptos/listado', methods=['POST'])
@login_required
def api_conceptos_listado():
    """sp_pr_listarconceptos_web: listado maestro de conceptos."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    descripcion = str(body.get('descripcion') or body.get('busqueda') or body.get('q') or '').strip()

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listarconceptos_web @company=?, @descripcion=?",
            (cia, descripcion or None),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = [_concepto_lista_dict(r) for r in rows]
        return jsonify({"rows": resultado, "total": len(resultado)})
    except Exception as e:
        logging.exception("api_conceptos_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/conceptos/obtener', methods=['POST'])
@login_required
def api_conceptos_obtener():
    """sp_pr_obtenerconcepto_web: detalle para edición."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    concept = str(body.get('concept') or '').strip()

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not concept:
        return jsonify({"error": "Seleccione un concepto."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_obtenerconcepto_web @company=?, @concept=?",
            (cia, concept),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        detalle = _concepto_detalle_dict(rows[0] if rows else None)
        if not detalle:
            return jsonify({"error": "Concepto no encontrado."}), 404
        return jsonify(detalle)
    except Exception as e:
        logging.exception("api_conceptos_obtener")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/conceptos/guardar', methods=['POST'])
@login_required
def api_conceptos_guardar():
    """sp_pr_guardarconcepto_web: alta / edición de concepto."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    concept = str(body.get('concept') or '').strip()
    modo = str(body.get('modo') or ('U' if concept else 'I')).strip().upper()

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400

    description = str(body.get('description') or '').strip()
    formulacode = str(body.get('formulacode') or '').strip()
    concepttype = str(body.get('concepttype') or '').strip()
    if not description or not formulacode or not concepttype:
        return jsonify({"error": "Complete concepto, nemónico y tipo."}), 400

    conceptorder = body.get('conceptorder')
    reporden = body.get('reporden')
    try:
        conceptorder = int(conceptorder) if conceptorder not in (None, '') else None
    except Exception:
        conceptorder = None
    try:
        reporden = int(reporden) if reporden not in (None, '') else None
    except Exception:
        reporden = None

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_guardarconcepto_web "
            "@modo=?, @company=?, @concept=?, @description=?, @formulacode=?, @concepttype=?, "
            "@conceptcurrency=?, @flagismonetary=?, @printtext=?, @conceptorder=?, @status=?, "
            "@flagassign=?, @flagpayrollticket=?, @flagcontract=?, @pdt=?, @flagconceptdeclare=?, "
            "@reporden=?, @flaginsertar=?, @flagafectoafp=?, @flagafecto5ta=?, @xlastuser=?",
            (
                modo,
                cia,
                concept or None,
                description,
                formulacode,
                concepttype,
                str(body.get('conceptcurrency') or 'LO').strip().upper()[:2],
                _normalize_flag_yn(body.get('flagismonetary'), 'Y'),
                str(body.get('printtext') or description).strip(),
                conceptorder,
                str(body.get('status') or 'A').strip().upper()[:1],
                _normalize_flag_yn(body.get('flagassign')),
                _normalize_flag_yn(body.get('flagpayrollticket')),
                _normalize_flag_yn(body.get('flagcontract')),
                str(body.get('pdt') or '').strip() or None,
                _normalize_flag_yn(body.get('flagconceptdeclare')),
                reporden,
                str(body.get('flaginsertar') or 'N').strip().upper()[:1] or 'N',
                _normalize_flag_yn(body.get('flagafectoafp')),
                _normalize_flag_yn(body.get('flagafecto5ta')),
                _xlastuser_id(),
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        conn.commit()
        row = rows[0] if rows else {}
        concept_id = str(row.get('concept') or concept or '').strip()
        return jsonify({
            "ok": True,
            "concept": concept_id,
            "modo": _jsonable_value(row.get('modo')) or modo,
            "mensaje": _jsonable_value(row.get('mensaje')) or 'Guardado correctamente.',
        })
    except Exception as e:
        logging.exception("api_conceptos_guardar")
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/conceptos/eliminar', methods=['POST'])
@login_required
def api_conceptos_eliminar():
    """sp_pr_eliminarconcepto_web: elimina concepto si no está en uso."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    concept = str(body.get('concept') or '').strip()

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not concept:
        return jsonify({"error": "Seleccione un concepto."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_eliminarconcepto_web @company=?, @concept=?",
            (cia, concept),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        conn.commit()
        row = rows[0] if rows else {}
        return jsonify({
            "ok": True,
            "concept": _jsonable_value(row.get('concept')) or concept,
            "mensaje": _jsonable_value(row.get('mensaje')) or 'Concepto eliminado correctamente.',
        })
    except Exception as e:
        logging.exception("api_conceptos_eliminar")
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        err = str(e)
        if 'RAISERROR' in err or '50000' in err:
            parts = err.split(']')
            if len(parts) > 1:
                err = parts[-1].strip(" ()'\"")
        return jsonify({"error": err}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/concept-types')
@login_required
def api_concept_types():
    """sp_pr_selectorconcepttype_web: tipos de concepto."""
    cia = str(request.args.get('cia') or request.args.get('company') or '').strip() or None
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectorconcepttype_web @cia=?", (cia,))
        rows = _dicts_first_nonempty_resultset(cursor)
        return jsonify([
            {
                "id": _jsonable_value(r.get('id')),
                "text": _jsonable_value(r.get('text')),
                "shortname": _jsonable_value(r.get('shortname')),
            }
            for r in rows
        ])
    except Exception as e:
        logging.exception("api_concept_types")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/plame/archivo-14')
@login_required
def plame_archivo14_page():
    return render_template('plame_archivo14.html')


@app.route('/plame/archivo-15')
@login_required
def plame_archivo15_page():
    return render_template('plame_archivo15.html')


@app.route('/plame/archivo-18')
@login_required
def plame_archivo18_page():
    return render_template('plame_archivo18.html')


@app.route('/afp/declaracion')
@login_required
def declaracion_afp_page():
    return render_template('declaracion_afp.html')


@app.route('/afp/control-pagos')
@login_required
def control_pagos_afp_page():
    return render_template('reporte_control_pagos_afp.html')


@app.route('/reporte_control_pagos_afp', methods=['POST'])
@login_required
def reporte_control_pagos_afp_post():
    """sp_pr_control_pagos_afp_web: resumen de pagos AFP por planilla y AFP."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or body.get('payrolltype') or '').strip()
    period_raw = str(body.get('period') or '').strip().replace('-', '').replace('/', '')
    period = period_raw[:6] if len(period_raw) >= 6 else period_raw

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not payroll_type:
        return jsonify({"error": "Debe indicar tipo de planilla."}), 400
    if not period or len(period) != 6 or not period.isdigit():
        return jsonify({"error": "Debe indicar un periodo válido (YYYYMM)."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_control_pagos_afp_web @company=?, @payrolltype=?, @period=?",
            (cia, payroll_type, period),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            fixed = _float_sp_cell(r.get('fixedamountlo'))
            variable = _float_sp_cell(r.get('variableamountlo'))
            insured = _float_sp_cell(r.get('insuredamountlo'))
            comision = _float_sp_cell(r.get('arcomisionamountlo'))
            aporte = round(fixed + variable, 2)
            retenciones = round(insured + comision, 2)
            total = round(aporte + retenciones, 2)
            cantidad_raw = r.get('cantidad')
            try:
                cantidad = int(cantidad_raw) if cantidad_raw is not None else 0
            except Exception:
                cantidad = 0
            resultado.append({
                "tipoplanilla": '' if r.get('tipoplanilla') is None else str(r.get('tipoplanilla')).strip(),
                "afpname": '' if r.get('afpname') is None else str(r.get('afpname')).strip(),
                "cantidad": cantidad,
                "fixedamountlo": fixed,
                "insuredamountlo": insured,
                "employercontributionlo": _float_sp_cell(r.get('employercontributionlo')),
                "variableamountlo": variable,
                "arcomisionamountlo": comision,
                "aporte": aporte,
                "retenciones": retenciones,
                "total": total,
            })
        return jsonify({
            "rows": resultado,
            "meta": {
                "period": period,
                "period_label": f"{period[:4]}-{period[4:6]}",
            },
        })
    except Exception as e:
        logging.exception("reporte_control_pagos_afp_post")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/plame/archivo-26')
@login_required
def plame_archivo26_page():
    return render_template('plame_archivo26.html')


@app.route('/plame/validar')
@login_required
def plame_validar_page():
    return render_template('plame_validar.html')


def _plame_sunat_obtener_carga(cursor, cia, period):
    cursor.execute(
        "EXEC sp_pr_plame_sunat_obtener_carga_web @cia=?, @period=?",
        (cia, period),
    )
    rows = _dicts_first_nonempty_resultset(cursor)
    return rows[0] if rows else None


def _plame_sunat_carga_a_json(row):
    if not row:
        return None
    uploaded_at = row.get('uploadedat')
    if isinstance(uploaded_at, datetime):
        uploaded_at = uploaded_at.strftime('%Y-%m-%d %H:%M:%S')
    return {
        'cargaid': row.get('cargaid'),
        'company': row.get('company'),
        'period': row.get('period'),
        'ruc': row.get('ruc'),
        'employername': row.get('employername'),
        'periodosunat': row.get('periodosunat'),
        'filer01name': row.get('filer01name'),
        'filer04name': row.get('filer04name'),
        'filer05name': row.get('filer05name'),
        'rowsr01': int(row.get('rowsr01') or 0),
        'rowsr04': int(row.get('rowsr04') or 0),
        'rowsr05': int(row.get('rowsr05') or 0),
        'uploadedat': uploaded_at,
        'uploadedby': row.get('uploadedby'),
    }


def _plame_sunat_validar_archivos(parsed_bundle, filenames, cia, period):
    metas = [parsed_bundle[t]['meta'] for t in ARCHIVOS_SUNAT]
    rucs = {m.get('ruc') for m in metas if m.get('ruc')}
    if len(rucs) != 1:
        return 'Los tres XML deben corresponder al mismo RUC.'
    periodos = {m.get('period_yyyymm') for m in metas if m.get('period_yyyymm')}
    if len(periodos) != 1:
        return 'Los tres XML deben corresponder al mismo periodo tributario.'
    periodo_xml = next(iter(periodos))
    if periodo_xml != period:
        return (
            f'El periodo del XML ({periodo_xml}) no coincide con el seleccionado ({period}).'
        )
    for tipo in ARCHIVOS_SUNAT:
        fn_info = parse_filename(filenames.get(tipo))
        if fn_info and fn_info.get('period') != period:
            return f'El nombre de {tipo} no coincide con el periodo seleccionado.'
        if fn_info and fn_info.get('tipo') != tipo:
            return f'El archivo {filenames.get(tipo)} no corresponde a {tipo}.'
    return None


def _plame_sunat_guardar_carga(conn, cia, period, parsed_bundle, filenames, username):
    meta = parsed_bundle['R01']['meta']
    cursor = conn.cursor()
    cursor.execute(
        "EXEC sp_pr_plame_sunat_eliminar_carga_web @cia=?, @period=?",
        (cia, period),
    )
    _drain_pyodbc_cursor(cursor)
    cursor.execute(
        """
        INSERT INTO PR_PlameSunatCarga (
            Company, Period, Ruc, EmployerName, PeriodoSunat,
            FileR01Name, FileR04Name, FileR05Name,
            RowsR01, RowsR04, RowsR05, UploadedBy
        )
        OUTPUT INSERTED.CargaId
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            cia,
            period,
            meta.get('ruc') or '',
            (meta.get('employername') or '')[:300],
            (meta.get('periodosunat') or '')[:7],
            (filenames.get('R01') or '')[:260],
            (filenames.get('R04') or '')[:260],
            (filenames.get('R05') or '')[:260],
            parsed_bundle['R01']['rows_count'],
            parsed_bundle['R04']['rows_count'],
            parsed_bundle['R05']['rows_count'],
            (username or '')[:50],
        ),
    )
    row = cursor.fetchone()
    if not row or row[0] is None:
        _drain_pyodbc_cursor(cursor)
        row = _plame_sunat_obtener_carga(cursor, cia, period)
        carga_id = int(row.get('cargaid') or 0) if row else 0
    else:
        carga_id = int(row[0])
    if not carga_id:
        raise RuntimeError('No se pudo registrar la cabecera de la carga SUNAT.')

    insert_sql = """
        INSERT INTO PR_PlameSunatFila (
            CargaId, Archivo, NumFila, TipoDoc, DocumentNumber,
            LastName1, LastName2, Names, Situacion, MontosJson
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    params = []
    for tipo in ARCHIVOS_SUNAT:
        for fila in parsed_bundle[tipo]['filas']:
            params.append((
                carga_id,
                fila['archivo'],
                fila['num_fila'],
                fila['tipodoc'],
                fila['documentnumber'],
                fila['lastname1'][:100],
                fila['lastname2'][:100],
                fila['names'][:200],
                fila['situacion'][:80],
                fila['montos_json'],
            ))
    batch_size = 50
    for i in range(0, len(params), batch_size):
        cursor.executemany(insert_sql, params[i:i + batch_size])
    conn.commit()
    return carga_id


@app.route('/api/plame/validar/estado', methods=['GET'])
@login_required
def api_plame_validar_estado():
    """Estado de la carga SUNAT R01/R04/R05 para compañía y periodo."""
    cia = str(request.args.get('cia') or '').strip()
    period = _plame_period_yyyymm(request.args.get('period'))
    if not cia or len(period) != 6:
        return jsonify({'carga': None})
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        row = _plame_sunat_obtener_carga(cursor, cia, period)
        return jsonify({'carga': _plame_sunat_carga_a_json(row)})
    except Exception as e:
        logging.exception('api_plame_validar_estado')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/validar/carga', methods=['POST'])
@login_required
def api_plame_validar_carga():
    """Sube y parsea los XML SUNAT R01, R04 y R05 para un periodo."""
    cia = str(request.form.get('cia') or '').strip()
    period = _plame_period_yyyymm(request.form.get('period'))
    err = _plame_validar_params({'cia': cia, 'period': period})
    if err:
        return jsonify({'error': err}), 400

    file_map = {
        'R01': request.files.get('file_r01'),
        'R04': request.files.get('file_r04'),
        'R05': request.files.get('file_r05'),
    }
    for tipo in ARCHIVOS_SUNAT:
        f = file_map.get(tipo)
        if not f or not f.filename:
            return jsonify({'error': f'Debe seleccionar el archivo {tipo}.'}), 400

    parsed_bundle = {}
    filenames = {}
    try:
        for tipo in ARCHIVOS_SUNAT:
            f = file_map[tipo]
            filenames[tipo] = f.filename
            raw = f.read()
            if not raw:
                return jsonify({'error': f'El archivo {tipo} está vacío.'}), 400
            parsed_bundle[tipo] = parse_sunat_xml(raw, tipo)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except ET.ParseError:
        return jsonify({'error': 'Uno de los archivos no es un XML válido.'}), 400
    except Exception as e:
        logging.exception('api_plame_validar_carga parse')
        return jsonify({'error': f'Error al leer XML: {e}'}), 400

    err = _plame_sunat_validar_archivos(parsed_bundle, filenames, cia, period)
    if err:
        return jsonify({'error': err}), 400

    conn = None
    try:
        conn = get_db_connection()
        username = str(getattr(current_user, 'username', '') or getattr(current_user, 'id', ''))
        carga_id = _plame_sunat_guardar_carga(
            conn, cia, period, parsed_bundle, filenames, username,
        )
        cursor = conn.cursor()
        row = _plame_sunat_obtener_carga(cursor, cia, period)
        return jsonify({
            'ok': True,
            'cargaid': carga_id,
            'carga': _plame_sunat_carga_a_json(row),
            'mensaje': 'Archivos SUNAT cargados correctamente.',
            'ejecutar_validacion': True,
        })
    except Exception as e:
        logging.exception('api_plame_validar_carga')
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


def _plame_validar_neto_r01_ejecutar(cursor, cia, period):
    cursor.execute(
        "EXEC sp_pr_plame_validar_neto_r01_web @cia=?, @period=?",
        (cia, period),
    )
    sets = _dicts_collect_nonempty_resultsets(cursor)
    resumen = sets[0][0] if len(sets) > 0 and sets[0] else {}
    filas = sets[1] if len(sets) > 1 else []
    return resumen, filas


def _plame_validar_neto_r01_json(resumen, filas):
    def num(val):
        try:
            return round(float(val or 0), 2)
        except (TypeError, ValueError):
            return 0.0

    def int_val(val):
        try:
            return int(val or 0)
        except (TypeError, ValueError):
            return 0

    res = resumen or {}
    out_filas = []
    for r in filas or []:
        out_filas.append({
            'tipodoc': _jsonable_value(r.get('tipodoc')),
            'documentnumber': _jsonable_value(r.get('documentnumber')),
            'nombre': _jsonable_value(r.get('nombre')),
            'neto_sunat': num(r.get('neto_sunat')),
            'neto_planilla': num(r.get('neto_planilla')),
            'diferencia': num(r.get('diferencia')),
            'estado': _jsonable_value(r.get('estado')),
        })
    tiene_diferencias = any(
        f.get('estado') != 'OK' for f in out_filas
    )
    filas_detalle = [
        f for f in out_filas
        if f.get('estado') in ('DIFERENCIA', 'SOLO_SUNAT')
    ]
    return {
        'resumen': {
            'total_filas': int_val(res.get('total_filas')),
            'coinciden': int_val(res.get('coinciden')),
            'con_diferencia': int_val(res.get('con_diferencia')),
            'solo_sunat': int_val(res.get('solo_sunat')),
            'solo_planilla': int_val(res.get('solo_planilla')),
            'total_neto_sunat': num(res.get('total_neto_sunat')),
            'total_neto_planilla': num(res.get('total_neto_planilla')),
            'total_diferencia': num(res.get('total_diferencia')),
        },
        'filas': filas_detalle,
        'tiene_diferencias': tiene_diferencias,
    }


def _plame_validar_r04_ejecutar(cursor, cia, period):
    cursor.execute(
        'EXEC sp_pr_plame_validar_r04_web '
        '@cia=?, @period=?, @payroll_all=?, @payroll=?, @cesados=?',
        (cia, period, 'Y', None, 'T'),
    )
    sets = _dicts_collect_nonempty_resultsets(cursor)
    resumen = sets[0] if len(sets) > 0 else []
    filas = sets[1] if len(sets) > 1 else []
    return resumen, filas


def _plame_validar_r04_json(resumen_filas, detalle_filas):
    def num(val):
        try:
            return round(float(val or 0), 2)
        except (TypeError, ValueError):
            return 0.0

    def int_val(val):
        try:
            return int(val or 0)
        except (TypeError, ValueError):
            return 0

    conceptos = []
    for r in resumen_filas or []:
        conceptos.append({
            'concepto': _jsonable_value(r.get('concepto')),
            'concepto_nombre': _jsonable_value(r.get('concepto_nombre')),
            'total_filas': int_val(r.get('total_filas')),
            'coinciden': int_val(r.get('coinciden')),
            'con_diferencia': int_val(r.get('con_diferencia')),
            'solo_sunat': int_val(r.get('solo_sunat')),
            'solo_planilla': int_val(r.get('solo_planilla')),
            'total_sunat': num(r.get('total_sunat')),
            'total_planilla': num(r.get('total_planilla')),
            'total_diferencia': num(r.get('total_diferencia')),
        })

    out_filas = []
    for r in detalle_filas or []:
        estado = str(r.get('estado') or '').strip()
        if estado not in ('DIFERENCIA', 'SOLO_SUNAT'):
            continue
        out_filas.append({
            'concepto': _jsonable_value(r.get('concepto')),
            'concepto_nombre': _jsonable_value(r.get('concepto_nombre')),
            'documentnumber': _jsonable_value(r.get('documentnumber')),
            'nombre': _jsonable_value(r.get('nombre')),
            'monto_sunat': num(r.get('monto_sunat')),
            'monto_planilla': num(r.get('monto_planilla')),
            'diferencia': num(r.get('diferencia')),
            'estado': estado,
        })

    tiene_diferencias = any(
        c.get('con_diferencia', 0) > 0
        or c.get('solo_sunat', 0) > 0
        or c.get('solo_planilla', 0) > 0
        for c in conceptos
    )
    return {
        'conceptos': conceptos,
        'filas': out_filas,
        'tiene_diferencias': tiene_diferencias,
    }


def _plame_validar_r05_ejecutar(cursor, cia, period):
    cursor.execute(
        'EXEC sp_pr_plame_validar_r05_web '
        '@cia=?, @period=?, @payroll_all=?, @payroll=?, @cesados=?',
        (cia, period, 'Y', None, 'T'),
    )
    sets = _dicts_collect_nonempty_resultsets(cursor)
    resumen = sets[0][0] if len(sets) > 0 and sets[0] else {}
    filas = sets[1] if len(sets) > 1 else []
    return resumen, filas


def _plame_validar_r05_json(resumen, filas):
    def num(val):
        try:
            return round(float(val or 0), 2)
        except (TypeError, ValueError):
            return 0.0

    def int_val(val):
        try:
            return int(val or 0)
        except (TypeError, ValueError):
            return 0

    res = resumen or {}
    out_filas = []
    for r in filas or []:
        out_filas.append({
            'tipodoc': _jsonable_value(r.get('tipodoc')),
            'documentnumber': _jsonable_value(r.get('documentnumber')),
            'nombre': _jsonable_value(r.get('nombre')),
            'essalud_sunat': num(r.get('essalud_sunat')),
            'essalud_planilla': num(r.get('essalud_planilla')),
            'diferencia': num(r.get('diferencia')),
            'estado': _jsonable_value(r.get('estado')),
        })
    tiene_diferencias = any(
        f.get('estado') != 'OK' for f in out_filas
    )
    filas_detalle = [
        f for f in out_filas
        if f.get('estado') in ('DIFERENCIA', 'SOLO_SUNAT')
    ]
    return {
        'resumen': {
            'total_filas': int_val(res.get('total_filas')),
            'coinciden': int_val(res.get('coinciden')),
            'con_diferencia': int_val(res.get('con_diferencia')),
            'solo_sunat': int_val(res.get('solo_sunat')),
            'solo_planilla': int_val(res.get('solo_planilla')),
            'total_essalud_sunat': num(res.get('total_essalud_sunat')),
            'total_essalud_planilla': num(res.get('total_essalud_planilla')),
            'total_diferencia': num(res.get('total_diferencia')),
        },
        'filas': filas_detalle,
        'tiene_diferencias': tiene_diferencias,
    }


@app.route('/api/plame/validar/r05', methods=['POST'])
@login_required
def api_plame_validar_r05():
    """Compara ESSALUD Seguro de Salud (R05 SUNAT) vs planilla (FormulaCode ESSALUD)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({'error': err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        if not _plame_sunat_obtener_carga(cursor, p['cia'], p['period']):
            return jsonify({
                'error': 'No hay carga SUNAT para este periodo. Suba los archivos R01, R04 y R05 primero.',
            }), 400
        _drain_pyodbc_cursor(cursor)
        resumen, filas = _plame_validar_r05_ejecutar(cursor, p['cia'], p['period'])
        return jsonify(_plame_validar_r05_json(resumen, filas))
    except Exception as e:
        logging.exception('api_plame_validar_r05')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/validar/r04', methods=['POST'])
@login_required
def api_plame_validar_r04():
    """Compara tributos R04 (AFP, ONP, 5ta) vs planilla por FormulaCode."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({'error': err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        if not _plame_sunat_obtener_carga(cursor, p['cia'], p['period']):
            return jsonify({
                'error': 'No hay carga SUNAT para este periodo. Suba los archivos R01, R04 y R05 primero.',
            }), 400
        _drain_pyodbc_cursor(cursor)
        resumen, filas = _plame_validar_r04_ejecutar(cursor, p['cia'], p['period'])
        return jsonify(_plame_validar_r04_json(resumen, filas))
    except Exception as e:
        logging.exception('api_plame_validar_r04')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/validar/neto-r01', methods=['POST'])
@login_required
def api_plame_validar_neto_r01():
    """Compara Neto a pagar (R01 SUNAT) vs Neto planilla (FormulaCode NETO)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({'error': err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        if not _plame_sunat_obtener_carga(cursor, p['cia'], p['period']):
            return jsonify({
                'error': 'No hay carga SUNAT para este periodo. Suba los archivos R01, R04 y R05 primero.',
            }), 400
        _drain_pyodbc_cursor(cursor)
        resumen, filas = _plame_validar_neto_r01_ejecutar(cursor, p['cia'], p['period'])
        resultado = _plame_validar_neto_r01_json(resumen, filas)
        return jsonify(resultado)
    except Exception as e:
        logging.exception('api_plame_validar_neto_r01')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/archivo-14/listado', methods=['POST'])
@login_required
def api_plame_archivo14_listado():
    """sp_pr_listado_plame14_web: preview jornada laboral / sobretiempo PLAME."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listado_plame14_web @cia=?, @period=?",
            (p['cia'], p['period']),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        validaciones, personas_incidencia = _plame_validar_archivo14_incidencias(cursor, p)
        resultado = []
        for r in rows:
            wh = r.get('workinghours')
            eh = r.get('extrahours')
            try:
                wh_num = float(wh) if wh is not None else 0.0
            except Exception:
                wh_num = 0.0
            try:
                eh_num = float(eh) if eh is not None else 0.0
            except Exception:
                eh_num = 0.0
            person = _jsonable_value(r.get('person'))
            person_key = str(person or '').strip()
            sin_horas = wh_num <= 0
            incidencia = sin_horas or person_key in personas_incidencia
            resultado.append({
                "person": person,
                "documenttype": _jsonable_value(r.get('documenttype')),
                "documentnumber": _jsonable_value(r.get('documentnumber')),
                "name": _jsonable_value(r.get('name')),
                "workinghours": wh_num,
                "workingminutes": _jsonable_value(r.get('workingminutes')),
                "extrahours": eh_num,
                "extraminutes": _jsonable_value(r.get('extraminutes')),
                "selection": _jsonable_value(r.get('selection')),
                "incidencia": incidencia,
            })
        return jsonify({
            "rows": resultado,
            "total": len(resultado),
            "validaciones": validaciones,
        })
    except Exception as e:
        logging.exception("api_plame_archivo14_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/archivo-14/generar-txt', methods=['POST'])
@login_required
def api_plame_archivo14_generar_txt():
    """Genera TXT PLAME Archivo 14 (encoding latin-1)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    filas = _plame_rows_from_json(body)
    if not filas:
        return jsonify({"error": "Seleccione al menos un registro."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        ruc = _obtener_ruc_compania(cursor, p['cia'])
        if not ruc:
            return jsonify({"error": "No se encontró el RUC de la compañía en SY_Company."}), 400

        lineas = [_plame_linea_archivo14(row) for row in filas]
        contenido = '\r\n'.join(lineas)
        if lineas:
            contenido += '\r\n'

        filename = _plame_filename(ruc, p['period'], '14')
        resp = Response(
            contenido.encode('latin-1', errors='replace'),
            mimetype='text/plain; charset=iso-8859-1',
        )
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        logging.exception("api_plame_archivo14_generar_txt")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/archivo-15/listado', methods=['POST'])
@login_required
def api_plame_archivo15_listado():
    """sp_pr_listado_plame15_web: días subsidiados y no laborados (.snl)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listado_plame15_web @cia=?, @period=?",
            (p['cia'], p['period']),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            try:
                days_num = int(float(r.get('days') or 0))
            except (TypeError, ValueError):
                days_num = 0
            resultado.append({
                "person": _jsonable_value(r.get('person')),
                "documenttype": _jsonable_value(r.get('documenttype')),
                "documentnumber": _jsonable_value(r.get('documentnumber')),
                "name": _jsonable_value(r.get('name')),
                "suspensiontype": _jsonable_value(r.get('suspensiontype')),
                "suspensionname": _jsonable_value(r.get('suspensionname')),
                "days": days_num,
                "selection": _jsonable_value(r.get('selection')),
            })
        return jsonify({"rows": resultado, "total": len(resultado)})
    except Exception as e:
        logging.exception("api_plame_archivo15_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/archivo-15/generar-txt', methods=['POST'])
@login_required
def api_plame_archivo15_generar_txt():
    """Genera TXT PLAME Archivo 15 (.snl, encoding latin-1)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    filas = _plame_rows_archivo15_from_json(body)
    if not filas:
        return jsonify({"error": "Seleccione al menos un registro."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        ruc = _obtener_ruc_compania(cursor, p['cia'])
        if not ruc:
            return jsonify({"error": "No se encontró el RUC de la compañía en SY_Company."}), 400

        lineas = [_plame_linea_archivo15(row) for row in filas]
        contenido = '\r\n'.join(lineas)
        if lineas:
            contenido += '\r\n'

        filename = _plame_filename(ruc, p['period'], '15')
        resp = Response(
            contenido.encode('latin-1', errors='replace'),
            mimetype='text/plain; charset=iso-8859-1',
        )
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        logging.exception("api_plame_archivo15_generar_txt")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/archivo-18/listado', methods=['POST'])
@login_required
def api_plame_archivo18_listado():
    """sp_pr_listado_plame18_web: ingresos, tributos y descuentos (.rem)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_archivo18_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listado_plame18_web @cia=?, @period=?, @payroll_all=?, @payroll=?, @cesados=?",
            (p['cia'], p['period'], p['payroll_all'], p['payroll'] or None, p['cesados']),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        validaciones, personas_incidencia = _plame_validar_archivo18_incidencias(cursor, p)
        resultado = []
        for r in rows:
            try:
                cv = float(r.get('conceptvalue') or 0)
            except (TypeError, ValueError):
                cv = 0.0
            try:
                cl = float(r.get('conceptvaluelo') or 0)
            except (TypeError, ValueError):
                cl = 0.0
            person = _jsonable_value(r.get('person'))
            person_key = str(person or '').strip()
            incidencia = person_key in personas_incidencia
            resultado.append({
                "person": person,
                "documenttype": _jsonable_value(r.get('documenttype')),
                "documentnumber": _jsonable_value(r.get('documentnumber')),
                "name": _jsonable_value(r.get('name')),
                "pdt": _jsonable_value(r.get('pdt')),
                "conceptvalue": cv,
                "conceptvaluelo": cl,
                "selection": _jsonable_value(r.get('selection')),
                "incidencia": incidencia,
            })
        return jsonify({
            "rows": resultado,
            "total": len(resultado),
            "validaciones": validaciones,
        })
    except Exception as e:
        logging.exception("api_plame_archivo18_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/archivo-18/generar-txt', methods=['POST'])
@login_required
def api_plame_archivo18_generar_txt():
    """Genera TXT PLAME Archivo 18 (.rem, encoding latin-1)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_archivo18_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    filas = _plame_rows_archivo18_from_json(body)
    if not filas:
        return jsonify({"error": "Seleccione al menos un registro."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        ruc = _obtener_ruc_compania(cursor, p['cia'])
        if not ruc:
            return jsonify({"error": "No se encontró el RUC de la compañía en SY_Company."}), 400

        lineas = [_plame_linea_archivo18(row) for row in filas]
        contenido = '\r\n'.join(lineas)
        if lineas:
            contenido += '\r\n'

        filename = _plame_filename(ruc, p['period'], '18')
        resp = Response(
            contenido.encode('latin-1', errors='replace'),
            mimetype='text/plain; charset=iso-8859-1',
        )
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        logging.exception("api_plame_archivo18_generar_txt")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/archivo-26/listado', methods=['POST'])
@login_required
def api_plame_archivo26_listado():
    """sp_pr_listado_plame26_web: indicador aporte Asegura tu pensión (.toc)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listado_plame26_web @cia=?, @period=?",
            (p['cia'], p['period']),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            resultado.append({
                "person": _jsonable_value(r.get('person')),
                "documenttype": _jsonable_value(r.get('documenttype')),
                "documentnumber": _jsonable_value(r.get('documentnumber')),
                "name": _jsonable_value(r.get('name')),
                "pensionmembership": _jsonable_value(r.get('pensionmembership')),
                "accidentinsurance": _jsonable_value(r.get('accidentinsurance')),
                "typeaporte": _jsonable_value(r.get('typeaporte')),
                "isdomiciled": _jsonable_value(r.get('isdomiciled')),
                "selection": _jsonable_value(r.get('selection')),
            })
        return jsonify({"rows": resultado, "total": len(resultado)})
    except Exception as e:
        logging.exception("api_plame_archivo26_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/plame/archivo-26/generar-txt', methods=['POST'])
@login_required
def api_plame_archivo26_generar_txt():
    """Genera TXT PLAME Archivo 26 (.toc, encoding latin-1)."""
    body = request.get_json(silent=True) or {}
    p = _plame_params_from_json(body)
    err = _plame_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    filas = _plame_rows_archivo26_from_json(body)
    if not filas:
        return jsonify({"error": "Seleccione al menos un registro."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        ruc = _obtener_ruc_compania(cursor, p['cia'])
        if not ruc:
            return jsonify({"error": "No se encontró el RUC de la compañía en SY_Company."}), 400

        lineas = [_plame_linea_archivo26(row) for row in filas]
        contenido = '\r\n'.join(lineas)
        if lineas:
            contenido += '\r\n'

        filename = _plame_filename(ruc, p['period'], '26')
        resp = Response(
            contenido.encode('latin-1', errors='replace'),
            mimetype='text/plain; charset=iso-8859-1',
        )
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        logging.exception("api_plame_archivo26_generar_txt")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/declaracion-afp/listado', methods=['POST'])
@login_required
def api_declaracion_afp_listado():
    """sp_pr_listado_declaracion_afp_web: reporte analítico Declaración AFP."""
    body = request.get_json(silent=True) or {}
    p = _declaracion_afp_params_from_json(body)
    err = _declaracion_afp_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        sync_afp = _declaracion_afp_sincronizar_planilla(cursor, conn, p)
        resultado = _declaracion_afp_ejecutar_listado(cursor, p)
        filas, validaciones = _declaracion_afp_validaciones_completas(cursor, resultado, p)
        validaciones = _declaracion_afp_validaciones_sync_afp(validaciones, sync_afp)
        return jsonify({
            "rows": filas,
            "total": len(filas),
            "validaciones": validaciones,
            "puede_generar_xlsx": len(filas) > 0,
            "sync_afp": {
                "actualizado": sync_afp.get('actualizado'),
                "filas_afp": sync_afp.get('filas_afp'),
                "mensaje": sync_afp.get('mensaje'),
            },
        })
    except Exception as e:
        logging.exception("api_declaracion_afp_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/declaracion-afp/validar-resumen', methods=['POST'])
@login_required
def api_declaracion_afp_validar_resumen():
    """Resumen de verificación AFPnet sin generar Excel ni actualizar datos."""
    body = request.get_json(silent=True) or {}
    p = _declaracion_afp_params_from_json(body)
    err = _declaracion_afp_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    filas = _declaracion_afp_rows_from_json(body)
    if not filas:
        return jsonify({"error": "No hay registros en el listado para validar."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _declaracion_afp_sincronizar_planilla(cursor, conn, p)
        filas = _declaracion_afp_ejecutar_listado(cursor, p)
        if not filas:
            return jsonify({"error": "No hay registros para validar después de sincronizar AFP."}), 400
        montos_rows, planilla_row, planilla_trabajadores = _declaracion_afp_ejecutar_resumen_planilla(cursor, p)
        resumen = _declaracion_afp_build_resumen(
            montos_rows, planilla_row, filas, planilla_trabajadores
        )
        validaciones = _declaracion_afp_validar_regimen_pension_planilla(cursor, p)
        validaciones.extend(_declaracion_afp_validar_jubilados_filas(filas))
        return jsonify({
            'resumen': resumen,
            'tiene_diferencias': _declaracion_afp_resumen_tiene_diferencias(resumen),
            'validaciones': validaciones,
            'rows': filas,
        })
    except Exception as e:
        logging.exception("api_declaracion_afp_validar_resumen")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/declaracion-afp/generar-xlsx', methods=['POST'])
@login_required
def api_declaracion_afp_generar_xlsx():
    """Genera archivo Excel AFPnet (16 columnas, sin cabeceras)."""
    body = request.get_json(silent=True) or {}
    p = _declaracion_afp_params_from_json(body)
    err = _declaracion_afp_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _declaracion_afp_sincronizar_planilla(cursor, conn, p)
        filas = _declaracion_afp_ejecutar_listado(cursor, p)
        if not filas:
            return jsonify({"error": "No hay registros para generar el archivo AFPnet."}), 400

        montos_rows, planilla_row, planilla_trabajadores = _declaracion_afp_ejecutar_resumen_planilla(cursor, p)
        resumen = _declaracion_afp_build_resumen(
            montos_rows, planilla_row, filas, planilla_trabajadores
        )
        filas, validaciones = _declaracion_afp_validaciones_completas(cursor, filas, p)

        ruc = _obtener_ruc_compania(cursor, p['cia']) or '00000000000'

        buf = _declaracion_afp_generar_xlsx_bytes(filas)
        filename = f'AFPNET_{p["period"]}_{ruc}.xlsx'
        tiene_diferencias = _declaracion_afp_resumen_tiene_diferencias(resumen)
        return jsonify({
            'filename': filename,
            'file_base64': base64.b64encode(buf.getvalue()).decode('ascii'),
            'resumen': resumen if tiene_diferencias else None,
            'tiene_diferencias': tiene_diferencias,
            'rows': filas,
            'validaciones': validaciones,
        })
    except Exception as e:
        logging.exception("api_declaracion_afp_generar_xlsx")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/pago-haberes/telecredito')
@login_required
def pago_haberes_telecredito_page():
    return render_template('pago_haberes_telecredito.html')


@app.route('/api/pago-haberes/telecredito/listado', methods=['POST'])
@login_required
def api_pago_haberes_telecredito_listado():
    """sp_pr_listatelecredito_web: trabajadores con abono Telecrédito para el periodo/concepto."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    cesados = _normalize_cesados_telecredito(body.get('cesados'))

    log_sp = (
        '[telecredito listado] EXEC sp_pr_listatelecredito_web '
        f'@par_company={p["cia"]!r} @par_currency={p["currency"]!r} @par_concept={p["concept"]!r} '
        f'@par_payrolltype={p["payrolltype"]!r} @par_period={p["period"]!r} '
        f'@par_processtype={p["processtype"]!r} @par_paydate={p["paydate"].strftime("%Y-%m-%d %H:%M:%S")!r} '
        f'@cesados={cesados!r}'
    )
    logging.info(log_sp)
    print(log_sp, flush=True)

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listatelecredito_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @par_paydate=?, @cesados=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], p['paydate'], cesados,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        filas_pantalla = []
        filas_detalle = []
        for r in rows:
            person = str(r.get('person') or '').strip()
            dni = str(r.get('dni') or '').strip()
            nombre = str(r.get('nombre') or '').strip()
            tipodoc = str(r.get('tipodoc') or '').strip()
            importe = r.get('importe')
            try:
                importe_num = float(importe) if importe is not None else 0.0
            except Exception:
                importe_num = 0.0
            filas_pantalla.append([dni, tipodoc, nombre])
            filas_detalle.append({
                "person": person,
                "dni": dni,
                "tipodoc": tipodoc,
                "nombre": nombre,
                "importe": importe_num,
            })
        log_result = f'[telecredito listado] registros devueltos={len(filas_detalle)}'
        logging.info(log_result)
        print(log_result, flush=True)
        return jsonify({
            "headers": ['DNI', 'Tipo doc.', 'Nombre'],
            "data": filas_pantalla,
            "rows": filas_detalle,
            "meta": {
                "total": len(filas_detalle),
                "paydate": p['paydate'].strftime('%d/%m/%Y'),
            },
        })
    except Exception as e:
        logging.exception("api_pago_haberes_telecredito_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/pago-haberes/telecredito/generar-txt', methods=['POST'])
@login_required
def api_pago_haberes_telecredito_generar_txt():
    """sp_pr_generar_telecredito_web → archivo TXT Telecrédito BCP."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    persons = _telecredito_persons_from_json(body)
    if not persons:
        return jsonify({"error": "Seleccione al menos un trabajador."}), 400

    log_sp = (
        '[telecredito generar] EXEC sp_pr_generar_telecredito_web '
        f'@par_company={p["cia"]!r} @par_currency={p["currency"]!r} @par_concept={p["concept"]!r} '
        f'@par_payrolltype={p["payrolltype"]!r} @par_period={p["period"]!r} '
        f'@par_processtype={p["processtype"]!r} @par_paydate={p["paydate"].strftime("%Y-%m-%d %H:%M:%S")!r} '
        f'trabajadores_seleccionados={len(persons)}'
    )
    logging.info(log_sp)
    print(log_sp, flush=True)

    conn = None
    t0 = time.perf_counter()
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        t_conn = time.perf_counter()
        _telecredito_cargar_personas_temp(cursor, persons)
        t_temp = time.perf_counter()
        cursor.execute(
            "EXEC sp_pr_generar_telecredito_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @par_paydate=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], p['paydate'],
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        t_sp = time.perf_counter()
        lineas = []
        for r in rows:
            txt = str(r.get('linea_txt') or '').rstrip('\r\n')
            if txt:
                lineas.append(txt)

        detalle_count = max(len(lineas) - 1, 0)
        t_done = time.perf_counter()
        log_result = (
            f'[telecredito generar] seleccionados={len(persons)} '
            f'detalle_txt={detalle_count} '
            f'ms_conexion={int((t_conn - t0) * 1000)} '
            f'ms_temp={int((t_temp - t_conn) * 1000)} '
            f'ms_sp={int((t_sp - t_temp) * 1000)} '
            f'ms_total={int((t_done - t0) * 1000)}'
        )
        logging.info(log_result)
        print(log_result, flush=True)

        if len(lineas) < 2:
            campos = _telecredito_campos_faltantes(
                conn, p['cia'], lineas[1:] if len(lineas) > 1 else [], p['currency']
            )
            return jsonify({
                "error": (
                    f"No se pudo generar el archivo (sin líneas de detalle). "
                    f"Seleccionados: {len(persons)}."
                ),
                "campos_faltantes": campos,
            }), 400

        contenido = '\r\n'.join(lineas) + '\r\n'
        filename = _telecredito_filename(p['period'])

        resp = Response(
            contenido.encode('latin-1', errors='replace'),
            mimetype='text/plain; charset=iso-8859-1',
        )
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        logging.exception("api_pago_haberes_telecredito_generar_txt")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/pago-haberes/interbank/listado', methods=['POST'])
@login_required
def api_pago_haberes_interbank_listado():
    """sp_pr_listainterbank_web: trabajadores con abono Interbank para el periodo/concepto."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    cesados = _normalize_cesados_telecredito(body.get('cesados'))

    log_sp = (
        '[interbank listado] EXEC sp_pr_listainterbank_web '
        f'@par_company={p["cia"]!r} @par_currency={p["currency"]!r} @par_concept={p["concept"]!r} '
        f'@par_payrolltype={p["payrolltype"]!r} @par_period={p["period"]!r} '
        f'@par_processtype={p["processtype"]!r} @par_paydate={p["paydate"].strftime("%Y-%m-%d %H:%M:%S")!r} '
        f'@cesados={cesados!r}'
    )
    logging.info(log_sp)
    print(log_sp, flush=True)

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listainterbank_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @par_paydate=?, @cesados=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], p['paydate'], cesados,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        filas_pantalla = []
        filas_detalle = []
        for r in rows:
            person = str(r.get('person') or '').strip()
            dni = str(r.get('dni') or '').strip()
            nombre = str(r.get('nombre') or '').strip()
            tipodoc = str(r.get('tipodoc') or '').strip()
            importe = r.get('importe')
            try:
                importe_num = float(importe) if importe is not None else 0.0
            except Exception:
                importe_num = 0.0
            filas_pantalla.append([dni, tipodoc, nombre])
            filas_detalle.append({
                "person": person,
                "dni": dni,
                "tipodoc": tipodoc,
                "nombre": nombre,
                "importe": importe_num,
            })
        log_result = f'[interbank listado] registros devueltos={len(filas_detalle)}'
        logging.info(log_result)
        print(log_result, flush=True)
        return jsonify({
            "headers": ['DNI', 'Tipo doc.', 'Nombre'],
            "data": filas_pantalla,
            "rows": filas_detalle,
            "meta": {
                "total": len(filas_detalle),
                "paydate": p['paydate'].strftime('%d/%m/%Y'),
            },
        })
    except Exception as e:
        logging.exception("api_pago_haberes_interbank_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/pago-haberes/interbank/generar-txt', methods=['POST'])
@login_required
def api_pago_haberes_interbank_generar_txt():
    """sp_pr_generar_interbank_web → archivo TXT Interbank."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    persons = _telecredito_persons_from_json(body)
    if not persons:
        return jsonify({"error": "Seleccione al menos un trabajador."}), 400

    log_sp = (
        '[interbank generar] EXEC sp_pr_generar_interbank_web '
        f'@par_company={p["cia"]!r} @par_currency={p["currency"]!r} @par_concept={p["concept"]!r} '
        f'@par_payrolltype={p["payrolltype"]!r} @par_period={p["period"]!r} '
        f'@par_processtype={p["processtype"]!r} @par_paydate={p["paydate"].strftime("%Y-%m-%d %H:%M:%S")!r} '
        f'trabajadores_seleccionados={len(persons)}'
    )
    logging.info(log_sp)
    print(log_sp, flush=True)

    conn = None
    t0 = time.perf_counter()
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        t_conn = time.perf_counter()
        _pago_haberes_cargar_personas_temp(cursor, persons, 'InterbankPersonas')
        t_temp = time.perf_counter()
        cursor.execute(
            "EXEC sp_pr_generar_interbank_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @par_paydate=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], p['paydate'],
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        t_sp = time.perf_counter()
        lineas = []
        for r in rows:
            txt = str(r.get('linea_txt') or '').rstrip('\r\n')
            if txt:
                lineas.append(txt)

        detalle_count = max(len(lineas) - 1, 0)
        t_done = time.perf_counter()
        log_result = (
            f'[interbank generar] seleccionados={len(persons)} '
            f'detalle_txt={detalle_count} '
            f'ms_conexion={int((t_conn - t0) * 1000)} '
            f'ms_temp={int((t_temp - t_conn) * 1000)} '
            f'ms_sp={int((t_sp - t_temp) * 1000)} '
            f'ms_total={int((t_done - t0) * 1000)}'
        )
        logging.info(log_result)
        print(log_result, flush=True)

        if len(lineas) < 2:
            return jsonify({
                "error": (
                    f"No se pudo generar el archivo (sin líneas de detalle). "
                    f"Seleccionados: {len(persons)}."
                ),
            }), 400

        contenido = '\r\n'.join(lineas) + '\r\n'
        filename = _interbank_filename(p['period'])

        resp = Response(
            contenido.encode('latin-1', errors='replace'),
            mimetype='text/plain; charset=iso-8859-1',
        )
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        logging.exception("api_pago_haberes_interbank_generar_txt")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/pago-haberes/interbank')
@login_required
def pago_haberes_interbank_page():
    return render_template('pago_haberes_interbank.html')


@app.route('/pago-haberes/bbva')
@login_required
def pago_haberes_bbva_page():
    return render_template('pago_haberes_bbva.html')


@app.route('/pago-haberes/banbif')
@login_required
def pago_haberes_banbif_page():
    return render_template('pago_haberes_banbif.html')


@app.route('/api/pago-haberes/continental/listado', methods=['POST'])
@login_required
def api_pago_haberes_continental_listado():
    """sp_pr_listacontinental_web: trabajadores con abono Continental/BBVA."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    cesados = _normalize_cesados_telecredito(body.get('cesados'))
    todos_bancos = _normalize_todos_bancos_banbif(body.get('todos_bancos'))

    log_sp = (
        '[continental listado] EXEC sp_pr_listacontinental_web '
        f'@par_company={p["cia"]!r} @par_currency={p["currency"]!r} @par_concept={p["concept"]!r} '
        f'@par_payrolltype={p["payrolltype"]!r} @par_period={p["period"]!r} '
        f'@par_processtype={p["processtype"]!r} @par_paydate={p["paydate"].strftime("%Y-%m-%d %H:%M:%S")!r} '
        f'@cesados={cesados!r} @todos_bancos={todos_bancos!r}'
    )
    logging.info(log_sp)
    print(log_sp, flush=True)

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listacontinental_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @par_paydate=?, @cesados=?, @todos_bancos=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], p['paydate'], cesados, todos_bancos,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        filas_detalle = []
        for r in rows:
            person = str(r.get('person') or '').strip()
            dni = str(r.get('dni') or '').strip()
            nombre = str(r.get('nombre') or '').strip()
            tipodoc = str(r.get('tipodoc') or '').strip()
            importe = r.get('importe')
            try:
                importe_num = float(importe) if importe is not None else 0.0
            except Exception:
                importe_num = 0.0
            filas_detalle.append({
                "person": person,
                "dni": dni,
                "tipodoc": tipodoc,
                "nombre": nombre,
                "banco": str(r.get('banco') or '').strip(),
                "importe": importe_num,
            })
        log_result = f'[continental listado] registros devueltos={len(filas_detalle)} todos_bancos={todos_bancos}'
        logging.info(log_result)
        print(log_result, flush=True)
        headers = ['DNI', 'Tipo doc.', 'Nombre']
        if todos_bancos == 'Y':
            headers.append('Banco')
        headers.append('Importe')
        data_rows = []
        for r in filas_detalle:
            fila = [r['dni'], r['tipodoc'], r['nombre']]
            if todos_bancos == 'Y':
                fila.append(r['banco'])
            fila.append(r['importe'])
            data_rows.append(fila)
        return jsonify({
            "headers": headers,
            "data": data_rows,
            "rows": filas_detalle,
            "meta": {
                "total": len(filas_detalle),
                "paydate": p['paydate'].strftime('%d/%m/%Y'),
                "todos_bancos": todos_bancos == 'Y',
            },
        })
    except Exception as e:
        logging.exception("api_pago_haberes_continental_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/pago-haberes/continental/generar-txt', methods=['POST'])
@login_required
def api_pago_haberes_continental_generar_txt():
    """sp_pr_generar_continental_web → archivo TXT Continental/BBVA."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    persons = _telecredito_persons_from_json(body)
    if not persons:
        return jsonify({"error": "Seleccione al menos un trabajador."}), 400

    todos_bancos = _normalize_todos_bancos_banbif(body.get('todos_bancos'))

    log_sp = (
        '[continental generar] EXEC sp_pr_generar_continental_web '
        f'@par_company={p["cia"]!r} @par_currency={p["currency"]!r} @par_concept={p["concept"]!r} '
        f'@par_payrolltype={p["payrolltype"]!r} @par_period={p["period"]!r} '
        f'@par_processtype={p["processtype"]!r} @par_paydate={p["paydate"].strftime("%Y-%m-%d %H:%M:%S")!r} '
        f'@todos_bancos={todos_bancos!r} trabajadores_seleccionados={len(persons)}'
    )
    logging.info(log_sp)
    print(log_sp, flush=True)

    conn = None
    t0 = time.perf_counter()
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        t_conn = time.perf_counter()
        _pago_haberes_cargar_personas_temp(cursor, persons, 'ContinentalPersonas')
        t_temp = time.perf_counter()
        cursor.execute(
            "EXEC sp_pr_generar_continental_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @par_paydate=?, @todos_bancos=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], p['paydate'], todos_bancos,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        t_sp = time.perf_counter()
        lineas = []
        for r in rows:
            txt = str(r.get('linea_txt') or '').rstrip('\r\n')
            if txt:
                lineas.append(txt)

        detalle_count = max(len(lineas) - 1, 0)
        t_done = time.perf_counter()
        log_result = (
            f'[continental generar] seleccionados={len(persons)} todos_bancos={todos_bancos} '
            f'detalle_txt={detalle_count} '
            f'ms_conexion={int((t_conn - t0) * 1000)} '
            f'ms_temp={int((t_temp - t_conn) * 1000)} '
            f'ms_sp={int((t_sp - t_temp) * 1000)} '
            f'ms_total={int((t_done - t0) * 1000)}'
        )
        logging.info(log_result)
        print(log_result, flush=True)

        if len(lineas) < 2:
            return jsonify({
                "error": (
                    f"No se pudo generar el archivo (sin líneas de detalle). "
                    f"Seleccionados: {len(persons)}."
                ),
            }), 400

        contenido = '\r\n'.join(lineas) + '\r\n'
        filename = _continental_filename(p['period'])

        resp = Response(
            contenido.encode('latin-1', errors='replace'),
            mimetype='text/plain; charset=iso-8859-1',
        )
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        logging.exception("api_pago_haberes_continental_generar_txt")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/pago-haberes/banbif/listado', methods=['POST'])
@login_required
def api_pago_haberes_banbif_listado():
    """sp_pr_listabanbif_web: trabajadores con abono BANBIF."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    cesados = _normalize_cesados_telecredito(body.get('cesados'))
    todos_bancos = _normalize_todos_bancos_banbif(body.get('todos_bancos'))

    log_sp = (
        '[banbif listado] EXEC sp_pr_listabanbif_web '
        f'@par_company={p["cia"]!r} @par_currency={p["currency"]!r} @par_concept={p["concept"]!r} '
        f'@par_payrolltype={p["payrolltype"]!r} @par_period={p["period"]!r} '
        f'@par_processtype={p["processtype"]!r} @par_paydate={p["paydate"].strftime("%Y-%m-%d %H:%M:%S")!r} '
        f'@cesados={cesados!r} @todos_bancos={todos_bancos!r}'
    )
    logging.info(log_sp)
    print(log_sp, flush=True)

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listabanbif_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @par_paydate=?, @cesados=?, @todos_bancos=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], p['paydate'], cesados, todos_bancos,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        filas_detalle = []
        for r in rows:
            person = str(r.get('person') or '').strip()
            dni = str(r.get('dni') or '').strip()
            nombre = str(r.get('nombre') or '').strip()
            tipodoc = str(r.get('tipodoc') or '').strip()
            importe = r.get('importe')
            try:
                importe_num = float(importe) if importe is not None else 0.0
            except Exception:
                importe_num = 0.0
            filas_detalle.append({
                "person": person,
                "dni": dni,
                "tipodoc": tipodoc,
                "nombre": nombre,
                "banco": str(r.get('banco') or '').strip(),
                "importe": importe_num,
            })
        log_result = f'[banbif listado] registros devueltos={len(filas_detalle)} todos_bancos={todos_bancos}'
        logging.info(log_result)
        print(log_result, flush=True)
        headers = ['DNI', 'Tipo doc.', 'Nombre']
        if todos_bancos == 'Y':
            headers.append('Banco')
        headers.append('Importe')
        data_rows = []
        for r in filas_detalle:
            fila = [r['dni'], r['tipodoc'], r['nombre']]
            if todos_bancos == 'Y':
                fila.append(r['banco'])
            fila.append(r['importe'])
            data_rows.append(fila)
        return jsonify({
            "headers": headers,
            "data": data_rows,
            "rows": filas_detalle,
            "meta": {
                "total": len(filas_detalle),
                "paydate": p['paydate'].strftime('%d/%m/%Y'),
                "todos_bancos": todos_bancos == 'Y',
            },
        })
    except Exception as e:
        logging.exception("api_pago_haberes_banbif_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/pago-haberes/banbif/generar-txt', methods=['POST'])
@login_required
def api_pago_haberes_banbif_generar_txt():
    """sp_pr_generar_banbif_web → archivo TXT BANBIF."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    persons = _telecredito_persons_from_json(body)
    if not persons:
        return jsonify({"error": "Seleccione al menos un trabajador."}), 400

    todos_bancos = _normalize_todos_bancos_banbif(body.get('todos_bancos'))

    log_sp = (
        '[banbif generar] EXEC sp_pr_generar_banbif_web '
        f'@par_company={p["cia"]!r} @par_currency={p["currency"]!r} @par_concept={p["concept"]!r} '
        f'@par_payrolltype={p["payrolltype"]!r} @par_period={p["period"]!r} '
        f'@par_processtype={p["processtype"]!r} @par_paydate={p["paydate"].strftime("%Y-%m-%d %H:%M:%S")!r} '
        f'@todos_bancos={todos_bancos!r} trabajadores_seleccionados={len(persons)}'
    )
    logging.info(log_sp)
    print(log_sp, flush=True)

    conn = None
    t0 = time.perf_counter()
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        t_conn = time.perf_counter()
        _pago_haberes_cargar_personas_temp(cursor, persons, 'BanbifPersonas')
        t_temp = time.perf_counter()
        cursor.execute(
            "EXEC sp_pr_generar_banbif_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @par_paydate=?, @todos_bancos=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], p['paydate'], todos_bancos,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        t_sp = time.perf_counter()
        lineas = []
        for r in rows:
            txt = str(r.get('linea_txt') or '').rstrip('\r\n')
            if txt:
                lineas.append(txt)

        t_done = time.perf_counter()
        log_result = (
            f'[banbif generar] seleccionados={len(persons)} todos_bancos={todos_bancos} '
            f'detalle_txt={len(lineas)} '
            f'ms_conexion={int((t_conn - t0) * 1000)} '
            f'ms_temp={int((t_temp - t_conn) * 1000)} '
            f'ms_sp={int((t_sp - t_temp) * 1000)} '
            f'ms_total={int((t_done - t0) * 1000)}'
        )
        logging.info(log_result)
        print(log_result, flush=True)

        if not lineas:
            return jsonify({
                "error": (
                    f"No se pudo generar el archivo (sin líneas de detalle). "
                    f"Seleccionados: {len(persons)}."
                ),
            }), 400

        contenido = '\r\n'.join(lineas) + '\r\n'
        filename = _banbif_filename(p['period'])

        resp = Response(
            contenido.encode('latin-1', errors='replace'),
            mimetype='text/plain; charset=iso-8859-1',
        )
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        logging.exception("api_pago_haberes_banbif_generar_txt")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/generar_boletas')
@login_required
def generar_boletas_page():
    return render_template('generar_boletas.html')


@app.route('/liquidaciones/certificado_trabajo')
@login_required
def certificado_trabajo_page():
    return render_template('certificado_trabajo.html')


@app.route('/get_lista_certificado_trabajo', methods=['POST'])
@login_required
def get_lista_certificado_trabajo():
    """sp_pr_listadocertificadotrabajo_web — listado liquidación del periodo."""
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    person = str(body.get('person') or '0').strip() or '0'
    nombre = str(body.get('nombre') or body.get('busqueda') or body.get('name') or '').strip() or None

    if not cia or not payroll_type or not period:
        return jsonify({'error': 'Faltan compañía, tipo de planilla o periodo.'}), 400

    try:
        return jsonify(_listar_trabajadores_liquidacion(cia, payroll_type, period, person, nombre))
    except Exception as e:
        logging.exception('get_lista_certificado_trabajo')
        return jsonify({'error': str(e)}), 500


@app.route('/liquidaciones/certificado_retiro_cts')
@login_required
def certificado_retiro_cts_page():
    return render_template('certificado_retiro_cts.html')


@app.route('/get_lista_certificado_retiro_cts', methods=['POST'])
@login_required
def get_lista_certificado_retiro_cts():
    """sp_pr_listadocertificadotrabajo_web — mismo listado liquidación del periodo."""
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    person = str(body.get('person') or '0').strip() or '0'
    nombre = str(body.get('nombre') or body.get('busqueda') or body.get('name') or '').strip() or None

    if not cia or not payroll_type or not period:
        return jsonify({'error': 'Faltan compañía, tipo de planilla o periodo.'}), 400

    try:
        return jsonify(_listar_trabajadores_liquidacion(cia, payroll_type, period, person, nombre))
    except Exception as e:
        logging.exception('get_lista_certificado_retiro_cts')
        return jsonify({'error': str(e)}), 500


@app.route('/preview_certificado_retiro_cts')
@login_required
def preview_certificado_retiro_cts():
    params = request.args
    person = str(params.get('person') or '').strip()
    period = _normalize_pr_period(params.get('period'))
    try:
        pdf_buffer = generar_pdf_certificado_retiro_cts(params)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logging.exception('preview_certificado_retiro_cts')
        return jsonify({'error': str(e)}), 500
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=_certificado_retiro_cts_pdf_filename(person, period),
    )


@app.route('/procesar_certificados_retiro_cts_masivo', methods=['POST'])
@login_required
def procesar_certificados_retiro_cts_masivo():
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    modo = str(body.get('modo') or '').strip().lower()
    seleccionados = body.get('trabajadores') or []
    if modo not in ('zip', 'mail'):
        return jsonify({'error': 'Modo inválido. Use zip o mail.'}), 400
    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'No hay trabajadores seleccionados.'}), 400
    if not cia or not payroll_type or not period:
        return jsonify({'error': 'Faltan filtros para procesar certificados.'}), 400

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    if not ids:
        return jsonify({'error': 'No hay IDs válidos para procesar.'}), 400

    if modo == 'zip':
        company_name = str(body.get('company_name') or cia).strip()
        safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name).strip('_') or 'compania'
        period_yyyymm = period[:6] if len(period) >= 6 else period
        safe_period = re.sub(r'[^A-Za-z0-9_\\-]+', '_', period_yyyymm).strip('_') or 'periodo'
        nombre_zip = f'certificados_retiro_cts_{safe_company.lower()}_{safe_period}.zip'
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for pid in ids:
                pdf_data = generar_pdf_certificado_retiro_cts(
                    {
                        'person': pid,
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                    }
                )
                zf.writestr(_certificado_retiro_cts_pdf_filename(pid, period), pdf_data.getvalue())
        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            download_name=nombre_zip,
            as_attachment=True,
        )

    return jsonify(
        {
            'status': 'pending',
            'message': 'Use el modo Enviar por Email desde la pantalla.',
            'total': len(ids),
        }
    ), 202


@app.route('/descargar_zip_certificados_retiro_cts')
@login_required
def descargar_zip_certificados_retiro_cts():
    ensure_user_session()
    cia = session.get('company')
    payroll_type = (request.args.get('payroll_type') or '').strip()
    period = _normalize_pr_period(request.args.get('period'))
    company_name = (request.args.get('company_name') or '').strip()
    trabajadores_raw = (request.args.get('trabajadores') or '').strip()
    seleccionados = [x.strip() for x in trabajadores_raw.split(',') if x.strip()]

    if not (cia and payroll_type and period):
        flash('Faltan filtros para generar el ZIP de certificados.', 'warning')
        return redirect(url_for('certificado_retiro_cts_page'))

    try:
        empleados = _listar_trabajadores_liquidacion(cia, payroll_type, period, '0', None)
    except Exception:
        logging.exception('descargar_zip_certificados_retiro_cts listado')
        empleados = []

    if not empleados:
        flash('No hay certificados para procesar en este periodo.', 'warning')
        return redirect(url_for('certificado_retiro_cts_page'))

    if seleccionados:
        wanted = set(seleccionados)
        empleados = [e for e in empleados if str(e.get('person') or '').strip() in wanted]
        if not empleados:
            flash('La selección no contiene trabajadores válidos para el periodo indicado.', 'warning')
            return redirect(url_for('certificado_retiro_cts_page'))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for emp in empleados:
            person_id = str(emp.get('person') or '').strip()
            if not person_id:
                continue
            try:
                pdf_io = generar_pdf_certificado_retiro_cts(
                    {
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                        'person': person_id,
                    }
                )
                zip_file.writestr(_certificado_retiro_cts_pdf_filename(person_id, period), pdf_io.getvalue())
            except Exception:
                logging.exception('descargar_zip_certificados_retiro_cts persona=%s', person_id)
                continue

    zip_buffer.seek(0)
    safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name or cia).strip('_') or 'COMPANIA'
    safe_period = re.sub(r'[^0-9]+', '', period) or 'PERIODO'
    safe_payroll = re.sub(r'[^A-Za-z0-9_\\-]+', '_', payroll_type).strip('_') or 'PLANILLA'
    nombre_zip = f'Certificados_Retiro_CTS_{safe_company}_{safe_period}_{safe_payroll}.zip'
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=nombre_zip,
    )


@app.route('/enviar_certificados_retiro_cts_masivo', methods=['POST'])
@login_required
def enviar_certificados_retiro_cts_masivo():
    data = request.get_json(silent=True) or {}
    ensure_user_session()
    cia = session.get('company')
    payroll_type = str(data.get('payroll_type') or '').strip()
    period = _normalize_pr_period(data.get('period'))
    seleccionados = data.get('empleados', data.get('trabajadores', []))

    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'Debe enviar una lista de empleados.'}), 400
    if not (cia and payroll_type and period):
        return jsonify({'error': 'Faltan filtros para envío de certificados.'}), 400

    try:
        empleados_periodo = _listar_trabajadores_liquidacion(cia, payroll_type, period, '0', None)
    except Exception:
        logging.exception('enviar_certificados_retiro_cts_masivo listado')
        empleados_periodo = []

    by_person = {}
    for e in empleados_periodo:
        pid = str(e.get('person') or '').strip()
        if pid:
            by_person[pid] = e

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    total = len(ids)
    if total == 0:
        return jsonify({'error': 'No hay códigos de empleado válidos.'}), 400

    def generar_progreso_envio():
        enviados = 0
        errores = 0
        for idx, emp_code in enumerate(ids, start=1):
            emp = by_person.get(emp_code, {})
            emp_nombre = str(emp.get('nombre') or emp_code).strip()
            emp_email = str(emp.get('email') or '').strip()

            if not emp_email:
                errores += 1
                motivo = 'Sin email'
                yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'status': 'Error', 'detalle': motivo, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"
                continue

            try:
                pdf_buffer = generar_pdf_certificado_retiro_cts(
                    {
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                        'person': emp_code,
                    }
                )
                exito, msg = enviar_correo_certificado_retiro_cts(
                    destinatario=emp_email,
                    nombre_empleado=emp_nombre,
                    periodo=period,
                    sexo=emp.get('sex', 0),
                    pdf_io=pdf_buffer,
                    person=emp_code,
                )
                if exito:
                    enviados += 1
                    status = 'Enviado'
                    detalle = msg
                    motivo = ''
                else:
                    errores += 1
                    status = 'Error'
                    detalle = msg or 'No se pudo enviar el correo.'
                    motivo = detalle
            except Exception as e:
                logging.exception('enviar_certificados_retiro_cts_masivo persona=%s', emp_code)
                errores += 1
                status = 'Error'
                detalle = str(e)
                motivo = detalle

            yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'email': emp_email, 'status': status, 'detalle': detalle, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"

        yield f"data: {json.dumps({'done': True, 'enviados': enviados, 'errores': errores, 'total': total})}\n\n"

    return Response(
        stream_with_context(generar_progreso_envio()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no',
        },
    )


@app.route('/liquidaciones/formato_liquidacion')
@login_required
def formato_liquidacion_page():
    return render_template('formato_liquidacion.html')


@app.route('/get_lista_formato_liquidacion', methods=['POST'])
@login_required
def get_lista_formato_liquidacion():
    """sp_pr_listadocertificadotrabajo_web — listado liquidación del periodo."""
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    person = str(body.get('person') or '0').strip() or '0'
    nombre = str(body.get('nombre') or body.get('busqueda') or body.get('name') or '').strip() or None

    if not cia or not payroll_type or not period:
        return jsonify({'error': 'Faltan compañía, tipo de planilla o periodo.'}), 400

    try:
        return jsonify(_listar_trabajadores_liquidacion(cia, payroll_type, period, person, nombre))
    except Exception as e:
        logging.exception('get_lista_formato_liquidacion')
        return jsonify({'error': str(e)}), 500


@app.route('/preview_formato_liquidacion')
@login_required
def preview_formato_liquidacion():
    params = request.args
    person = str(params.get('person') or '').strip()
    period = _normalize_pr_period(params.get('period'))
    try:
        pdf_buffer = generar_pdf_formato_liquidacion(params)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logging.exception('preview_formato_liquidacion')
        return jsonify({'error': str(e)}), 500
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=_formato_liquidacion_pdf_filename(person, period),
    )


@app.route('/procesar_formatos_liquidacion_masivo', methods=['POST'])
@login_required
def procesar_formatos_liquidacion_masivo():
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    modo = str(body.get('modo') or '').strip().lower()
    seleccionados = body.get('trabajadores') or []
    if modo not in ('zip', 'mail'):
        return jsonify({'error': 'Modo inválido. Use zip o mail.'}), 400
    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'No hay trabajadores seleccionados.'}), 400
    if not cia or not payroll_type or not period:
        return jsonify({'error': 'Faltan filtros para procesar formatos.'}), 400

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    if not ids:
        return jsonify({'error': 'No hay IDs válidos para procesar.'}), 400

    if modo == 'zip':
        company_name = str(body.get('company_name') or cia).strip()
        safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name).strip('_') or 'compania'
        period_yyyymm = period[:6] if len(period) >= 6 else period
        safe_period = re.sub(r'[^A-Za-z0-9_\\-]+', '_', period_yyyymm).strip('_') or 'periodo'
        nombre_zip = f'formatos_liquidacion_{safe_company.lower()}_{safe_period}.zip'
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for pid in ids:
                pdf_data = generar_pdf_formato_liquidacion(
                    {
                        'person': pid,
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                    }
                )
                zf.writestr(_formato_liquidacion_pdf_filename(pid, period), pdf_data.getvalue())
        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            download_name=nombre_zip,
            as_attachment=True,
        )

    return jsonify(
        {
            'status': 'pending',
            'message': 'Use el modo Enviar por Email desde la pantalla.',
            'total': len(ids),
        }
    ), 202


@app.route('/descargar_zip_formatos_liquidacion')
@login_required
def descargar_zip_formatos_liquidacion():
    ensure_user_session()
    cia = session.get('company')
    payroll_type = (request.args.get('payroll_type') or '').strip()
    period = _normalize_pr_period(request.args.get('period'))
    company_name = (request.args.get('company_name') or '').strip()
    trabajadores_raw = (request.args.get('trabajadores') or '').strip()
    seleccionados = [x.strip() for x in trabajadores_raw.split(',') if x.strip()]

    if not (cia and payroll_type and period):
        flash('Faltan filtros para generar el ZIP de formatos.', 'warning')
        return redirect(url_for('formato_liquidacion_page'))

    try:
        empleados = _listar_trabajadores_liquidacion(cia, payroll_type, period, '0', None)
    except Exception:
        logging.exception('descargar_zip_formatos_liquidacion listado')
        empleados = []

    if not empleados:
        flash('No hay formatos para procesar en este periodo.', 'warning')
        return redirect(url_for('formato_liquidacion_page'))

    if seleccionados:
        wanted = set(seleccionados)
        empleados = [e for e in empleados if str(e.get('person') or '').strip() in wanted]
        if not empleados:
            flash('La selección no contiene trabajadores válidos para el periodo indicado.', 'warning')
            return redirect(url_for('formato_liquidacion_page'))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for emp in empleados:
            person_id = str(emp.get('person') or '').strip()
            if not person_id:
                continue
            try:
                pdf_io = generar_pdf_formato_liquidacion(
                    {
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                        'person': person_id,
                    }
                )
                zip_file.writestr(_formato_liquidacion_pdf_filename(person_id, period), pdf_io.getvalue())
            except Exception:
                logging.exception('descargar_zip_formatos_liquidacion persona=%s', person_id)
                continue

    zip_buffer.seek(0)
    safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name or cia).strip('_') or 'COMPANIA'
    safe_period = re.sub(r'[^0-9]+', '', period) or 'PERIODO'
    safe_payroll = re.sub(r'[^A-Za-z0-9_\\-]+', '_', payroll_type).strip('_') or 'PLANILLA'
    nombre_zip = f'Formatos_Liquidacion_{safe_company}_{safe_period}_{safe_payroll}.zip'
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=nombre_zip,
    )


@app.route('/enviar_formatos_liquidacion_masivo', methods=['POST'])
@login_required
def enviar_formatos_liquidacion_masivo():
    data = request.get_json(silent=True) or {}
    ensure_user_session()
    cia = session.get('company')
    payroll_type = str(data.get('payroll_type') or '').strip()
    period = _normalize_pr_period(data.get('period'))
    seleccionados = data.get('empleados', data.get('trabajadores', []))

    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'Debe enviar una lista de empleados.'}), 400
    if not (cia and payroll_type and period):
        return jsonify({'error': 'Faltan filtros para envío de formatos.'}), 400

    try:
        empleados_periodo = _listar_trabajadores_liquidacion(cia, payroll_type, period, '0', None)
    except Exception:
        logging.exception('enviar_formatos_liquidacion_masivo listado')
        empleados_periodo = []

    by_person = {}
    for e in empleados_periodo:
        pid = str(e.get('person') or '').strip()
        if pid:
            by_person[pid] = e

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    total = len(ids)
    if total == 0:
        return jsonify({'error': 'No hay códigos de empleado válidos.'}), 400

    def generar_progreso_envio():
        enviados = 0
        errores = 0
        for idx, emp_code in enumerate(ids, start=1):
            emp = by_person.get(emp_code, {})
            emp_nombre = str(emp.get('nombre') or emp_code).strip()
            emp_email = str(emp.get('email') or '').strip()

            if not emp_email:
                errores += 1
                motivo = 'Sin email'
                yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'status': 'Error', 'detalle': motivo, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"
                continue

            try:
                pdf_buffer = generar_pdf_formato_liquidacion(
                    {
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                        'person': emp_code,
                    }
                )
                exito, msg = enviar_correo_formato_liquidacion(
                    destinatario=emp_email,
                    nombre_empleado=emp_nombre,
                    periodo=period,
                    sexo=emp.get('sex', 0),
                    pdf_io=pdf_buffer,
                    person=emp_code,
                )
                if exito:
                    enviados += 1
                    status = 'Enviado'
                    detalle = msg
                    motivo = ''
                else:
                    errores += 1
                    status = 'Error'
                    detalle = msg or 'No se pudo enviar el correo.'
                    motivo = detalle
            except Exception as e:
                logging.exception('enviar_formatos_liquidacion_masivo persona=%s', emp_code)
                errores += 1
                status = 'Error'
                detalle = str(e)
                motivo = detalle

            yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'email': emp_email, 'status': status, 'detalle': detalle, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"

        yield f"data: {json.dumps({'done': True, 'enviados': enviados, 'errores': errores, 'total': total})}\n\n"

    return Response(
        stream_with_context(generar_progreso_envio()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no',
        },
    )


@app.route('/preview_certificado_trabajo')
@login_required
def preview_certificado_trabajo():
    params = request.args
    person = str(params.get('person') or '').strip()
    period = _normalize_pr_period(params.get('period'))
    try:
        pdf_buffer = generar_pdf_certificado_trabajo(params)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logging.exception('preview_certificado_trabajo')
        return jsonify({'error': str(e)}), 500
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=_certificado_trabajo_pdf_filename(person, period),
    )


@app.route('/procesar_certificados_trabajo_masivo', methods=['POST'])
@login_required
def procesar_certificados_trabajo_masivo():
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    modo = str(body.get('modo') or '').strip().lower()
    seleccionados = body.get('trabajadores') or []
    if modo not in ('zip', 'mail'):
        return jsonify({'error': 'Modo inválido. Use zip o mail.'}), 400
    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'No hay trabajadores seleccionados.'}), 400
    if not cia or not payroll_type or not period:
        return jsonify({'error': 'Faltan filtros para procesar certificados.'}), 400

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    if not ids:
        return jsonify({'error': 'No hay IDs válidos para procesar.'}), 400

    if modo == 'zip':
        company_name = str(body.get('company_name') or cia).strip()
        safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name).strip('_') or 'compania'
        period_yyyymm = period[:6] if len(period) >= 6 else period
        safe_period = re.sub(r'[^A-Za-z0-9_\\-]+', '_', period_yyyymm).strip('_') or 'periodo'
        nombre_zip = f'certificados_trabajo_{safe_company.lower()}_{safe_period}.zip'
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for pid in ids:
                pdf_data = generar_pdf_certificado_trabajo(
                    {
                        'person': pid,
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                    }
                )
                zf.writestr(_certificado_trabajo_pdf_filename(pid, period), pdf_data.getvalue())
        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            download_name=nombre_zip,
            as_attachment=True,
        )

    return jsonify(
        {
            'status': 'pending',
            'message': 'Use el modo Enviar por Email desde la pantalla.',
            'total': len(ids),
        }
    ), 202


@app.route('/descargar_zip_certificados_trabajo')
@login_required
def descargar_zip_certificados_trabajo():
    ensure_user_session()
    cia = session.get('company')
    payroll_type = (request.args.get('payroll_type') or '').strip()
    period = _normalize_pr_period(request.args.get('period'))
    company_name = (request.args.get('company_name') or '').strip()
    trabajadores_raw = (request.args.get('trabajadores') or '').strip()
    seleccionados = [x.strip() for x in trabajadores_raw.split(',') if x.strip()]

    if not (cia and payroll_type and period):
        flash('Faltan filtros para generar el ZIP de certificados.', 'warning')
        return redirect(url_for('certificado_trabajo_page'))

    conn = None
    empleados = []
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'EXEC sp_pr_listadocertificadotrabajo_web @cia=?, @payrolltype=?, @period=?, @person=?, @nombre=?',
            (cia, payroll_type, period, '0', None),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        empleados = rows or []
    except Exception:
        logging.exception('descargar_zip_certificados_trabajo listado')
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    if not empleados:
        flash('No hay certificados para procesar en este periodo.', 'warning')
        return redirect(url_for('certificado_trabajo_page'))

    if seleccionados:
        wanted = set(seleccionados)
        empleados = [e for e in empleados if str(e.get('person') or '').strip() in wanted]
        if not empleados:
            flash('La selección no contiene trabajadores válidos para el periodo indicado.', 'warning')
            return redirect(url_for('certificado_trabajo_page'))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for emp in empleados:
            person_id = str(emp.get('person') or '').strip()
            if not person_id:
                continue
            try:
                pdf_io = generar_pdf_certificado_trabajo(
                    {
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                        'person': person_id,
                    }
                )
                zip_file.writestr(_certificado_trabajo_pdf_filename(person_id, period), pdf_io.getvalue())
            except Exception:
                logging.exception('descargar_zip_certificados_trabajo persona=%s', person_id)
                continue

    zip_buffer.seek(0)
    safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name or cia).strip('_') or 'COMPANIA'
    safe_period = re.sub(r'[^0-9]+', '', period) or 'PERIODO'
    safe_payroll = re.sub(r'[^A-Za-z0-9_\\-]+', '_', payroll_type).strip('_') or 'PLANILLA'
    nombre_zip = f'Certificados_Trabajo_{safe_company}_{safe_period}_{safe_payroll}.zip'
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=nombre_zip,
    )


@app.route('/enviar_certificados_trabajo_masivo', methods=['POST'])
@login_required
def enviar_certificados_trabajo_masivo():
    data = request.get_json(silent=True) or {}
    ensure_user_session()
    cia = session.get('company')
    payroll_type = str(data.get('payroll_type') or '').strip()
    period = _normalize_pr_period(data.get('period'))
    seleccionados = data.get('empleados', data.get('trabajadores', []))

    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'Debe enviar una lista de empleados.'}), 400
    if not (cia and payroll_type and period):
        return jsonify({'error': 'Faltan filtros para envío de certificados.'}), 400

    conn = None
    empleados_periodo = []
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'EXEC sp_pr_listadocertificadotrabajo_web @cia=?, @payrolltype=?, @period=?, @person=?, @nombre=?',
            (cia, payroll_type, period, '0', None),
        )
        empleados_periodo = _dicts_first_nonempty_resultset(cursor) or []
    except Exception:
        logging.exception('enviar_certificados_trabajo_masivo listado')
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    by_person = {}
    for e in empleados_periodo:
        pid = str(e.get('person') or '').strip()
        if pid:
            by_person[pid] = e

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    total = len(ids)
    if total == 0:
        return jsonify({'error': 'No hay códigos de empleado válidos.'}), 400

    def generar_progreso_envio():
        enviados = 0
        errores = 0
        for idx, emp_code in enumerate(ids, start=1):
            emp = by_person.get(emp_code, {})
            emp_nombre = str(emp.get('nombre') or emp_code).strip()
            emp_email = str(emp.get('email') or '').strip()

            if not emp_email:
                errores += 1
                motivo = 'Sin email'
                yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'status': 'Error', 'detalle': motivo, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"
                continue

            try:
                pdf_buffer = generar_pdf_certificado_trabajo(
                    {
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'period': period,
                        'person': emp_code,
                    }
                )
                exito, msg = enviar_correo_certificado_trabajo(
                    destinatario=emp_email,
                    nombre_empleado=emp_nombre,
                    periodo=period,
                    sexo=emp.get('sex', 0),
                    pdf_io=pdf_buffer,
                    person=emp_code,
                )
                if exito:
                    enviados += 1
                    status = 'Enviado'
                    detalle = msg
                    motivo = ''
                else:
                    errores += 1
                    status = 'Error'
                    detalle = msg or 'No se pudo enviar el correo.'
                    motivo = detalle
            except Exception as e:
                logging.exception('enviar_certificados_trabajo_masivo persona=%s', emp_code)
                errores += 1
                status = 'Error'
                detalle = str(e)
                motivo = detalle

            yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'email': emp_email, 'status': status, 'detalle': detalle, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"

        yield f"data: {json.dumps({'done': True, 'enviados': enviados, 'errores': errores, 'total': total})}\n\n"

    return Response(
        stream_with_context(generar_progreso_envio()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no',
        },
    )


@app.route('/impuesto_renta/certificado_quinta')
@login_required
def certificado_quinta_page():
    return render_template(
        'certificado_quinta.html',
        anio_actual=date.today().year,
        fecha_hoy=date.today().strftime('%Y-%m-%d'),
    )


@app.route('/impuesto_renta/calculo_quinta_trabajador')
@login_required
def calculo_quinta_trabajador_page():
    return render_template('calculo_quinta_trabajador.html')


@app.route('/get_calculo_quinta_trabajador', methods=['POST'])
@login_required
def get_calculo_quinta_trabajador():
    """sp_pr_5ta_trabajador_web: seguimiento de cálculo de 5ta por trabajador."""
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    process = str(body.get('process') or body.get('processtype') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    person = str(body.get('person') or body.get('employee') or '').strip()

    if not cia or not payroll_type or not process or not period or not person:
        return jsonify({'error': 'Faltan compañía, planilla, proceso, periodo o trabajador.'}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout(cursor)
        cursor.execute(
            'EXEC sp_pr_5ta_trabajador_web @company=?, @payrolltype=?, @process=?, @period=?, @person=?',
            (cia, payroll_type, process, period, person),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
    except Exception as e:
        logging.exception('get_calculo_quinta_trabajador')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    if not rows:
        return jsonify({'error': 'No se encontraron datos para los filtros indicados.'}), 404

    raw = rows[0]
    reporte = _armar_reporte_quinta_trabajador(raw, period)
    empresa = _get_company_header_quinta(cia)
    return jsonify(
        {
            'empresa': empresa,
            'reporte': reporte,
            'filtros': {
                'cia': cia,
                'payroll_type': payroll_type,
                'process': process,
                'period': period,
                'person': person,
            },
            'fecha_emision': datetime.now().strftime('%d-%m-%Y %H:%M:%S'),
        }
    )


@app.route('/get_lista_certificado_quinta', methods=['POST'])
@login_required
def get_lista_certificado_quinta():
    """
    sp_pr_listadocertificadoquinta_web @cia, @payrolltype, @anio, @person.

    Lista trabajadores con al menos una planilla en cualquier proceso del año indicado.
    """
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    anio = str(body.get('anio') or body.get('year') or '').strip()
    person = str(body.get('person') or '0').strip() or '0'

    if not cia or not payroll_type or not anio:
        return jsonify({'error': 'Faltan compañía, tipo de planilla o año.'}), 400
    if len(anio) != 4 or not anio.isdigit():
        return jsonify({'error': 'Año inválido. Use cuatro dígitos (ej. 2026).'}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'EXEC sp_pr_listadocertificadoquinta_web @cia=?, @payrolltype=?, @anio=?, @person=?',
            (cia, payroll_type, anio, person),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        trabajadores = []
        for r in rows:
            fi = _jsonable_value(r.get('fechaingreso'))
            fc = _jsonable_value(r.get('fechacese'))
            trabajadores.append(
                {
                    'person': str(r.get('person') or '').strip(),
                    'nombre': str(r.get('nombre') or '').strip(),
                    'email': str(r.get('email') or '').strip(),
                    'ingreso': fi if fi is not None else '',
                    'cese': fc if fc is not None else '',
                }
            )
        return jsonify(trabajadores)
    except Exception as e:
        logging.exception('get_lista_certificado_quinta')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/get_detalle_calculo_certificado_quinta', methods=['POST'])
@login_required
def get_detalle_calculo_certificado_quinta():
    """
    sp_pr_detallecalculocertificadoquinta_web: detalle sueldos/asignaciones + utilidades.
    """
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    anio = str(body.get('anio') or body.get('year') or '').strip()
    person = str(body.get('person') or '').strip()
    nombre = str(body.get('nombre') or '').strip()

    if not cia or not payroll_type or not anio or not person:
        return jsonify({'error': 'Faltan compañía, tipo de planilla, año o trabajador.'}), 400
    if len(anio) != 4 or not anio.isdigit():
        return jsonify({'error': 'Año inválido. Use cuatro dígitos (ej. 2026).'}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'EXEC sp_pr_detallecalculocertificadoquinta_web @cia=?, @payrolltype=?, @anio=?, @person=?',
            (cia, payroll_type, anio, person),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        detalle = []
        total_sueldos = 0.0
        total_utilidades = 0.0
        for r in rows:
            try:
                importe = float(r.get('importe') or 0)
            except (TypeError, ValueError):
                importe = 0.0
            tipo_linea = str(r.get('tipo_linea') or '').strip().upper()
            if tipo_linea == 'UTILIDADES':
                total_utilidades += importe
            else:
                total_sueldos += importe
            detalle.append(
                {
                    'proceso': str(r.get('proceso') or '').strip(),
                    'proceso_descripcion': str(r.get('proceso_descripcion') or '').strip(),
                    'periodo': _jsonable_value(r.get('periodo')) or '',
                    'concepto': str(r.get('concepto') or '').strip(),
                    'formulacode': str(r.get('formulacode') or '').strip(),
                    'importe': importe,
                    'tipo_linea': tipo_linea or 'SUELDOS',
                }
            )
        return jsonify(
            {
                'person': person,
                'nombre': nombre,
                'anio': anio,
                'rows': detalle,
                'total_sueldos_asignaciones': total_sueldos,
                'total_utilidades': total_utilidades,
            }
        )
    except Exception as e:
        logging.exception('get_detalle_calculo_certificado_quinta')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/preview_certificado_quinta')
@login_required
def preview_certificado_quinta():
    params = request.args
    person = str(params.get('person') or '').strip()
    anio = str(params.get('anio') or params.get('year') or '').strip()
    try:
        pdf_buffer = generar_pdf_certificado_quinta(params)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logging.exception('preview_certificado_quinta')
        return jsonify({'error': str(e)}), 500
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=_certificado_quinta_pdf_filename(person, anio),
    )


@app.route('/procesar_certificados_quinta_masivo', methods=['POST'])
@login_required
def procesar_certificados_quinta_masivo():
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    anio = str(body.get('anio') or body.get('year') or '').strip()
    fecha_emision = str(body.get('fecha_emision') or '').strip()
    modo = str(body.get('modo') or '').strip().lower()
    seleccionados = body.get('trabajadores') or []
    if modo not in ('zip', 'mail'):
        return jsonify({'error': 'Modo inválido. Use zip o mail.'}), 400
    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'No hay trabajadores seleccionados.'}), 400
    if not cia or not payroll_type or not anio:
        return jsonify({'error': 'Faltan filtros para procesar certificados de quinta.'}), 400
    if len(anio) != 4 or not anio.isdigit():
        return jsonify({'error': 'Año inválido. Use cuatro dígitos (ej. 2026).'}), 400

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    if not ids:
        return jsonify({'error': 'No hay IDs válidos para procesar.'}), 400

    if modo == 'zip':
        company_name = str(body.get('company_name') or cia).strip()
        safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name).strip('_') or 'compania'
        safe_anio = re.sub(r'[^0-9]+', '', anio) or 'anio'
        nombre_zip = f'certificados_quinta_{safe_company.lower()}_{safe_anio}.zip'
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for pid in ids:
                pdf_data = generar_pdf_certificado_quinta(
                    {
                        'person': pid,
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'anio': anio,
                        'fecha_emision': fecha_emision,
                    }
                )
                zf.writestr(_certificado_quinta_pdf_filename(pid, anio), pdf_data.getvalue())
        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            download_name=nombre_zip,
            as_attachment=True,
        )

    return jsonify(
        {
            'status': 'pending',
            'message': 'Modo envío por Email pendiente de integración.',
            'total': len(ids),
        }
    ), 202


@app.route('/descargar_zip_certificados_quinta')
@login_required
def descargar_zip_certificados_quinta():
    ensure_user_session()
    cia = str(request.args.get('cia') or session.get('company') or '').strip()
    payroll_type = (request.args.get('payroll_type') or '').strip()
    anio = str(request.args.get('anio') or request.args.get('year') or '').strip()
    fecha_emision = str(request.args.get('fecha_emision') or '').strip()
    company_name = (request.args.get('company_name') or '').strip()
    trabajadores_raw = (request.args.get('trabajadores') or '').strip()
    seleccionados = [x.strip() for x in trabajadores_raw.split(',') if x.strip()]

    if not (cia and payroll_type and anio):
        flash('Faltan filtros para generar el ZIP de certificados de quinta.', 'warning')
        return redirect(url_for('certificado_quinta_page'))
    if len(anio) != 4 or not anio.isdigit():
        flash('Año inválido para generar el ZIP.', 'warning')
        return redirect(url_for('certificado_quinta_page'))

    empleados = get_listado_certificado_quinta(cia, payroll_type, anio, '0')
    if not empleados:
        flash('No hay certificados para procesar en este año.', 'warning')
        return redirect(url_for('certificado_quinta_page'))

    if seleccionados:
        wanted = set(seleccionados)
        empleados = [e for e in empleados if str(e.get('person') or '').strip() in wanted]
        if not empleados:
            flash('La selección no contiene trabajadores válidos para el año indicado.', 'warning')
            return redirect(url_for('certificado_quinta_page'))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for emp in empleados:
            person_id = str(emp.get('person') or '').strip()
            if not person_id:
                continue
            try:
                params = {
                    'cia': cia,
                    'payroll_type': payroll_type,
                    'anio': anio,
                    'person': person_id,
                    'fecha_emision': fecha_emision,
                }
                pdf_io = generar_pdf_certificado_quinta(params)
                nombre_pdf = _certificado_quinta_pdf_filename(person_id, anio)
                zip_file.writestr(nombre_pdf, pdf_io.getvalue())
            except Exception:
                logging.exception('descargar_zip_certificados_quinta persona=%s', person_id)
                continue

    zip_buffer.seek(0)
    safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name or cia).strip('_') or 'COMPANIA'
    safe_anio = re.sub(r'[^0-9]+', '', anio) or 'ANIO'
    safe_payroll = re.sub(r'[^A-Za-z0-9_\\-]+', '_', payroll_type).strip('_') or 'PLANILLA'
    nombre_zip = f'CertificadosQuinta_{safe_company}_{safe_anio}_{safe_payroll}.zip'
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=nombre_zip,
    )


@app.route('/enviar_certificados_quinta_masivo', methods=['POST'])
@login_required
def enviar_certificados_quinta_masivo():
    data = request.get_json(silent=True) or {}
    ensure_user_session()
    cia = str(data.get('cia') or session.get('company') or '').strip()
    payroll_type = str(data.get('payroll_type') or '').strip()
    anio = str(data.get('anio') or data.get('year') or '').strip()
    fecha_emision = str(data.get('fecha_emision') or '').strip()
    seleccionados = data.get('empleados', data.get('trabajadores', []))

    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'Debe enviar una lista de empleados.'}), 400
    if not (cia and payroll_type and anio):
        return jsonify({'error': 'Faltan filtros para envío de certificados de quinta.'}), 400
    if len(anio) != 4 or not anio.isdigit():
        return jsonify({'error': 'Año inválido. Use cuatro dígitos (ej. 2026).'}), 400

    empleados_periodo = get_listado_certificado_quinta(cia, payroll_type, anio, '0')
    by_person = {}
    for e in empleados_periodo:
        pid = str(e.get('person') or '').strip()
        if pid:
            by_person[pid] = e

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    total = len(ids)
    if total == 0:
        return jsonify({'error': 'No hay códigos de empleado válidos.'}), 400

    def generar_progreso_envio():
        enviados = 0
        errores = 0
        for idx, emp_code in enumerate(ids, start=1):
            emp = by_person.get(emp_code, {})
            emp_nombre = str(emp.get('nombre') or emp_code).strip()
            emp_email = str(emp.get('email') or '').strip()

            if not emp_email:
                errores += 1
                motivo = 'Sin email'
                yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'status': 'Error', 'detalle': motivo, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"
                continue

            try:
                pdf_buffer = generar_pdf_certificado_quinta(
                    {
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'anio': anio,
                        'person': emp_code,
                        'fecha_emision': fecha_emision,
                    }
                )
                exito, msg = enviar_correo_certificado_quinta(
                    destinatario=emp_email,
                    nombre_empleado=emp_nombre,
                    anio=anio,
                    sexo=emp.get('sex', emp.get('sexo', 0)),
                    pdf_io=pdf_buffer,
                    person=emp_code,
                )
                if exito:
                    enviados += 1
                    status = 'Enviado'
                    detalle = msg
                    motivo = ''
                else:
                    errores += 1
                    status = 'Error'
                    detalle = msg or 'No se pudo enviar el correo.'
                    motivo = detalle
            except Exception as e:
                logging.exception('enviar_certificados_quinta_masivo persona=%s', emp_code)
                errores += 1
                status = 'Error'
                detalle = str(e)
                motivo = detalle

            yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'email': emp_email, 'status': status, 'detalle': detalle, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"

        yield f"data: {json.dumps({'done': True, 'enviados': enviados, 'errores': errores, 'total': total})}\n\n"

    return Response(
        stream_with_context(generar_progreso_envio()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no',
        },
    )


@app.route('/get_lista_boletas', methods=['POST'])
@login_required
def get_lista_boletas():
    """
    sp_pr_listadogenerarboletas_web @cia, @payrolltype, @processtype, @period, @person.

    Nota BD: el filtro para listar todos con @person = '0' debe ser
    ``(@person = '0' OR PR_EmployeePayRoll.Person = @person)``.
    Si el SP usa ``(@person = '0' AND Person = @person)``, no devolverá filas al listar todos.
    """
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    processtype = str(body.get('process') or body.get('processtype') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    person = str(body.get('person') or '0').strip() or '0'
    nombre = str(body.get('nombre') or body.get('busqueda') or body.get('name') or '').strip() or None

    if not cia or not payroll_type or not processtype or not period:
        return jsonify({'error': 'Faltan compañía, tipo de planilla, proceso o periodo.'}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'EXEC sp_pr_listadogenerarboletas_web @cia=?, @payrolltype=?, @processtype=?, @period=?, @person=?, @nombre=?',
            (cia, payroll_type, processtype, period, person, nombre),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        trabajadores = []
        for r in rows:
            fi = _jsonable_value(r.get('fechaingreso'))
            fc = _jsonable_value(r.get('fechacese'))
            trabajadores.append(
                {
                    'person': str(r.get('person') or '').strip(),
                    'nombre': str(r.get('nombre') or '').strip(),
                    'email': str(r.get('email') or '').strip(),
                    'ingreso': fi if fi is not None else '',
                    'cese': fc if fc is not None else '',
                }
            )
        return jsonify(trabajadores)
    except Exception as e:
        logging.exception('get_lista_boletas')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/preview_boleta')
@login_required
def preview_boleta():
    params = request.args
    person = str(params.get('person') or '').strip()
    period = _normalize_pr_period(params.get('period'))
    try:
        pdf_buffer = generar_pdf_en_memoria(params)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logging.exception('preview_boleta')
        return jsonify({'error': str(e)}), 500
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=_boleta_pdf_filename(person, period),
    )


@app.route('/procesar_boletas_masivo', methods=['POST'])
@login_required
def procesar_boletas_masivo():
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    process = str(body.get('process') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    modo = str(body.get('modo') or '').strip().lower()
    seleccionados = body.get('trabajadores') or []
    if modo not in ('zip', 'mail'):
        return jsonify({'error': 'Modo inválido. Use zip o mail.'}), 400
    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'No hay trabajadores seleccionados.'}), 400
    if not cia or not payroll_type or not process or not period:
        return jsonify({'error': 'Faltan filtros para procesar boletas.'}), 400

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    if not ids:
        return jsonify({'error': 'No hay IDs válidos para procesar.'}), 400

    if modo == 'zip':
        company_name = str(body.get('company_name') or cia).strip()
        safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name).strip('_') or 'compania'
        # Periodo de BD viene como yyyymmdd; pediste nombre con yyyymm.
        period_yyyymm = period[:6] if len(period) >= 6 else period
        safe_period = re.sub(r'[^A-Za-z0-9_\\-]+', '_', period_yyyymm).strip('_') or 'periodo'
        nombre_zip = f'boletas_{safe_company.lower()}_{safe_period}.zip'
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for pid in ids:
                pdf_data = generar_pdf_en_memoria(
                    {
                        'person': pid,
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'process': process,
                        'period': period,
                    }
                )
                zf.writestr(_boleta_pdf_filename(pid, period), pdf_data.getvalue())
        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            download_name=nombre_zip,
            as_attachment=True,
        )

    # Stub controlado para modo correo (pendiente integración real de SMTP/servicio).
    return jsonify(
        {
            'status': 'pending',
            'message': 'Modo envío por Email pendiente de integración.',
            'total': len(ids),
        }
    ), 202


@app.route('/descargar_zip_boletas')
@login_required
def descargar_zip_boletas():
    ensure_user_session()
    cia = session.get('company')
    payroll_type = (request.args.get('payroll_type') or '').strip()
    processtype = (request.args.get('process') or '').strip()
    period = _normalize_pr_period(request.args.get('period'))
    company_name = (request.args.get('company_name') or '').strip()
    trabajadores_raw = (request.args.get('trabajadores') or '').strip()
    seleccionados = [x.strip() for x in trabajadores_raw.split(',') if x.strip()]

    if not (cia and payroll_type and processtype and period):
        flash('Faltan filtros para generar el ZIP de boletas.', 'warning')
        return redirect(url_for('generar_boletas_page'))

    empleados = get_listado_generar_boletas(cia, payroll_type, processtype, period, '0')
    if not empleados:
        flash('No hay boletas para procesar en este periodo.', 'warning')
        return redirect(url_for('generar_boletas_page'))

    # Si se envía selección, limita a esos códigos.
    if seleccionados:
        wanted = set(seleccionados)
        empleados = [e for e in empleados if str(e.get('person') or e.get('employeecode') or '').strip() in wanted]
        if not empleados:
            flash('La selección no contiene boletas válidas para el periodo indicado.', 'warning')
            return redirect(url_for('generar_boletas_page'))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for emp in empleados:
            person_id = str(emp.get('person') or emp.get('employeecode') or '').strip()
            if not person_id:
                continue
            try:
                params = {
                    'cia': cia,
                    'payroll_type': payroll_type,
                    'process': processtype,
                    'period': period,
                    'person': person_id,
                }
                pdf_io = generar_pdf_en_memoria(params)
                nombre_pdf = _boleta_pdf_filename(person_id, period)
                zip_file.writestr(nombre_pdf, pdf_io.getvalue())
            except Exception as e:
                logging.exception('descargar_zip_boletas persona=%s', person_id)
                continue

    zip_buffer.seek(0)
    safe_company = re.sub(r'[^A-Za-z0-9_\\-]+', '_', company_name or cia).strip('_') or 'COMPANIA'
    safe_period = re.sub(r'[^A-Za-z0-9_\\-]+', '_', period).strip('_') or 'PERIODO'
    safe_payroll = re.sub(r'[^A-Za-z0-9_\\-]+', '_', payroll_type).strip('_') or 'PLANILLA'
    nombre_zip = f'Boletas_{safe_company}_{safe_period}_{safe_payroll}.zip'
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=nombre_zip,
    )


@app.route('/enviar_boletas_masivo', methods=['POST'])
@login_required
def enviar_boletas_masivo():
    data = request.get_json(silent=True) or {}
    ensure_user_session()
    cia = session.get('company')
    payroll_type = str(data.get('payroll_type') or '').strip()
    process = str(data.get('process') or '').strip()
    period = _normalize_pr_period(data.get('period'))
    seleccionados = data.get('empleados', data.get('trabajadores', []))

    if not isinstance(seleccionados, list) or not seleccionados:
        return jsonify({'error': 'Debe enviar una lista de empleados.'}), 400
    if not (cia and payroll_type and process and period):
        return jsonify({'error': 'Faltan filtros para envío de boletas.'}), 400

    # Trae email/nombre del mismo SP de listado para el periodo.
    empleados_periodo = get_listado_generar_boletas(cia, payroll_type, process, period, '0')
    by_person = {}
    for e in empleados_periodo:
        pid = str(e.get('person') or e.get('employeecode') or '').strip()
        if pid:
            by_person[pid] = e

    ids = [str(x).strip() for x in seleccionados if str(x).strip()]
    total = len(ids)
    if total == 0:
        return jsonify({'error': 'No hay códigos de empleado válidos.'}), 400

    def generar_progreso_envio():
        enviados = 0
        errores = 0
        for idx, emp_code in enumerate(ids, start=1):
            emp = by_person.get(emp_code, {})
            emp_nombre = str(emp.get('nombre') or emp.get('fullname') or emp_code).strip()
            emp_email = str(emp.get('email') or '').strip()

            if not emp_email:
                errores += 1
                motivo = 'Sin email'
                yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'status': 'Error', 'detalle': motivo, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"
                continue

            try:
                pdf_buffer = generar_pdf_en_memoria(
                    {
                        'cia': cia,
                        'payroll_type': payroll_type,
                        'process': process,
                        'period': period,
                        'person': emp_code,
                    }
                )
                exito, msg = enviar_correo_boleta(
                    destinatario=emp_email,
                    nombre_empleado=emp_nombre,
                    periodo=period,
                    sexo=emp.get('sex', emp.get('sexo', 0)),
                    pdf_io=pdf_buffer,
                    person=emp_code,
                )
                if exito:
                    enviados += 1
                    status = 'Enviado'
                    detalle = msg
                    motivo = ''
                else:
                    errores += 1
                    status = 'Error'
                    detalle = msg or 'No se pudo enviar el correo.'
                    motivo = detalle
            except Exception as e:
                logging.exception('enviar_boletas_masivo persona=%s', emp_code)
                errores += 1
                status = 'Error'
                detalle = str(e)
                motivo = detalle

            yield f"data: {json.dumps({'empleado': emp_nombre, 'codigo': emp_code, 'email': emp_email, 'status': status, 'detalle': detalle, 'motivo': motivo, 'actual': idx, 'total': total, 'progreso': int((idx / total) * 100)})}\n\n"

        yield f"data: {json.dumps({'done': True, 'enviados': enviados, 'errores': errores, 'total': total})}\n\n"

    return Response(
        stream_with_context(generar_progreso_envio()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no',
        },
    )


# ==========================================
# APIS PARA SELECTORES EN CASCADA (stored procedures)
# ==========================================


@app.route('/api/selectores/companias')
@login_required
def api_companias():
    """sp_pr_selectorcompanias_web → Company, description (@cia para el resto)."""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectorcompanias_web")
        rows = cursor.fetchall()
        data = [{"id": r.Company, "text": r.description} for r in rows]
        return jsonify(data)
    except Exception:
        logging.exception("api_companias")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/bancos')
@login_required
def api_bancos():
    """sp_pr_selectorbancos_web @cia → bank, name."""
    cia = request.args.get('cia')
    if not cia:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectorbancos_web @cia=?", (cia,))
        col_names = [str(c[0]).strip() for c in (cursor.description or [])]
        rows = cursor.fetchall()
        data = []
        for row in rows:
            rd = _row_dict_from_columns(col_names, row)
            data.append({
                "id": rd.get("bank"),
                "text": rd.get("name"),
            })
        return jsonify(data)
    except Exception:
        logging.exception("api_bancos")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/unidades')
@login_required
def api_unidades():
    """sp_pr_selectorunidades_web → replicationunit, description."""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectorunidades_web")
        col_names = [str(c[0]).strip() for c in (cursor.description or [])]
        rows = cursor.fetchall()
        data = []
        for row in rows:
            rd = _row_dict_from_columns(col_names, row)
            data.append({
                "id": rd.get("replicationunit") or rd.get("ReplicationUnit"),
                "text": rd.get("description") or rd.get("Description"),
            })
        return jsonify(data)
    except Exception:
        logging.exception("api_unidades")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/conceptos')
@login_required
def api_conceptos():
    """sp_pr_selectorconceptos_web @cia → concept, description."""
    cia = request.args.get('cia')
    if not cia:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectorconceptos_web @cia=?", (cia,))
        col_names = [str(c[0]).strip() for c in (cursor.description or [])]
        rows = cursor.fetchall()
        data = []
        for row in rows:
            rd = _row_dict_from_columns(col_names, row)
            data.append({
                "id": rd.get("concept"),
                "text": rd.get("description"),
            })
        return jsonify(data)
    except Exception:
        logging.exception("api_conceptos")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/planillas')
@login_required
def api_planillas():
    """sp_pr_selectorplanillas_web @cia → payrolltype, tipoplanilla"""
    cia = request.args.get('cia')
    if not cia:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectorplanillas_web @cia=?", (cia,))
        rows = cursor.fetchall()
        data = [{"id": r.payrolltype, "text": r.tipoplanilla} for r in rows]
        return jsonify(data)
    except Exception:
        logging.exception("api_planillas")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/procesos')
@login_required
def api_procesos():
    """sp_pr_selectorprocesos_web @cia, @payrolltype → processtype, proceso"""
    cia = request.args.get('cia')
    payrolltype = request.args.get('payrolltype')
    if not cia or not payrolltype:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorprocesos_web @cia=?, @payrolltype=?",
            (cia, payrolltype),
        )
        rows = cursor.fetchall()
        data = [{"id": r.processtype, "text": r.proceso} for r in rows]
        return jsonify(data)
    except Exception:
        logging.exception("api_procesos")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/periodos')
@login_required
def api_periodos():
    """sp_pr_selectorperiodos_web @cia, @payrolltype, @processtype → period, periodo"""
    cia = request.args.get('cia')
    payrolltype = request.args.get('payrolltype')
    processtype = request.args.get('processtype')
    if not all([cia, payrolltype, processtype]):
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorperiodos_web @cia=?, @payrolltype=?, @processtype=?",
            (cia, payrolltype, processtype),
        )
        rows = cursor.fetchall()
        data = [{"id": r.period, "text": r.periodo} for r in rows]
        return jsonify(data)
    except Exception:
        logging.exception("api_periodos")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/concepto-neto')
@login_required
def api_concepto_neto():
    """sp_pr_selectorconceptoneto_web: concepto Neto a recibir (FormulaCode = NETO)."""
    cia = request.args.get('cia', '').strip()
    if not cia:
        return jsonify({"concept": "", "description": ""})
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectorconceptoneto_web @cia=?", (cia,))
        rows = _dicts_first_nonempty_resultset(cursor)
        concept = ''
        description = ''
        if rows:
            concept = str(rows[0].get('concept') or '').strip()
            description = str(rows[0].get('description') or '').strip()
        return jsonify({"concept": concept, "description": description})
    except Exception:
        logging.exception("api_concepto_neto")
        return jsonify({"concept": "", "description": ""})
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/periodo-activo')
@login_required
def api_periodo_activo():
    """sp_pr_selectorperiodoactivo_web: periodo activo en PR_ProcessControl (Status = A)."""
    cia = request.args.get('cia', '').strip()
    payrolltype = request.args.get('payrolltype', '').strip()
    processtype = request.args.get('processtype', '').strip()
    if not all([cia, payrolltype, processtype]):
        return jsonify({"prperiod": ""})
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorperiodoactivo_web @cia=?, @payrolltype=?, @processtype=?",
            (cia, payrolltype, processtype),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        prperiod = ''
        if rows:
            raw = rows[0].get('prperiod')
            prperiod = _normalize_pr_period(raw) or str(raw or '').strip()
        return jsonify({"prperiod": prperiod})
    except Exception:
        logging.exception("api_periodo_activo")
        return jsonify({"prperiod": ""})
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/afp')
@login_required
def api_selectores_afp():
    """sp_pr_selectorafp_web @cia → afp, description."""
    cia = request.args.get('cia')
    if not cia:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        return jsonify(_selector_items_from_sp(cursor, 'EXEC sp_pr_selectorafp_web @cia=?', (cia,)))
    except Exception:
        logging.exception("api_selectores_afp")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/periodos-plame')
@login_required
def api_periodos_plame():
    """sp_pr_selectorperiodos_plame_web @cia → prperiod (YYYYMM), description (YYYY-MM)."""
    cia = request.args.get('cia')
    if not cia:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorperiodos_plame_web @cia=?",
            (cia,),
        )
        desc = cursor.description
        rows = cursor.fetchall()
        if not rows or not desc:
            return jsonify([])
        cols = [str(c[0] or '').strip().lower() for c in desc]
        data = []
        for row in rows:
            rd = {cols[i]: row[i] for i in range(len(cols))}
            pid = rd.get('prperiod')
            txt = rd.get('description')
            pid_s = str(pid).strip() if pid is not None else ''
            txt_s = str(txt).strip() if txt is not None else pid_s
            if pid_s:
                data.append({"id": pid_s, "text": txt_s})
        return jsonify(data)
    except Exception:
        logging.exception("api_periodos_plame")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/periodos-asig')
@login_required
def api_periodos_asig():
    """sp_pr_selectorperiodos_asig_web @cia, @payrolltype → PRPERIOD (id), description (text)."""
    cia = request.args.get('cia')
    payrolltype = request.args.get('payrolltype')
    if not cia or not payrolltype:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorperiodos_asig_web @cia=?, @payrolltype=?",
            (cia, payrolltype),
        )
        desc = cursor.description
        rows = cursor.fetchall()
        if not rows or not desc:
            return jsonify([])
        cols = [str(c[0] or '').strip().lower() for c in desc]
        data = []
        for row in rows:
            rd = {cols[i]: row[i] for i in range(len(cols))}
            pid = rd.get('prperiod')
            txt = rd.get('description')
            pid_s = str(pid).strip() if pid is not None else ''
            txt_s = str(txt).strip() if txt is not None else pid_s
            if pid_s:
                data.append({"id": pid_s, "text": txt_s})
        return jsonify(data)
    except Exception:
        logging.exception("api_periodos_asig")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/trabajadores/listado', methods=['POST'])
@login_required
def api_trabajadores_listado():
    """sp_pr_listatrabajadores_web: listado con filtros incl. rango fecha de ingreso."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '0').strip() or '0'
    person = str(body.get('person') or '0').strip() or '0'
    docnro = str(body.get('docnro') or body.get('dni') or '').strip()
    nombre = str(body.get('nombre') or body.get('name') or '').strip()
    estado = _normalize_estado_trabajador(body.get('estado'))
    salarybank = str(body.get('salarybank') or body.get('salary_bank') or '0').strip() or '0'
    cesados = _normalize_cesados_telecredito(body.get('cesados'))
    fecha_ingreso_all, fecha_ingreso_desde, fecha_ingreso_hasta = _trabajadores_fecha_ingreso_from_json(body)

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if fecha_ingreso_all == 'N':
        if not fecha_ingreso_desde or not fecha_ingreso_hasta:
            return jsonify({"error": "Indique fecha de ingreso desde y hasta."}), 400
        if fecha_ingreso_desde > fecha_ingreso_hasta:
            return jsonify({"error": "La fecha de ingreso desde no puede ser mayor que hasta."}), 400

    headers_es = [
        'Tipo planilla',
        'Código',
        'Nombre',
        'Estado',
        'Tipo documento',
        'Nro. documento',
        'F. ingreso',
        'F. cese',
        'Cargo',
        'Acciones',
    ]
    keys_datos = [
        'tipoplanilla', 'codigo', 'nombre', 'estado', 'tipodocumento', 'numerodocumento',
        'fechaingreso', 'fechacese', 'cargo',
    ]
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        fecha_desde_sql = _sql_date_str_param(fecha_ingreso_desde) if fecha_ingreso_all == 'N' else ''
        fecha_hasta_sql = _sql_date_str_param(fecha_ingreso_hasta) if fecha_ingreso_all == 'N' else ''
        cursor.execute(
            "EXEC sp_pr_listatrabajadores_web "
            "@cia=?, @payrolltype=?, @person=?, @docnro=?, @nombre=?, @estado=?, "
            "@salarybank=?, @cesados=?, @fecha_ingreso_all=?, @fecha_ingreso_desde=?, @fecha_ingreso_hasta=?",
            (
                cia, payrolltype, person, docnro, nombre, estado, salarybank, cesados,
                fecha_ingreso_all, fecha_desde_sql, fecha_hasta_sql,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        rows_meta = []
        for r in rows:
            fila = [_jsonable_value(r.get(key)) for key in keys_datos]
            fila.append('')
            resultado.append(fila)
            rows_meta.append({
                'person': str(r.get('person') or '').strip(),
                'codigo': str(r.get('codigo') or '').strip(),
            })
        return jsonify({"headers": headers_es, "data": resultado, "rows": rows_meta})
    except Exception as e:
        logging.exception("api_trabajadores_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/trabajadores')
@login_required
def api_trabajadores():
    """sp_pr_selectorpersonas_web @cia → Person, Name"""
    cia = request.args.get('cia')
    if not cia:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectorpersonas_web @cia=?", (cia,))
        rows = cursor.fetchall()
        data = [{"id": r.Person, "text": r.Name} for r in rows]
        return jsonify(data)
    except Exception:
        logging.exception("api_trabajadores")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/selectores/tipos-descanso-medico')
@login_required
def api_tipos_descanso_medico():
    """sp_pr_selectortipos_dm_web @cia → MedicalRestType, Description."""
    cia = request.args.get('cia')
    if not cia:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("EXEC sp_pr_selectortipos_dm_web @cia=?", (cia,))
        col_names = [str(c[0]).strip() for c in (cursor.description or [])]
        rows = cursor.fetchall()
        data = []
        for row in rows:
            rd = _row_dict_from_columns(col_names, row)
            data.append(
                {
                    "id": rd.get("medicalresttype"),
                    "text": rd.get("description"),
                }
            )
        return jsonify(data)
    except Exception:
        logging.exception("api_tipos_descanso_medico")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


# ==========================================
# API REPORTE PRINCIPAL
# ==========================================


@app.route('/api/reportes/promedio-liquidaciones', methods=['POST'])
@login_required
def api_reporte_promedio_liq():
    """SP_PR_ReportePromedioLiquidacion @cia, @payrolltype, @period, @person (varchar)."""
    params = _report_params_from_json(request)
    if not params:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC SP_PR_ReportePromedioLiquidacion @cia=?, @payrolltype=?, @period=?, @person=?",
            params,
        )
        columns, rows = _fetch_first_nonempty_resultset(cursor)
        if not rows:
            return jsonify([])
        data = [{col: _jsonable_value(val) for col, val in zip(columns, row)} for row in rows]
        return jsonify(data)
    except Exception:
        logging.exception("api_reporte_promedio_liq params=%s", params)
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


def _row_dict_lower(cursor, row):
    """Convierte una fila pyodbc en dict con claves en minúsculas."""
    if not cursor.description:
        return {}
    return {
        str(col[0]).strip().lower(): row[i]
        for i, col in enumerate(cursor.description)
    }


def _row_dict_from_columns(column_names, row):
    """Igual que _row_dict_lower pero con nombres ya capturados (tras nextset)."""
    return {
        str(column_names[i]).strip().lower(): row[i]
        for i in range(len(column_names))
    }


def _drain_all_cursor_resultsets(cursor):
    """Consume todos los lotes devueltos por un SP (SET NOCOUNT off, varios SELECT, etc.)."""
    while True:
        if cursor.description:
            try:
                cursor.fetchall()
            except Exception:
                pass
        if not cursor.nextset():
            break


def _fetch_last_query_resultset(cursor):
    """
    SPs con CREATE/INSERT/UPDATE antes del SELECT no dejan un result set en el primer lote;
    pyodbc exige no hacer fetchall() si no hay consulta. Tomamos el último lote con description.
    """
    last_cols = None
    last_rows = None
    while True:
        if cursor.description:
            last_cols = [str(c[0]).strip() for c in cursor.description]
            last_rows = cursor.fetchall()
        if not cursor.nextset():
            break
    return last_cols or [], last_rows or []


def _float_sp_cell(value):
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


@app.route('/reporte-resumen-total')
@login_required
def reporte_resumen_total():
    return render_template('reporte_resumen_total.html')


@app.route('/reporte_resumen_total', methods=['POST'])
@login_required
def reporte_resumen_total_post():
    """sp_pr_reporteplame_total_web: resumen por concepto y tipo (Mensual, Semanal, …)."""
    body = request.get_json(silent=True) or {}
    cia = (body.get('cia') or '').strip()
    payroll_type = (body.get('payroll_type') or '').strip()
    period = (body.get('period') or '').strip()

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not payroll_type or not period:
        return jsonify({"error": "Debe indicar tipo de planilla y periodo."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_reporteplame_total_web @cia=?, @payrolltype=?, @period=?, @person=?",
            (cia, payroll_type, period, None),
        )
        col_names, rows = _fetch_last_query_resultset(cursor)
        resumen = []
        for row in rows:
            rd = _row_dict_from_columns(col_names, row)
            mensual = _float_sp_cell(rd.get('mensual'))
            semanal = _float_sp_cell(rd.get('semanal'))
            liquida = _float_sp_cell(rd.get('liquida'))
            vacaciones = _float_sp_cell(rd.get('vacaciones'))
            cts = _float_sp_cell(rd.get('cts'))
            grati = _float_sp_cell(rd.get('grati'))
            total_fila = mensual + semanal + liquida + vacaciones + cts + grati

            tipo_raw = rd.get('tipo')
            tipo = tipo_raw.strip() if isinstance(tipo_raw, str) else (str(tipo_raw).strip() if tipo_raw is not None else '')

            pdt_val = rd.get('pdt')
            concepto_val = rd.get('concepto')

            resumen.append(
                {
                    "tipo": tipo,
                    "pdt": '' if pdt_val is None else str(pdt_val).strip(),
                    "concepto": '' if concepto_val is None else str(concepto_val).strip(),
                    "mensual": mensual,
                    "semanal": semanal,
                    "liquida": liquida,
                    "vacaciones": vacaciones,
                    "cts": cts,
                    "grati": grati,
                    "total": total_fila,
                }
            )
        return jsonify(resumen)
    except Exception as e:
        logging.exception("reporte_resumen_total_post")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/reporte_planilla_vertical', methods=['POST'])
@login_required
def reporte_planilla_vertical_post():
    """
    sp_pr_reporteplamevertical_web @cia, @payrolltype, @process, @period, @person, @salarybank,
    @fecha_ingreso_all, @fecha_ingreso_desde, @fecha_ingreso_hasta.
    Cabeceras dinámicas desde xx_plamevertical2 + PR_Concept; datos desde xx_reporteplanilla.
    """
    body = request.get_json(silent=True) or {}
    cia = (body.get('cia') or '').strip()
    payroll_type = (body.get('payroll_type') or body.get('payrolltype') or '').strip()
    process = (body.get('process') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    person = (body.get('person') or '0').strip() or '0'
    salarybank = str(body.get('salarybank') if body.get('salarybank') is not None else body.get('salary_bank') or '').strip()
    fecha_ingreso_all, fecha_ingreso_desde, fecha_ingreso_hasta = _trabajadores_fecha_ingreso_from_json(body)

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not payroll_type or not process or not period:
        return jsonify({"error": "Debe indicar tipo de planilla, proceso y periodo."}), 400
    if fecha_ingreso_all == 'N':
        if not fecha_ingreso_desde or not fecha_ingreso_hasta:
            return jsonify({"error": "Indique fecha de ingreso desde y hasta."}), 400
        if fecha_ingreso_desde > fecha_ingreso_hasta:
            return jsonify({"error": "La fecha de ingreso desde no puede ser mayor que hasta."}), 400

    static_headers_es = [
        'Código',
        'Nombre',
        'F.Ingreso',
        'F.Cese',
        'Cargo',
        'AFP',
        'C.Costo',
        'Cod.Costo',
        'Unidad',
        'TipoPago',
        'Perfil',
        'Horas',
        'Banco',
        'Num. Cuenta',
    ]
    static_keys = [
        'person',
        'name',
        'entrydate',
        'ceasedate',
        'position',
        'afp',
        'ccname',
        'costcenter',
        'unidad',
        'tipopago',
        'profile',
        'horas',
        'banco',
        'numcuenta',
    ]

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        fecha_desde_sql = _sql_date_str_param(fecha_ingreso_desde) if fecha_ingreso_all == 'N' else ''
        fecha_hasta_sql = _sql_date_str_param(fecha_ingreso_hasta) if fecha_ingreso_all == 'N' else ''
        cursor.execute(
            "EXEC sp_pr_reporteplamevertical_web "
            "@cia=?, @payrolltype=?, @process=?, @period=?, @person=?, @salarybank=?, "
            "@fecha_ingreso_all=?, @fecha_ingreso_desde=?, @fecha_ingreso_hasta=?",
            (
                cia, payroll_type, process, period, person, salarybank,
                fecha_ingreso_all, fecha_desde_sql, fecha_hasta_sql,
            ),
        )
        _drain_all_cursor_resultsets(cursor)

        cursor.execute(
            """
            SELECT DISTINCT UPPER(PR_Concept.PrintText) AS conceptname, PR_Concept.reporden
            FROM xx_plamevertical2
            INNER JOIN PR_Concept ON (
                xx_plamevertical2.conceptname = PR_Concept.Description
                AND PR_Concept.Company = ?
            )
            ORDER BY PR_Concept.reporden ASC, 1 ASC
            """,
            (cia,),
        )
        concept_rows = cursor.fetchall()
        conceptos_dinamicos = []
        for crow in concept_rows:
            cname = crow[0] if crow[0] is not None else ''
            cname = str(cname).strip()
            if cname:
                conceptos_dinamicos.append(cname)

        headers = list(static_headers_es) + conceptos_dinamicos
        num_concepts = len(conceptos_dinamicos)

        # Mismo SELECT que el SP (@grupo = 'N'): no usar SELECT * sobre la tabla,
        # porque position/costcenter almacenan IDs; el SP expone descripción y CCCode.
        concept_cols_sql = ", ".join(f"concept{str(i).zfill(2)}" for i in range(1, 66))
        sql_datos = f"""
            SELECT
                person,
                name,
                entrydate,
                ceasedate,
                (SELECT Description FROM PR_Position WHERE Position = xx_reporteplanilla.position) AS position,
                afp,
                (SELECT Description FROM AC_CostCenter WHERE CostCenter = xx_reporteplanilla.costcenter) AS ccname,
                (SELECT CCCode FROM AC_CostCenter WHERE CostCenter = xx_reporteplanilla.costcenter) AS costcenter,
                (SELECT Description FROM SY_ReplicationUnit
                 INNER JOIN SY_Person ON (SY_ReplicationUnit.ReplicationUnit = SY_Person.ReplicationUnit)
                 WHERE SY_Person.Person = xx_reporteplanilla.person) AS unidad,
                (SELECT CASE WHEN ISNULL(SY_Person.isrecruiter, 'N') = 'Y' THEN 'H' ELSE 'P' END
                 FROM sy_person WHERE person = xx_reporteplanilla.person) AS tipopago,
                (SELECT description FROM PR_AccountProfile
                 INNER JOIN PR_Employee ON (
                     PR_AccountProfile.AccountProfile = PR_Employee.AccountProfile
                     AND PR_AccountProfile.company = ?
                     AND PR_Employee.Person = xx_reporteplanilla.person)) AS profile,
                (SELECT SUM(hourday) FROM PR_REGISTERHOUR
                 WHERE period = ? AND Company = ? AND person = xx_reporteplanilla.person) AS horas,
                CASE WHEN (
                    SELECT ShortName FROM PR_ProcessType
                    WHERE Company = ? AND ProcessType = ?
                ) = 'CTS' THEN (
                    SELECT name FROM ERP_Bank
                    INNER JOIN PR_Employee ON (
                        ERP_Bank.Bank = PR_Employee.CTSBank
                        AND ERP_Bank.company = ?
                        AND PR_Employee.Person = xx_reporteplanilla.person)
                ) ELSE (
                    SELECT name FROM ERP_Bank
                    INNER JOIN PR_Employee ON (
                        ERP_Bank.Bank = PR_Employee.SalaryBank
                        AND ERP_Bank.company = ?
                        AND PR_Employee.Person = xx_reporteplanilla.person)
                ) END AS banco,
                CASE WHEN (
                    SELECT ShortName FROM PR_ProcessType
                    WHERE Company = ? AND ProcessType = ?
                ) = 'CTS' THEN (
                    SELECT CTSAccount FROM PR_Employee
                    WHERE PR_Employee.Person = xx_reporteplanilla.person AND PR_Employee.Company = ?
                ) ELSE (
                    SELECT salaryaccount FROM PR_Employee
                    WHERE PR_Employee.Person = xx_reporteplanilla.person AND PR_Employee.Company = ?
                ) END AS numcuenta,
                {concept_cols_sql}
            FROM xx_reporteplanilla
            ORDER BY name
        """
        params_datos = (
            cia,
            period,
            cia,
            cia,
            process,
            cia,
            cia,
            cia,
            process,
            cia,
            cia,
        )
        cursor.execute(sql_datos, params_datos)
        desc = cursor.description
        if not desc:
            return jsonify({"headers": headers, "data": []})
        col_names = [str(c[0]).strip().lower() for c in desc]
        rows = cursor.fetchall()

        resultado = []
        for row in rows:
            rd = {col_names[i]: row[i] for i in range(len(col_names))}
            fila = []
            for key in static_keys:
                fila.append(_jsonable_value(rd.get(key)))
            for i in range(num_concepts):
                cn = f"concept{str(i + 1).zfill(2)}"
                fila.append(_float_sp_cell(rd.get(cn)))
            resultado.append(fila)

        return jsonify({"headers": headers, "data": resultado})
    except Exception as e:
        logging.exception("reporte_planilla_vertical_post")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/reporte_vacaciones_detalle', methods=['POST'])
@login_required
def reporte_vacaciones_detalle_post():
    """sp_pr_r019_vacationdetail_web @cia, @payrolltype, @period, @person."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or '').strip()
    payroll_type = str(body.get('payroll_type') or body.get('payrolltype') or '').strip()
    period_raw = body.get('period')
    ps = str(period_raw).strip() if period_raw is not None else ''
    if ps == '' or ps == '0':
        period = '0'
    else:
        period = _normalize_pr_period(period_raw)
    person = str(body.get('person') or '0').strip() or '0'

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not payroll_type:
        return jsonify({"error": "Debe indicar tipo de planilla."}), 400

    headers_es = [
        'Periodo',
        'Código',
        'Nombre',
        'Fecha inicio',
        'Fecha fin',
        'Días',
        'Año control',
        'Cargo',
    ]
    keys_datos = ['person', 'name', 'datebegin', 'dateend', 'days', 'controlyear', 'cargo']

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout_report(cursor)
        cursor.execute(
            "EXEC sp_pr_r019_vacationdetail_web @cia=?, @payrolltype=?, @period=?, @person=?",
            (cia, payroll_type, period, person),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            fila = [_fmt_periodo_yyyy_mm(r.get('prperiod'))]
            for key in keys_datos:
                val = r.get(key)
                if key == 'days' and val is not None:
                    try:
                        fila.append(int(round(float(val))))
                    except Exception:
                        fila.append(_jsonable_value(val))
                else:
                    fila.append(_jsonable_value(val))
            resultado.append(fila)
        return jsonify({"headers": headers_es, "data": resultado})
    except Exception as e:
        logging.exception("reporte_vacaciones_detalle_post")
        return jsonify({"error": _reporte_sql_error_message(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/reporte_saldo_vacaciones', methods=['POST'])
@login_required
def reporte_saldo_vacaciones_post():
    """sp_pr_saldovacaciones_web @company, @payrolltype, @date, @person, @cesados."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    payroll_type = str(body.get('payroll_type') or body.get('payrolltype') or '').strip()
    person = str(body.get('person') or '0').strip() or '0'
    fecha_dt = _parse_report_date(body.get('fecha') or body.get('date'))
    cesados = _normalize_cesados_saldo_vacaciones(body.get('cesados'))

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not payroll_type:
        return jsonify({"error": "Debe indicar tipo de planilla."}), 400

    anios = _anios_saldo_vacaciones(fecha_dt)
    headers_es = [
        'Tipo planilla',
        'Código',
        'Nombre',
        'Unidad',
        'F. ingreso',
        'F. cese',
        f'Saldo {anios[0]}',
        f'Saldo {anios[1]}',
        f'Saldo {anios[2]}',
        f'Saldo {anios[3]}',
        f'Saldo {anios[4]}',
        'Faltas',
        'Licencias',
        'Descansos',
        'Saldo',
    ]
    keys_datos = [
        'tipoplanillas', 'person', 'name', 'description', 'entrydate', 'ceasedate',
        'saldo1', 'saldo2', 'saldo3', 'saldo4', 'saldo5',
        'faltas', 'licencias', 'descansos', 'saldo',
    ]
    keys_numericos = {
        'saldo1', 'saldo2', 'saldo3', 'saldo4', 'saldo5',
        'faltas', 'licencias', 'descansos', 'saldo',
    }

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout_report(cursor)
        cursor.execute(
            "EXEC sp_pr_saldovacaciones_web @company=?, @payrolltype=?, @date=?, @person=?, @cesados=?",
            (cia, payroll_type, fecha_dt, person, cesados),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            fila = []
            for key in keys_datos:
                val = r.get(key)
                if key in keys_numericos and val is not None:
                    try:
                        fila.append(float(val))
                    except Exception:
                        fila.append(_jsonable_value(val))
                else:
                    fila.append(_jsonable_value(val))
            resultado.append(fila)
        return jsonify({
            "headers": headers_es,
            "data": resultado,
            "meta": {
                "fecha": fecha_dt.strftime('%d/%m/%Y'),
                "anios": anios,
            },
        })
    except Exception as e:
        logging.exception("reporte_saldo_vacaciones_post")
        return jsonify({"error": _reporte_sql_error_message(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/asignacion-conceptos/listado', methods=['POST'])
@login_required
def api_asignacion_conceptos_listado():
    """sp_pr_listaasignacionconceptos_web: listado de conceptos asignados a trabajadores."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '').strip()
    period_raw = body.get('period')
    ps = str(period_raw).strip() if period_raw is not None else ''
    period = '0' if ps in ('', '0') else ps
    concept_raw = body.get('concept')
    cs = str(concept_raw).strip() if concept_raw is not None else ''
    concept = '0' if cs in ('', '0') else cs
    person_raw = body.get('person') or body.get('trabajador') or body.get('empleado') or ''
    ps_person = str(person_raw).strip() if person_raw is not None else ''
    person = '0' if ps_person in ('', '0') else ps_person
    nombre = str(body.get('nombre') or body.get('name') or '').strip()
    cesados = _normalize_cesados_telecredito(body.get('cesados'))
    frecuencytype = _normalize_tipo_concepto_asig(
        body.get('frecuencytype') or body.get('tipo_concepto') or body.get('tipoconcepto')
    )
    replicationunit = _normalize_replicationunit_asig(
        body.get('replicationunit') or body.get('unidad') or body.get('repunit')
    )

    if not cia:
        return jsonify({"error": "Seleccione compañía."}), 400
    if not payrolltype:
        return jsonify({"error": "Seleccione tipo de planilla."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listaasignacionconceptos_web "
            "@par_company=?, @par_payrolltype=?, @par_period=?, @par_concept=?, "
            "@par_person=?, @nombre=?, @cesados=?, @par_frecuencytype=?, @par_replicationunit=?",
            (cia, payrolltype, period, concept, person, nombre, cesados, frecuencytype, replicationunit),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        total_valor = 0.0
        for r in rows:
            conceptvalue = r.get('conceptvalue')
            try:
                conceptvalue_num = float(conceptvalue) if conceptvalue is not None else 0.0
            except Exception:
                conceptvalue_num = 0.0
            total_valor += conceptvalue_num
            resultado.append({
                "person": _jsonable_value(r.get('person')),
                "nombre": _jsonable_value(r.get('nombre')),
                "employeecode": _jsonable_value(r.get('employeecode')),
                "company": _jsonable_value(r.get('company')),
                "concept": _jsonable_value(r.get('concept')),
                "conceptname": _jsonable_value(r.get('conceptname')),
                "payrolltype": _jsonable_value(r.get('payrolltype')),
                "prperiodstart": _jsonable_value(r.get('prperiodstart')),
                "prperiodend": _jsonable_value(r.get('prperiodend')),
                "conceptvalue": conceptvalue_num,
                "conceptcurrency": _jsonable_value(r.get('conceptcurrency')),
                "flagapplyformula": _jsonable_value(r.get('flagapplyformula')),
                "flagfrecuencytype": _jsonable_value(r.get('flagfrecuencytype')),
                "costcenter": _jsonable_value(r.get('costcenter')),
                "costcentercode": _jsonable_value(r.get('costcentercode')),
                "project": _jsonable_value(r.get('project')),
                "comments": _jsonable_value(r.get('comments')),
                "tareo": _jsonable_value(r.get('tareo')),
                "xlastuser": _jsonable_value(r.get('xlastuser')),
                "xlastdate": _jsonable_value(r.get('xlastdate')),
            })
        return jsonify({
            "rows": resultado,
            "total": len(resultado),
            "total_valor": total_valor,
        })
    except Exception as e:
        logging.exception("api_asignacion_conceptos_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/asignacion-conceptos/detalle', methods=['POST'])
@login_required
def api_asignacion_conceptos_detalle():
    """sp_pr_obtenerasignacionconcepto_web: detalle para edición."""
    body = request.get_json(silent=True) or {}
    pk = _asignacion_concepto_pk_from_json(body)
    if not pk['company'] or not pk['person'] or not pk['concept'] or not pk['payrolltype'] or not pk['prperiodstart']:
        return jsonify({"error": "Faltan datos de la asignación."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_obtenerasignacionconcepto_web "
            "@par_company=?, @par_person=?, @par_concept=?, "
            "@par_payrolltype=?, @par_prperiodstart=?, @par_costcenter=?",
            (
                pk['company'], pk['person'], pk['concept'],
                pk['payrolltype'], pk['prperiodstart'], pk['costcenter'],
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        if not rows:
            return jsonify({"error": "No se encontró la asignación."}), 404
        return jsonify({"row": _asignacion_concepto_detalle_dict(rows[0])})
    except Exception as e:
        logging.exception("api_asignacion_conceptos_detalle")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/asignacion-conceptos/guardar', methods=['POST'])
@login_required
def api_asignacion_conceptos_guardar():
    """sp_pr_guardarasignacionconcepto_web: alta o actualización."""
    body = request.get_json(silent=True) or {}
    modo_raw = str(body.get('modo') or 'I').strip().upper()
    modo = 'U' if modo_raw == 'U' else 'I'
    pk = _asignacion_concepto_pk_from_json(body)

    if not pk['company'] or not pk['person'] or not pk['concept'] or not pk['payrolltype'] or not pk['prperiodstart']:
        return jsonify({"error": "Complete compañía, empleado, concepto, tipo planilla y periodo inicio."}), 400

    prperiodend_raw = body.get('prperiodend') or body.get('period_end')
    prperiodend = _normalize_pr_period(prperiodend_raw) if prperiodend_raw not in (None, '') else ''
    flagfrecuencytype = _normalize_flagfrecuencytype_asig(body.get('flagfrecuencytype'))
    if flagfrecuencytype == 'P':
        prperiodend = ''

    conceptvalue_raw = body.get('conceptvalue')
    if conceptvalue_raw is None or str(conceptvalue_raw).strip() == '':
        return jsonify({"error": "Indique el valor del concepto."}), 400
    try:
        conceptvalue = float(conceptvalue_raw)
    except Exception:
        return jsonify({"error": "Valor del concepto inválido."}), 400

    conceptcurrency = _normalize_conceptcurrency_asig(body.get('conceptcurrency'))
    flagapplyformula = _normalize_flagapplyformula_asig(body.get('flagapplyformula'))
    xlastuser = str(getattr(current_user, 'username', '') or '')[:20]

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_guardarasignacionconcepto_web "
            "@modo=?, @par_company=?, @par_person=?, @par_concept=?, @par_payrolltype=?, "
            "@par_prperiodstart=?, @par_costcenter=?, @par_prperiodend=?, @par_conceptvalue=?, "
            "@par_conceptcurrency=?, @par_flagapplyformula=?, @par_flagfrecuencytype=?, @xlastuser=?",
            (
                modo, pk['company'], pk['person'], pk['concept'], pk['payrolltype'],
                pk['prperiodstart'], pk['costcenter'], prperiodend or None, conceptvalue,
                conceptcurrency, flagapplyformula, flagfrecuencytype, xlastuser,
            ),
        )
        _drain_pyodbc_cursor(cursor)
        conn.commit()
        return jsonify({"ok": True, "modo": modo})
    except Exception as e:
        logging.exception("api_asignacion_conceptos_guardar")
        err = str(e)
        if 'RAISERROR' in err or '50000' in err:
            parts = err.split(']')
            if len(parts) > 1:
                err = parts[-1].strip(" ()'\"")
        return jsonify({"error": err}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


def _jsonable_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    return _jsonable_value(value)


def _vacacion_empleado_dict(row):
    if not row:
        return {}
    return {
        'person': _jsonable_value(row.get('person')),
        'codigo': _jsonable_value(row.get('codigo')),
        'nombre': _jsonable_value(row.get('nombre')),
        'documento': _jsonable_value(row.get('documento')),
        'fechaingreso': _jsonable_value(row.get('fechaingreso')),
        'payrolltype': _jsonable_value(row.get('payrolltype')),
        'tipoplanilla': _jsonable_value(row.get('tipoplanilla')),
    }


def _vacacion_resumen_dict(row):
    if not row:
        return {'dias_acumulados': 0, 'dias_gozados': 0, 'dias_pendientes': 0}
    return {
        'dias_acumulados': int(row.get('dias_acumulados') or 0),
        'dias_gozados': int(row.get('dias_gozados') or 0),
        'dias_pendientes': int(row.get('dias_pendientes') or 0),
    }


def _vacacion_periodo_dict(row):
    return {
        'line': int(row.get('line') or 0),
        'controlyear': _jsonable_value(row.get('controlyear')),
        'periodo': _jsonable_value(row.get('periodo')),
        'dias': int(row.get('dias') or 0),
        'dias_adquiridos': int(row.get('dias_adquiridos') or 0),
        'consumidos': int(row.get('consumidos') or 0),
        'pendientes': int(row.get('pendientes') or 0),
        'pagados': int(row.get('pagados') or 0),
        'por_pagar': int(row.get('por_pagar') or 0),
        'inicio_provision': _jsonable_value(row.get('inicio_provision')),
        'inicio_derecho': _jsonable_value(row.get('inicio_derecho')),
        'fin_derecho': _jsonable_value(row.get('fin_derecho')),
        'limite_sin_indemnizacion': _jsonable_value(row.get('limite_sin_indemnizacion')),
        'status': _jsonable_value(row.get('status')),
        'estado_texto': _jsonable_value(row.get('estado_texto')),
        'usuario': _jsonable_value(row.get('usuario')),
        'fecha_modificacion': _jsonable_datetime(row.get('fecha_modificacion')),
    }


def _vacacion_detalle_dict(row):
    prperiod_raw = row.get('prperiod')
    consumo = row.get('consumo_efectivo') or _format_prperiod_mes(prperiod_raw)
    return {
        'line': int(row.get('line') or 0),
        'secuence': int(row.get('secuence') or 0),
        'prperiod': _jsonable_value(prperiod_raw),
        'consumo_efectivo': _jsonable_value(consumo),
        'fecha_inicio': _jsonable_value(row.get('fecha_inicio')),
        'fecha_fin': _jsonable_value(row.get('fecha_fin')),
        'dias': int(row.get('dias') or 0),
        'vacationtype': _jsonable_value(row.get('vacationtype')),
        'tipo_texto': _jsonable_value(row.get('tipo_texto')),
        'usuario': _jsonable_value(row.get('usuario')),
        'fecha_modificacion': _jsonable_datetime(row.get('fecha_modificacion')),
    }


@app.route('/api/vacaciones/trabajadores', methods=['POST'])
@login_required
def api_vacaciones_trabajadores():
    """sp_pr_vacaciones_listar_trabajadores_web: trabajadores activos para panel izquierdo."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '0').strip() or '0'
    busqueda = str(body.get('busqueda') or body.get('nombre') or body.get('q') or '').strip()

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_vacaciones_listar_trabajadores_web @company=?, @payrolltype=?, @busqueda=?",
            (cia, payrolltype, busqueda),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            resultado.append({
                'person': _jsonable_value(r.get('person')),
                'codigo': _jsonable_value(r.get('codigo')),
                'nombre': _jsonable_value(r.get('nombre')),
                'documento': _jsonable_value(r.get('documento')),
                'fechaingreso': _jsonable_value(r.get('fechaingreso')),
                'payrolltype': _jsonable_value(r.get('payrolltype')),
                'tipoplanilla': _jsonable_value(r.get('tipoplanilla')),
            })
        return jsonify({"rows": resultado, "total": len(resultado)})
    except Exception as e:
        logging.exception("api_vacaciones_trabajadores")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/vacaciones/obtener', methods=['POST'])
@login_required
def api_vacaciones_obtener():
    """sp_pr_vacaciones_obtener_trabajador_web: periodos y detalle de utilización."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    person = str(body.get('person') or '').strip()

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not person:
        return jsonify({"error": "Seleccione un trabajador."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_vacaciones_obtener_trabajador_web @company=?, @person=?",
            (cia, person),
        )
        sets = _dicts_collect_nonempty_resultsets(cursor, max_sets=6)
        empleado = _vacacion_empleado_dict(sets[0][0] if len(sets) > 0 and sets[0] else None)
        resumen = _vacacion_resumen_dict(sets[1][0] if len(sets) > 1 and sets[1] else None)
        periodos = [_vacacion_periodo_dict(r) for r in (sets[2] if len(sets) > 2 else [])]
        detalle = [_vacacion_detalle_dict(r) for r in (sets[3] if len(sets) > 3 else [])]
        return jsonify({
            "empleado": empleado,
            "resumen": resumen,
            "periodos": periodos,
            "detalle": detalle,
        })
    except Exception as e:
        logging.exception("api_vacaciones_obtener")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/vacaciones/guardar-detalle', methods=['POST'])
@login_required
def api_vacaciones_guardar_detalle():
    """sp_pr_vacaciones_guardar_detalle_web: alta de utilización de vacaciones."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    person = str(body.get('person') or '').strip()
    line_raw = body.get('line')
    prperiod = _normalize_pr_period_vacacion(body.get('prperiod') or body.get('consumo_efectivo') or '')
    vacationtype = str(body.get('vacationtype') or body.get('tipo') or 'D').strip().upper()[:1] or 'D'
    xlastuser = _xlastuser_id()

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not person:
        return jsonify({"error": "Seleccione un trabajador."}), 400
    if line_raw is None or str(line_raw).strip() == '':
        return jsonify({"error": "Seleccione un periodo vacacional."}), 400
    if not prperiod:
        return jsonify({"error": "Seleccione el periodo de consumo efectivo."}), 400

    try:
        line = int(line_raw)
    except Exception:
        return jsonify({"error": "Periodo vacacional inválido."}), 400

    fecha_inicio = _sql_date_str_param(body.get('fecha_inicio') or body.get('datebegin'))
    fecha_fin = _sql_date_str_param(body.get('fecha_fin') or body.get('dateend'))
    if not fecha_inicio or not fecha_fin:
        return jsonify({"error": "Indique fecha de inicio y término."}), 400
    if fecha_inicio > fecha_fin:
        return jsonify({"error": "La fecha de término no puede ser anterior a la de inicio."}), 400

    days_raw = body.get('dias') or body.get('days')
    days_param = None
    if days_raw is not None and str(days_raw).strip() != '':
        try:
            days_param = int(days_raw)
        except Exception:
            return jsonify({"error": "Días calculados inválidos."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_vacaciones_guardar_detalle_web "
            "@company=?, @person=?, @line=?, @prperiod=?, @datebegin=?, @dateend=?, "
            "@days=?, @vacationtype=?, @xlastuser=?",
            (
                cia, person, line, prperiod, fecha_inicio, fecha_fin,
                days_param, vacationtype, xlastuser,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        _drain_pyodbc_cursor(cursor)
        conn.commit()
        return jsonify({"ok": True, "row": rows[0] if rows else {}})
    except Exception as e:
        logging.exception("api_vacaciones_guardar_detalle")
        err = str(e)
        if 'RAISERROR' in err or '50000' in err:
            parts = err.split(']')
            if len(parts) > 1:
                err = parts[-1].strip(" ()'\"")
        return jsonify({"error": err}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/vacaciones/eliminar-detalle', methods=['POST'])
@login_required
def api_vacaciones_eliminar_detalle():
    """sp_pr_vacaciones_eliminar_detalle_web: elimina utilización de vacaciones."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    person = str(body.get('person') or '').strip()
    line_raw = body.get('line')
    secuence_raw = body.get('secuence')
    xlastuser = _xlastuser_id()

    if not cia or not person:
        return jsonify({"error": "Faltan datos del trabajador."}), 400
    if line_raw is None or secuence_raw is None:
        return jsonify({"error": "Seleccione un registro de utilización."}), 400

    try:
        line = int(line_raw)
        secuence = int(secuence_raw)
    except Exception:
        return jsonify({"error": "Registro de utilización inválido."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_vacaciones_eliminar_detalle_web "
            "@company=?, @person=?, @line=?, @secuence=?, @xlastuser=?",
            (cia, person, line, secuence, xlastuser),
        )
        _drain_pyodbc_cursor(cursor)
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logging.exception("api_vacaciones_eliminar_detalle")
        err = str(e)
        if 'RAISERROR' in err or '50000' in err:
            parts = err.split(']')
            if len(parts) > 1:
                err = parts[-1].strip(" ()'\"")
        return jsonify({"error": err}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/asignacion-conceptos/eliminar', methods=['POST'])
@login_required
def api_asignacion_conceptos_eliminar():
    """sp_pr_eliminarasignacionconcepto_web: elimina asignación por clave."""
    body = request.get_json(silent=True) or {}
    pk = _asignacion_concepto_pk_from_json(body)
    if not pk['company'] or not pk['person'] or not pk['concept'] or not pk['payrolltype'] or not pk['prperiodstart']:
        return jsonify({"error": "Faltan datos de la asignación."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_eliminarasignacionconcepto_web "
            "@par_company=?, @par_person=?, @par_concept=?, "
            "@par_payrolltype=?, @par_prperiodstart=?, @par_costcenter=?",
            (
                pk['company'], pk['person'], pk['concept'],
                pk['payrolltype'], pk['prperiodstart'], pk['costcenter'],
            ),
        )
        _drain_pyodbc_cursor(cursor)
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logging.exception("api_asignacion_conceptos_eliminar")
        err = str(e)
        if 'RAISERROR' in err or '50000' in err:
            parts = err.split(']')
            if len(parts) > 1:
                err = parts[-1].strip(" ()'\"")
        return jsonify({"error": err}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/reportes/listado-pagos', methods=['POST'])
@login_required
def api_reporte_listado_pagos():
    """sp_pr_reportelistadopagos_web: listado de pagos por trabajador (filtros Telecrédito sin fecha de pago)."""
    body = request.get_json(silent=True) or {}
    p = _telecredito_params_from_json(body)
    err = _telecredito_validar_params(p)
    if err:
        return jsonify({"error": err}), 400

    cesados = _normalize_cesados_telecredito(body.get('cesados'))
    salarybank = str(body.get('salarybank') or body.get('salary_bank') or '0').strip() or '0'

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_reportelistadopagos_web "
            "@par_company=?, @par_currency=?, @par_concept=?, "
            "@par_payrolltype=?, @par_period=?, @par_processtype=?, @cesados=?, @salarybank=?",
            (
                p['cia'], p['currency'], p['concept'], p['payrolltype'],
                p['period'], p['processtype'], cesados, salarybank,
            ),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            importe = r.get('importe')
            try:
                importe_num = float(importe) if importe is not None else 0.0
            except Exception:
                importe_num = 0.0
            resultado.append({
                "employeecode": _jsonable_value(r.get('employeecode')),
                "nombre": _jsonable_value(r.get('nombre')),
                "banco": _jsonable_value(r.get('banco')),
                "obra": _jsonable_value(r.get('obra')),
                "costcenter": _jsonable_value(r.get('costcenter')),
                "cuenta": _jsonable_value(r.get('cuenta')),
                "moneda": _jsonable_value(r.get('moneda')),
                "importe": importe_num,
            })
        return jsonify({"rows": resultado, "total": len(resultado)})
    except Exception as e:
        logging.exception("api_reporte_listado_pagos")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/reporte_log_calculo', methods=['POST'])
@login_required
def reporte_log_calculo_post():
    """sp_pr_reportelog_calculo_web @cia, @payrolltype, @process, @period, @person."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or '').strip()
    payroll_type = str(body.get('payroll_type') or body.get('payrolltype') or '').strip()
    process = str(body.get('process') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    person = str(body.get('person') or '0').strip() or '0'

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not payroll_type or not process or not period:
        return jsonify({"error": "Debe indicar tipo de planilla, proceso y periodo."}), 400

    headers_es = [
        'Código',
        'Nombre',
        'Fecha',
        'Concepto',
        'Cód. fórmula',
        'Importe',
        'Tipo concepto',
        'Tipo cálculo',
        'Insertar',
        'Afecto 5ta',
        'Afecto AFP',
        'Periodo inicio',
    ]
    keys_datos = [
        'person',
        'name',
        'fecha',
        'concepto',
        'formulacode',
        'importe',
        'tipoconcepto',
        'tipocalculo',
        'flaginsertar',
        'flagafecto5ta',
        'flagafectoafp',
        'periodbegin',
    ]

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_reportelog_calculo_web @cia=?, @payrolltype=?, @process=?, @period=?, @person=?",
            (cia, payroll_type, process, period, person),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            fila = []
            for key in keys_datos:
                val = r.get(key)
                if key == 'importe' and val is not None:
                    try:
                        fila.append(float(val))
                    except Exception:
                        fila.append(_jsonable_value(val))
                else:
                    fila.append(_jsonable_value(val))
            resultado.append(fila)
        return jsonify({"headers": headers_es, "data": resultado})
    except Exception as e:
        logging.exception("reporte_log_calculo_post")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/reporte_descansos_medicos_detalle', methods=['POST'])
@login_required
def reporte_descansos_medicos_detalle_post():
    """sp_pr_reportesdescansos_medicos_web @cia, @payrolltype, @period, @person, @medicalresttype."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or '').strip()
    payroll_type = str(body.get('payroll_type') or body.get('payrolltype') or '').strip()
    period_raw = body.get('period')
    ps = str(period_raw).strip() if period_raw is not None else ''
    if ps == '' or ps == '0':
        period = '0'
    else:
        period = _normalize_pr_period(period_raw)
    person = str(body.get('person') or '0').strip() or '0'
    mrt_raw = body.get('medicalresttype')
    mrs = str(mrt_raw).strip() if mrt_raw is not None else ''
    medicalresttype = '0' if mrs == '' or mrs == '0' else mrs

    if not cia:
        return jsonify({"error": "Seleccione una compañía."}), 400
    if not payroll_type:
        return jsonify({"error": "Debe indicar tipo de planilla."}), 400

    headers_es = [
        'Periodo',
        'Código',
        'Nombre',
        'Fecha inicio',
        'Fecha fin',
        'Días',
        'Tipo de descanso',
        'CITT',
    ]
    keys_datos = ['person', 'name', 'datebegin', 'dateend', 'days', 'description', 'citt']

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_reportesdescansos_medicos_web @cia=?, @payrolltype=?, @period=?, @person=?, @medicalresttype=?",
            (cia, payroll_type, period, person, medicalresttype),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            fila = [_fmt_periodo_yyyy_mm(r.get('prperiod'))]
            for key in keys_datos:
                val = r.get(key)
                if key == 'days' and val is not None:
                    try:
                        fila.append(int(round(float(val))))
                    except Exception:
                        fila.append(_jsonable_value(val))
                else:
                    fila.append(_jsonable_value(val))
            resultado.append(fila)
        return jsonify({"headers": headers_es, "data": resultado})
    except Exception as e:
        logging.exception("reporte_descansos_medicos_detalle_post")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


# ==========================================
# API Aperturar periodos — control de procesos
# ==========================================


def _format_prperiod_display(period_raw):
    """Formatea PRPeriod yyyymmdd como yyyy-mm-dd para la UI."""
    s = _normalize_pr_period(period_raw)
    if len(s) >= 8 and s[:8].isdigit():
        return f'{s[0:4]}-{s[4:6]}-{s[6:8]}'
    return str(period_raw or '').strip()


def _fecha_hora_tabla_json(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%d/%m/%Y %H:%M')
    if isinstance(val, date):
        return val.strftime('%d/%m/%Y')
    return str(val).strip()


@app.route('/api/aperturar-periodos/listado', methods=['POST'])
@login_required
def api_aperturar_periodos_listado():
    """sp_pr_listaprocesscontrol_apertura_web: procesos activos y pendientes de control."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '').strip()
    if not cia:
        return jsonify({"error": "Seleccione compañía."}), 400
    if not payrolltype:
        return jsonify({"error": "Seleccione tipo de planilla."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_listaprocesscontrol_apertura_web @cia=?, @payrolltype=?",
            (cia, payrolltype),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        resultado = []
        for r in rows:
            prperiod_raw = r.get('prperiod')
            prperiod = _normalize_pr_period(prperiod_raw) if prperiod_raw not in (None, '') else ''
            resultado.append({
                "processtype": _jsonable_value(r.get('processtype')),
                "description": _jsonable_value(r.get('description')),
                "company": _jsonable_value(r.get('company')),
                "payrolltype": _jsonable_value(r.get('payrolltype')),
                "prperiod": prperiod,
                "prperiod_display": _format_prperiod_display(prperiod) if prperiod else '',
                "processdate": _fecha_hora_tabla_json(r.get('processdate')),
                "status": _jsonable_value(r.get('status')),
                "statusdesc": _jsonable_value(r.get('statusdesc')),
            })
        return jsonify({"rows": resultado, "total": len(resultado)})
    except Exception as e:
        logging.exception("api_aperturar_periodos_listado")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/aperturar-periodos/periodos')
@login_required
def api_aperturar_periodos_periodos():
    """sp_pr_selectorperiodos_apertura_web: periodos configurados en PR_Period."""
    cia = request.args.get('cia', '').strip()
    payrolltype = request.args.get('payrolltype', '').strip()
    if not cia or not payrolltype:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorperiodos_apertura_web @cia=?, @payrolltype=?",
            (cia, payrolltype),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        data = []
        for r in rows:
            pid = _normalize_pr_period(r.get('prperiod'))
            if not pid:
                continue
            txt = str(r.get('description') or '').strip() or _format_prperiod_display(pid)
            data.append({"id": pid, "text": txt})
        return jsonify(data)
    except Exception:
        logging.exception("api_aperturar_periodos_periodos")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/aperturar-periodos/periodo-sugerido')
@login_required
def api_aperturar_periodos_periodo_sugerido():
    """sp_pr_selectorperiodoactivo_planilla_web: MAX periodo con status A/G."""
    cia = request.args.get('cia', '').strip()
    payrolltype = request.args.get('payrolltype', '').strip()
    if not cia or not payrolltype:
        return jsonify({"prperiod": ""})
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorperiodoactivo_planilla_web @cia=?, @payrolltype=?",
            (cia, payrolltype),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        prperiod = ''
        if rows:
            prperiod = _normalize_pr_period(rows[0].get('prperiod'))
        return jsonify({"prperiod": prperiod})
    except Exception:
        logging.exception("api_aperturar_periodos_periodo_sugerido")
        return jsonify({"prperiod": ""})
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


def _aperturar_periodos_procesos_from_body(body):
    raw = body.get('processtypes') or body.get('procesos') or body.get('process_types') or []
    if isinstance(raw, str):
        raw = [p.strip() for p in raw.split(',') if p.strip()]
    if not isinstance(raw, (list, tuple)):
        return []
    seen = set()
    out = []
    for item in raw:
        s = str(item or '').strip()
        if s and s not in seen:
            seen.add(s)
            out.append(s)
    return out


@app.route('/api/aperturar-periodos/aperturar', methods=['POST'])
@login_required
def api_aperturar_periodos_aperturar():
    """Apertura masiva de periodo por tipo de proceso seleccionado."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period') or body.get('prperiod'))
    processtypes = _aperturar_periodos_procesos_from_body(body)
    xlastuser = str(getattr(current_user, 'username', '') or '')[:20]

    if not cia:
        return jsonify({"error": "Seleccione compañía."}), 400
    if not payrolltype:
        return jsonify({"error": "Seleccione tipo de planilla."}), 400
    if not period:
        return jsonify({"error": "Seleccione el periodo a aperturar."}), 400
    if not processtypes:
        return jsonify({"error": "Seleccione al menos un tipo de proceso."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        for pt in processtypes:
            cursor.execute(
                "EXEC sp_pr_aperturarperiodo_proceso_web "
                "@cia=?, @payrolltype=?, @processtype=?, @period=?, @xlastuser=?",
                (cia, payrolltype, pt, period, xlastuser),
            )
            _drain_all_cursor_resultsets(cursor)
        conn.commit()
        return jsonify({
            "ok": True,
            "message": f"Periodo { _format_prperiod_display(period) } aperturado en {len(processtypes)} proceso(s).",
            "procesados": len(processtypes),
        })
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        logging.exception("api_aperturar_periodos_aperturar")
        msg = str(e)
        if '50001' in msg or 'RAISERROR' in msg.upper():
            msg = msg.split(']')[-1].strip() if ']' in msg else msg
        return jsonify({"error": msg or "No se pudo aperturar el periodo."}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/aperturar-periodos/cerrar', methods=['POST'])
@login_required
def api_aperturar_periodos_cerrar():
    """Cierra periodo activo (A/G → C) en los procesos seleccionados."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or body.get('company') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '').strip()
    processtypes = _aperturar_periodos_procesos_from_body(body)
    xlastuser = str(getattr(current_user, 'username', '') or '')[:20]

    if not cia:
        return jsonify({"error": "Seleccione compañía."}), 400
    if not payrolltype:
        return jsonify({"error": "Seleccione tipo de planilla."}), 400
    if not processtypes:
        return jsonify({"error": "Seleccione al menos un tipo de proceso."}), 400

    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        for pt in processtypes:
            cursor.execute(
                "EXEC sp_pr_cerrarperiodo_proceso_web "
                "@cia=?, @payrolltype=?, @processtype=?, @xlastuser=?",
                (cia, payrolltype, pt, xlastuser),
            )
            _drain_all_cursor_resultsets(cursor)
        conn.commit()
        return jsonify({
            "ok": True,
            "message": f"Periodo cerrado en {len(processtypes)} proceso(s).",
            "procesados": len(processtypes),
        })
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        logging.exception("api_aperturar_periodos_cerrar")
        return jsonify({"error": str(e) or "No se pudo cerrar el periodo."}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


# ==========================================
# API Procesar planilla (cálculo) — SPs dedicados
# ==========================================


@app.route('/api/procesar-planilla/procesos-calculo', methods=['POST'])
@login_required
def api_procesar_planilla_procesos():
    """sp_pr_selectorprocesoscalculo_web @cia, @payrolltype → PROCESSTYPE, DESCRIPTION."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or '').strip()
    payrolltype = str(body.get('payrolltype') or '').strip()
    if not cia or not payrolltype:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorprocesoscalculo_web @cia=?, @payrolltype=?",
            (cia, payrolltype),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        proc_names = {}
        try:
            cursor.execute(
                """
                SELECT ProcessType, ProcedureName
                FROM PR_ProcessType
                WHERE Company = ?
                """,
                (cia,),
            )
            for pr in _dicts_first_nonempty_resultset(cursor):
                pt = str(pr.get('ProcessType') or pr.get('processtype') or '').strip()
                pn = str(pr.get('ProcedureName') or pr.get('procedurename') or '').strip()
                if pt:
                    proc_names[pt] = pn
        except Exception:
            logging.debug('ProcedureName no disponible en PR_ProcessType', exc_info=True)

        data = [
            {
                "id": str(r.get("processtype") or "").strip(),
                "text": str(r.get("description") or "").strip(),
                "procedurename": proc_names.get(str(r.get("processtype") or "").strip(), ''),
            }
            for r in rows
            if str(r.get("processtype") or "").strip()
        ]
        return jsonify(data)
    except Exception:
        logging.exception("api_procesar_planilla_procesos")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/procesar-planilla/periodos-calculo')
@login_required
def api_procesar_planilla_periodos_list():
    """sp_pr_selectorperiodocalculo_web @cia, @processtype → PRPERIOD, description (lista ordenada en SP)."""
    cia = request.args.get('cia', '').strip()
    processtype = request.args.get('processtype', '').strip()
    if not cia or not processtype:
        return jsonify([])
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_selectorperiodocalculo_web @cia=?, @processtype=?",
            (cia, processtype),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        data = []
        for r in rows:
            raw = r.get("prperiod")
            pid = _normalize_pr_period(raw) or str(raw or "").strip()
            if not pid:
                continue
            data.append(
                {
                    "id": pid,
                    "text": str(r.get("description") or "").strip(),
                }
            )
        return jsonify(data)
    except Exception:
        logging.exception("api_procesar_planilla_periodos_list")
        return jsonify([])
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


def _fecha_tabla_json(val):
    """Serializa fecha/datetime para columnas de listados (JSON)."""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%d/%m/%Y')
    if isinstance(val, date):
        return val.strftime('%d/%m/%Y')
    return str(val).strip()


def _validar_calculo_planilla_mensajes(cursor, cia, payrolltype, processtype, period):
    """sp_pr_validar_calculo_web → lista de mensajes para el panel de validaciones."""
    cursor.execute(
        "EXEC sp_pr_validar_calculo_web "
        "@cia=?, @payrolltype=?, @processtype=?, @period=?",
        (cia, payrolltype, processtype, period),
    )
    rows = _dicts_first_nonempty_resultset(cursor)
    mensajes = []
    for r in rows:
        person = str(r.get('person') or '').strip()
        name = str(r.get('name') or '').strip()
        obs = str(r.get('observacion') or '').strip()
        if not obs:
            continue
        if person or name:
            etiqueta = ' — '.join([p for p in [person, name] if p])
            mensajes.append(f'{etiqueta}: {obs}')
        else:
            mensajes.append(obs)
    return mensajes


@app.route('/api/procesar-planilla/trabajadores-calculo', methods=['POST'])
@login_required
def api_procesar_planilla_trabajadores():
    """sp_pr_calcularplanillas_web @cia, @payrolltype, @processtype, @period, @cesados, @repunit."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or '').strip()
    payrolltype = str(body.get('payrolltype') or '').strip()
    processtype = str(body.get('processtype') or body.get('proceso') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    cesados = str(body.get('cesados') or 'T').strip().upper()
    if cesados not in ('T', 'Y', 'N'):
        cesados = 'T'
    repunit = str(body.get('repunit') or body.get('unidad') or '0').strip()
    if not repunit:
        repunit = '0'
    if len(repunit) > 20:
        repunit = repunit[:20]
    if not cia or not payrolltype or not processtype or not period:
        return jsonify({"error": "Faltan compañía, tipo de planilla, proceso o periodo."}), 400
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "EXEC sp_pr_calcularplanillas_web "
            "@cia=?, @payrolltype=?, @processtype=?, @period=?, @cesados=?, @repunit=?",
            (cia, payrolltype, processtype, period, cesados, repunit),
        )
        rows = _dicts_first_nonempty_resultset(cursor)
        trabajadores = [
            {
                "person": str(r.get("person") or "").strip(),
                "name": str(r.get("name") or "").strip(),
                "entrydate": _fecha_tabla_json(r.get("entrydate")),
                "ceasedate": _fecha_tabla_json(r.get("ceasedate")),
                "calculationdate": _fecha_hora_tabla_json(r.get("calculationdate")),
            }
            for r in rows
            if str(r.get("person") or "").strip()
        ]
        return jsonify(trabajadores)
    except Exception as e:
        logging.exception("api_procesar_planilla_trabajadores")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/procesar-planilla/validar-calculo', methods=['POST'])
@login_required
def api_procesar_planilla_validar_calculo():
    """sp_pr_validar_calculo_web: validaciones post-cálculo."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '').strip()
    processtype = str(body.get('processtype') or body.get('proceso') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    if not cia or not payrolltype or not processtype or not period:
        return jsonify({"error": "Faltan compañía, tipo de planilla, proceso o periodo."}), 400
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        validaciones = _validar_calculo_planilla_mensajes(
            cursor, cia, payrolltype, processtype, period
        )
        return jsonify({
            "validaciones": validaciones,
            "total": len(validaciones),
        })
    except Exception as e:
        logging.exception("api_procesar_planilla_validar_calculo")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/api/procesar-planilla/eliminar-calculo', methods=['POST'])
@login_required
def api_procesar_planilla_eliminar_calculo():
    """sp_pr_eliminar_calculo_planilla_web: elimina cálculo de trabajadores seleccionados."""
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or '').strip()
    payrolltype = str(body.get('payrolltype') or body.get('payroll_type') or '').strip()
    processtype = str(body.get('processtype') or body.get('proceso') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    seleccionados = body.get('trabajadores')

    if not isinstance(seleccionados, list) or len(seleccionados) == 0:
        return jsonify({"error": "Debe seleccionar al menos un trabajador."}), 400
    if not cia or not payrolltype or not processtype or not period:
        return jsonify({"error": "Faltan compañía, tipo de planilla, proceso o periodo."}), 400

    conn = None
    exitos = 0
    errores = []
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        for person_id in seleccionados:
            pid = str(person_id).strip()
            if not pid:
                continue
            try:
                cursor.execute(
                    "EXEC sp_pr_eliminar_calculo_planilla_web "
                    "@company=?, @payrolltype=?, @processtype=?, @period=?, @person=?",
                    (cia, payrolltype, processtype, period, pid),
                )
                _drain_pyodbc_cursor(cursor)
                conn.commit()
                exitos += 1
            except Exception as e_individual:
                try:
                    conn.rollback()
                except Exception:
                    pass
                err = str(e_individual)
                if 'RAISERROR' in err or '50000' in err:
                    parts = err.split(']')
                    if len(parts) > 1:
                        err = parts[-1].strip(" ()'\"")
                errores.append(f'{pid}: {err}')
                logging.warning('api_procesar_planilla_eliminar_calculo persona %s: %s', pid, e_individual)

        if errores and exitos == 0:
            return jsonify({"error": '; '.join(errores), "exitos": exitos, "errores": errores}), 500
        return jsonify({
            "ok": True,
            "exitos": exitos,
            "errores": errores,
            "mensaje": "Proceso concluido." if not errores else f"Proceso concluido con {len(errores)} error(es).",
        })
    except Exception as e:
        logging.exception("api_procesar_planilla_eliminar_calculo")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/ejecutar_calculo_planilla', methods=['POST'])
@login_required
def ejecutar_calculo_planilla():
    """
    Resuelve el SP en PR_ProcessType (ProcedureName) y lo ejecuta por cada person.
    Orden de parámetros del CALL: cia, payroll_type, processtype, period, person, user_id, tc.
    """
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    processtype = str(body.get('processtype') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    seleccionados = body.get('trabajadores')

    if not isinstance(seleccionados, list) or len(seleccionados) == 0:
        return jsonify({'error': 'Debe enviar una lista no vacía de trabajadores (person).'}), 400
    if not cia or not processtype or not payroll_type or not period:
        return jsonify({'error': 'Faltan compañía, tipo de planilla, proceso o periodo.'}), 400

    try:
        user_id = current_user.id
    except AttributeError:
        return jsonify({'error': 'Usuario no identificado.'}), 401

    tc = 3.0
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        _set_cursor_timeout_payroll(cursor)
        cursor.execute(
            """
            SELECT ProcedureName, Description
            FROM PR_ProcessType
            WHERE ProcessType = ? AND Company = ?
            """,
            (processtype, cia),
        )
        row = cursor.fetchone()
        proc_raw = None
        proceso_desc = processtype
        if row:
            proc_raw = getattr(row, 'ProcedureName', None)
            desc_raw = getattr(row, 'Description', None)
            if proc_raw is None and len(row) > 0:
                proc_raw = row[0]
            if desc_raw is None and len(row) > 1:
                desc_raw = row[1]
            if desc_raw is not None and str(desc_raw).strip():
                proceso_desc = str(desc_raw).strip()
        sp_name = _sanitize_dynamic_procedure_name(proc_raw)
        if not sp_name:
            return jsonify(
                {
                    'error': (
                        f'El proceso "{proceso_desc}" no tiene un procedimiento de cálculo '
                        f'configurado (ProcedureName).'
                    )
                }
            ), 400

        _drain_pyodbc_cursor(cursor)

        exitos = 0
        errores = []
        call_sql = f'{{CALL {sp_name} (?, ?, ?, ?, ?, ?, ?)}}'

        for person_id in seleccionados:
            pid = str(person_id).strip()
            if not pid:
                continue
            try:
                cursor.execute(
                    call_sql,
                    (cia, payroll_type, processtype, period, pid, user_id, tc),
                )
                _drain_pyodbc_cursor(cursor)
                conn.commit()
                exitos += 1
            except Exception as e_individual:
                if _is_transient_sql_error(e_individual):
                    logging.warning(
                        'ejecutar_calculo_planilla persona %s: error transitorio; reintentando 1 vez',
                        pid,
                    )
                    try:
                        try:
                            conn.close()
                        except Exception:
                            pass
                        conn = get_db_connection()
                        cursor = conn.cursor()
                        _set_cursor_timeout_payroll(cursor)
                        cursor.execute(
                            call_sql,
                            (cia, payroll_type, processtype, period, pid, user_id, tc),
                        )
                        _drain_pyodbc_cursor(cursor)
                        conn.commit()
                        exitos += 1
                        continue
                    except Exception as e_retry:
                        try:
                            conn.rollback()
                        except Exception:
                            pass
                        errores.append(f'Error en {pid}: {e_retry}')
                        logging.warning(
                            'ejecutar_calculo_planilla persona %s fallo en reintento: %s',
                            pid,
                            e_retry,
                        )
                else:
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    errores.append(f'Error en {pid}: {e_individual}')
                    logging.warning('ejecutar_calculo_planilla persona %s: %s', pid, e_individual)

        validaciones = []
        if exitos > 0:
            try:
                validaciones = _validar_calculo_planilla_mensajes(
                    cursor, cia, payroll_type, processtype, period
                )
            except Exception:
                logging.exception('validar_calculo_planilla tras ejecutar_calculo_planilla')

        status = 'success' if not errores else 'partial'
        n_errores = len(errores)
        message = f'Proceso terminado. Éxitos: {exitos}, Errores: {n_errores}.'
        return jsonify(
            {
                'status': status,
                'message': message,
                'exitos': exitos,
                'errores': n_errores,
                'procesados': exitos + n_errores,
                'detalles': errores,
                'validaciones': validaciones,
            }
        )
    except Exception as e:
        logging.exception('ejecutar_calculo_planilla')
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


@app.route('/ejecutar_calculo_streaming', methods=['POST'])
@login_required
def ejecutar_calculo_streaming():
    """
    Mismo orquestado que /ejecutar_calculo_planilla pero emite eventos SSE (text/event-stream)
    tras cada trabajador: data: {"progreso","actual","total"} y al final data: {"done",...}.
    """
    ensure_user_session()
    body = request.get_json(silent=True) or {}
    cia = str(body.get('cia') or session.get('company') or '').strip()
    processtype = str(body.get('processtype') or '').strip()
    payroll_type = str(body.get('payroll_type') or '').strip()
    period = _normalize_pr_period(body.get('period'))
    seleccionados = body.get('trabajadores')

    if not isinstance(seleccionados, list) or len(seleccionados) == 0:
        return jsonify({'error': 'Debe enviar una lista no vacía de trabajadores (person).'}), 400
    if not cia or not processtype or not payroll_type or not period:
        return jsonify({'error': 'Faltan compañía, tipo de planilla, proceso o periodo.'}), 400

    try:
        user_id = current_user.id
    except AttributeError:
        return jsonify({'error': 'Usuario no identificado.'}), 401

    lista = [str(x).strip() for x in seleccionados if str(x).strip()]
    total = len(lista)
    if total == 0:
        return jsonify({'error': 'No hay IDs de trabajador válidos en la lista.'}), 400

    tc = 3.0

    def generar_progreso():
        conn = None
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            _set_cursor_timeout_payroll(cursor)
            cursor.execute(
                """
                SELECT ProcedureName, Description
                FROM PR_ProcessType
                WHERE ProcessType = ? AND Company = ?
                """,
                (processtype, cia),
            )
            row = cursor.fetchone()
            proc_raw = None
            proceso_desc = processtype
            if row:
                proc_raw = getattr(row, 'ProcedureName', None)
                desc_raw = getattr(row, 'Description', None)
                if proc_raw is None and len(row) > 0:
                    proc_raw = row[0]
                if desc_raw is None and len(row) > 1:
                    desc_raw = row[1]
                if desc_raw is not None and str(desc_raw).strip():
                    proceso_desc = str(desc_raw).strip()
            sp_name = _sanitize_dynamic_procedure_name(proc_raw)
            if not sp_name:
                yield (
                    'data: '
                    + json.dumps(
                        {
                            'error': (
                                f'El proceso "{proceso_desc}" no tiene un procedimiento de cálculo '
                                f'configurado (ProcedureName).'
                            )
                        }
                    )
                    + '\n\n'
                )
                return

            _drain_pyodbc_cursor(cursor)

            exitos = 0
            errores = []
            call_sql = f'{{CALL {sp_name} (?, ?, ?, ?, ?, ?, ?)}}'

            for index, pid in enumerate(lista):
                try:
                    cursor.execute(
                        call_sql,
                        (cia, payroll_type, processtype, period, pid, user_id, tc),
                    )
                    _drain_pyodbc_cursor(cursor)
                    conn.commit()
                    exitos += 1
                    evento = {
                        'progreso': int(((index + 1) / total) * 100),
                        'actual': index + 1,
                        'total': total,
                        'person': pid,
                    }
                except Exception as e_individual:
                    if _is_transient_sql_error(e_individual):
                        logging.warning(
                            'ejecutar_calculo_streaming persona %s: error transitorio; reintentando 1 vez',
                            pid,
                        )
                        try:
                            try:
                                conn.close()
                            except Exception:
                                pass
                            conn = get_db_connection()
                            cursor = conn.cursor()
                            _set_cursor_timeout_payroll(cursor)
                            cursor.execute(
                                call_sql,
                                (cia, payroll_type, processtype, period, pid, user_id, tc),
                            )
                            _drain_pyodbc_cursor(cursor)
                            conn.commit()
                            exitos += 1
                            evento = {
                                'progreso': int(((index + 1) / total) * 100),
                                'actual': index + 1,
                                'total': total,
                            }
                        except Exception as e_retry:
                            try:
                                conn.rollback()
                            except Exception:
                                pass
                            msg = str(e_retry)
                            errores.append(f'Error en {pid}: {msg}')
                            logging.warning(
                                'ejecutar_calculo_streaming persona %s fallo en reintento: %s',
                                pid,
                                e_retry,
                            )
                            evento = {
                                'progreso': int(((index + 1) / total) * 100),
                                'actual': index + 1,
                                'total': total,
                                'detalle': msg,
                                'person': pid,
                            }
                    else:
                        try:
                            conn.rollback()
                        except Exception:
                            pass
                        msg = str(e_individual)
                        errores.append(f'Error en {pid}: {msg}')
                        logging.warning('ejecutar_calculo_streaming persona %s: %s', pid, e_individual)
                        evento = {
                            'progreso': int(((index + 1) / total) * 100),
                            'actual': index + 1,
                            'total': total,
                            'detalle': msg,
                            'person': pid,
                        }

                yield f'data: {json.dumps(evento)}\n\n'

            validaciones = []
            if exitos > 0:
                try:
                    validaciones = _validar_calculo_planilla_mensajes(
                        cursor, cia, payroll_type, processtype, period
                    )
                except Exception:
                    logging.exception('validar_calculo_planilla tras ejecutar_calculo_streaming')

            yield (
                'data: '
                + json.dumps(
                    {
                        'done': True,
                        'exitos': exitos,
                        'errores': len(errores),
                        'detalles': errores,
                        'validaciones': validaciones,
                    }
                )
                + '\n\n'
            )
        except Exception as e:
            logging.exception('ejecutar_calculo_streaming')
            yield f'data: {json.dumps({"error": str(e)})}\n\n'
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

    return Response(
        generar_progreso(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'X-Accel-Buffering': 'no',
        },
    )


# --- Rutas legacy (intranet): recuperar desde control de versiones al implementar Planillas ---
#
# @app.route('/datos-personales') → datos_personales
# @app.route('/resumen-ausencias') → resumen_ausencias
# @app.route('/solicitud-permisos') → solicitud_permisos
# @app.route('/documentos-personales') → documentos_personales
# @app.route('/descargar-archivo/<filename>') → descargar_archivo
# @app.route('/solicitudes-pendientes') → solicitudes_pendientes
# @app.route('/api/eventos') → api_eventos
# Helpers: fetch_pdf_file, get_sftp_client; imports: requests, paramiko, pdfkit, pyodbc, openpyxl, sendgrid, etc.

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)
