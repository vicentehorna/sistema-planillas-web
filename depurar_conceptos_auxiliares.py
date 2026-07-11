"""Depura conceptos AUXILIARES no utilizados y genera Excel.

Usa los SP versionados en sql/:
  - sp_pr_depurar_conceptos_auxiliares_web
  - sp_pr_extraer_nemonicos_literal_sp_web

Ejemplo:
  python depurar_conceptos_auxiliares.py --company BGT --database hm_aci2
  python depurar_conceptos_auxiliares.py --company SB01
"""
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from database import DatabaseConfig

OUT_DIR = Path(__file__).resolve().parent

CALC_SPS = [
    ("FIN_DE_MES", "MENSUAL", "sp_pr_calcular_finmes_persona"),
    ("LIQUIDACION", "LIQUIDACION", "sp_pr_calcular_liquidacion_persona"),
    ("GRATIFICACION", "GRATIFICACION", "sp_pr_calcular_gratificacion_persona"),
    ("PROVISION_CTS", "PROVISION CTS", "sp_pr_calcular_provcts_persona"),
    ("PROVISION_GRATIF", "PROVISION GRATIFICACION", "sp_pr_calcular_provgrati_persona"),
    ("PROVISION_VACACIONES", "PROVISION VACACIONES", "sp_pr_calcular_provvac_persona"),
]


def _style_header(ws, headers):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fetch_sp(cur, company, modo):
    cur.execute(
        "EXEC dbo.sp_pr_depurar_conceptos_auxiliares_web @company=?, @modo=?",
        (company, modo),
    )
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def _sheet_from_rows(wb, title, headers, rows, keys):
    ws = wb.create_sheet(title)
    _style_header(ws, headers)
    for row in rows:
        ws.append([row.get(k, "") for k in keys])
    ws.freeze_panes = "A2"
    return ws


def generate_excel(company, database):
    conn = DatabaseConfig.get_connection(database=database)
    cur = conn.cursor()

    resumen = _fetch_sp(cur, company, "RESUMEN")
    no_usados = _fetch_sp(cur, company, "NO_USADOS")
    no_usados_g123 = _fetch_sp(cur, company, "NO_USADOS_G123")
    detalle = _fetch_sp(cur, company, "DETALLE")
    g1 = _fetch_sp(cur, company, "G1")
    g2 = _fetch_sp(cur, company, "G2")
    g3 = _fetch_sp(cur, company, "G3")
    g4 = _fetch_sp(cur, company, "G4")

    sp_rows = []
    maestro_fc = {
        (r.get("formulacode") or "").upper()
        for r in detalle
        if (r.get("formulacode") or "").strip()
    }
    for proceso, proceso_nombre, sp_name in CALC_SPS:
        cur.execute("EXEC dbo.sp_pr_extraer_nemonicos_literal_sp_web @procedure_name=?", (sp_name,))
        codes = cur.fetchall()
        aux_n = sum(1 for c, _ in codes if c.upper() in maestro_fc)
        sp_rows.append(
            {
                "proceso": proceso,
                "proceso_nombre": proceso_nombre,
                "sp_name": sp_name,
                "refs": len(codes),
                "aux_maestro": aux_n,
            }
        )

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    _style_header(
        ws,
        [
            "Compañía",
            "Maestro AUX",
            "G1",
            "G2",
            "G3",
            "G4",
            "Usados G1-G3",
            "Usados G1-G4",
            "No usados G1-G3",
            "No usados definitivos",
        ],
    )
    if resumen:
        r = resumen[0]
        ws.append(
            [
                r.get("company", company),
                r.get("maestro_aux"),
                r.get("g1_formulas"),
                r.get("g2_asignacion"),
                r.get("g3_calculados"),
                r.get("g4_sp_calculo"),
                r.get("usados_g123"),
                r.get("usados_g1234"),
                r.get("no_usados_g123"),
                r.get("no_usados_definitivos"),
            ]
        )

    list_keys = ["formulacode", "nombre_concepto", "concept_id", "status"]
    list_headers = ["FormulaCode", "Nombre concepto", "Concept ID", "Status"]
    _sheet_from_rows(wb, "No utilizados definitivos", list_headers, no_usados, list_keys)
    _sheet_from_rows(wb, "No usados solo G1-G3", list_headers, no_usados_g123, list_keys)
    _sheet_from_rows(wb, "Grupo 1 formulas", list_headers, g1, list_keys)
    _sheet_from_rows(wb, "Grupo 2 asignacion", list_headers, g2, list_keys)
    _sheet_from_rows(wb, "Grupo 3 calculados", list_headers, g3, list_keys)
    _sheet_from_rows(wb, "Grupo 4 SP calculo", list_headers, g4, list_keys)

    ws_det = wb.create_sheet("Detalle maestro")
    _style_header(
        ws_det,
        [
            "FormulaCode",
            "Nombre concepto",
            "Concept ID",
            "Status",
            "G1",
            "G2",
            "G3",
            "G4",
            "Depurable",
        ],
    )
    for row in detalle:
        ws_det.append(
            [
                row.get("formulacode"),
                row.get("nombre_concepto"),
                row.get("concept_id"),
                row.get("status"),
                row.get("en_g1"),
                row.get("en_g2"),
                row.get("en_g3"),
                row.get("en_g4"),
                row.get("depurable"),
            ]
        )
    ws_det.freeze_panes = "A2"

    ws_sp = wb.create_sheet("SP calculo por proceso")
    _style_header(ws_sp, ["Proceso", "Nombre proceso", "SP", "Referencias", "AUX en maestro"])
    for row in sp_rows:
        ws_sp.append(
            [row["proceso"], row["proceso_nombre"], row["sp_name"], row["refs"], row["aux_maestro"]]
        )

    out = OUT_DIR / f"conceptos_auxiliares_no_utilizados_{company.lower()}_{database}.xlsx"
    wb.save(out)
    return out, resumen, no_usados


def main():
    parser = argparse.ArgumentParser(description="Depuración conceptos AUXILIARES")
    parser.add_argument("--company", default="BGT", help="Código compañía (BGT, SB01, ...)")
    parser.add_argument("--database", default="hm_aci2", help="Base de datos destino")
    args = parser.parse_args()

    out, resumen, no_usados = generate_excel(args.company.upper(), args.database)
    print(f"Compañía: {args.company.upper()} | Base: {args.database}")
    if resumen:
        r = resumen[0]
        print(
            f"Maestro={r.get('maestro_aux')} G1={r.get('g1_formulas')} G2={r.get('g2_asignacion')} "
            f"G3={r.get('g3_calculados')} G4={r.get('g4_sp_calculo')} "
            f"Definitivos={r.get('no_usados_definitivos')}"
        )
    print(f"No utilizados: {len(no_usados)}")
    print(f"Excel: {out}")


if __name__ == "__main__":
    main()
